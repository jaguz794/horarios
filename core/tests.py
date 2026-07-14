from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.test import Client
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.admin import UserSiteAccessAdmin
from core.access import get_accessible_sites_queryset, user_can_manage_all_sites
from core.models import Holiday, Site, UserSiteAccess
from schedules.models import EmployeeInitialBalance, ScheduleLine, WeeklySchedule
from schedules.calendar_utils import get_special_day_label, is_colombian_holiday
from schedules.services import rebuild_balances_for_employees_from_week

User = get_user_model()


class UserSiteAccessTests(TestCase):
    def setUp(self):
        self.site_a = Site.objects.create(code="007", name="JARDIN.I")
        self.site_b = Site.objects.create(code="008", name="SALADO")

    def test_site_user_only_sees_assigned_sites(self):
        user = User.objects.create_user(username="operador", password="secret")
        access = UserSiteAccess.objects.get(user=user)
        access.sites.add(self.site_a)

        visible_codes = list(
            get_accessible_sites_queryset(user).order_by("code").values_list("code", flat=True)
        )

        self.assertEqual(visible_codes, ["007"])
        self.assertFalse(user_can_manage_all_sites(user))

    def test_admin_role_can_manage_all_sites(self):
        user = User.objects.create_user(username="admin_local", password="secret")
        access = UserSiteAccess.objects.get(user=user)
        access.role = UserSiteAccess.Role.ADMIN
        access.save()

        visible_codes = list(
            get_accessible_sites_queryset(user).order_by("code").values_list("code", flat=True)
        )

        self.assertIn("007", visible_codes)
        self.assertIn("008", visible_codes)
        self.assertIn(Site.PERSONAL_VARIO_CODE, visible_codes)
        self.assertTrue(user_can_manage_all_sites(user))

    def test_user_without_assigned_sites_sees_none(self):
        user = User.objects.create_user(username="sin_sede", password="secret")

        visible_codes = list(
            get_accessible_sites_queryset(user).order_by("code").values_list("code", flat=True)
        )

        self.assertEqual(visible_codes, [])

    def test_site_user_never_sees_admin_only_site(self):
        personal_vario, _ = Site.objects.get_or_create(
            code=Site.PERSONAL_VARIO_CODE,
            defaults={
                "name": Site.PERSONAL_VARIO_NAME,
                "admin_only": True,
                "is_active": True,
            },
        )
        user = User.objects.create_user(username="operador_vario", password="secret")
        access = UserSiteAccess.objects.get(user=user)
        access.sites.add(personal_vario)

        visible_codes = list(
            get_accessible_sites_queryset(user).order_by("code").values_list("code", flat=True)
        )

        self.assertNotIn(Site.PERSONAL_VARIO_CODE, visible_codes)


class DashboardViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.site_a = Site.objects.create(code="007", name="JARDIN.I")
        self.site_b = Site.objects.create(code="008", name="SALADO")
        self.user = User.objects.create_user(username="admin_dashboard", password="secret")
        access = UserSiteAccess.objects.get(user=self.user)
        access.role = UserSiteAccess.Role.ADMIN
        access.save()
        self.schedule_a = WeeklySchedule.objects.create(
            site=self.site_a,
            week_start_date=date(2026, 6, 7),
            status=WeeklySchedule.Status.DRAFT,
        )
        self.schedule_b = WeeklySchedule.objects.create(
            site=self.site_b,
            week_start_date=date(2026, 6, 14),
            status=WeeklySchedule.Status.PUBLISHED,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule_a,
            employee_identifier="100",
            employee_name="Ana",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("8.00"),
            expected_weekly_hours=Decimal("12.00"),
            total_hours=Decimal("10.00"),
            weekly_hour_difference=Decimal("-2.00"),
            warnings_count=2,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule_a,
            employee_identifier="101",
            employee_name="Bruno",
            job_role_name="CAJERO",
            warnings_count=0,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule_b,
            employee_identifier="102",
            employee_name="Carla",
            job_role_name="SUPERVISOR",
            warnings_count=1,
        )

    def test_dashboard_shows_alert_tooltip_and_schedule_summaries(self):
        self.client.login(username="admin_dashboard", password="secret")

        response = self.client.get(
            reverse("dashboard"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen por horario")
        self.assertNotContains(response, "3 horario(s) con novedades")
        self.assertContains(response, "2 horario(s) con novedades")
        self.assertContains(response, "2 persona(s) cargadas.")
        self.assertContains(response, "2 alerta(s) en 1 colaborador(es).")
        self.assertContains(response, "Revisa objetivo semanal.")
        self.assertContains(response, "Ayuda")
        self.assertContains(response, "10vJTA2WoJ0HkF5ByLp35Gw1LKR2qs8zR/preview")

    def test_dashboard_uses_latest_balance_after_same_week_transfer(self):
        transfer_schedule_late = WeeklySchedule.objects.create(
            site=Site.objects.create(code="009", name="JARDIN.N"),
            week_start_date=date(2026, 6, 21),
            status=WeeklySchedule.Status.DRAFT,
        )
        transfer_schedule_early = WeeklySchedule.objects.create(
            site=Site.objects.create(code="010", name="RIOJA"),
            week_start_date=date(2026, 6, 21),
            status=WeeklySchedule.Status.DRAFT,
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="2005",
            employee_name="Empleado Dashboard",
            initial_day_balance=Decimal("6.00"),
        )
        ScheduleLine.objects.create(
            schedule=transfer_schedule_late,
            employee_identifier="2005",
            employee_name="Empleado Dashboard",
            job_role_name="AUXILIAR",
            day_4_shift_1="08:00-16:00",
            day_5_shift_1="08:00-16:00",
            day_6_shift_1="08:00-16:00",
        )
        ScheduleLine.objects.create(
            schedule=transfer_schedule_early,
            employee_identifier="2005",
            employee_name="Empleado Dashboard",
            job_role_name="AUXILIAR",
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_3_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )

        rebuild_balances_for_employees_from_week(date(2026, 6, 21), ["2005"])
        self.client.login(username="admin_dashboard", password="secret")
        response = self.client.get(
            reverse("dashboard"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["extra_day_total"], Decimal("3.00"))

    def test_dashboard_ignores_blank_draft_line_when_previous_balance_is_real(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="011", name="PRIOR"),
            week_start_date=date(2026, 6, 28),
            status=WeeklySchedule.Status.PUBLISHED,
        )
        current_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="012", name="ACTUAL"),
            week_start_date=date(2026, 7, 5),
            status=WeeklySchedule.Status.DRAFT,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="2010",
            employee_name="Empleado Con Saldo",
            job_role_name="AUXILIAR",
            accrued_day_balance=Decimal("2.00"),
            accrued_hour_balance=Decimal("6.00"),
            day_1_shift_1="08:00-16:00",
        )
        ScheduleLine.objects.create(
            schedule=current_schedule,
            employee_identifier="2010",
            employee_name="Empleado Con Saldo",
            job_role_name="AUXILIAR",
            accrued_day_balance=Decimal("-3.00"),
            accrued_hour_balance=Decimal("-44.00"),
        )

        self.client.login(username="admin_dashboard", password="secret")
        response = self.client.get(
            reverse("dashboard"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.context["extra_day_total"], Decimal("2.00"))
        self.assertGreaterEqual(response.context["extra_hour_total"], Decimal("6.00"))


class HolidayCalendarTests(TestCase):
    def test_seeded_manual_holiday_is_available_in_schedule_logic(self):
        configured_holiday = date(2026, 7, 13)

        self.assertTrue(Holiday.objects.filter(holiday_date=configured_holiday, is_active=True).exists())
        self.assertTrue(is_colombian_holiday(configured_holiday))
        self.assertEqual(get_special_day_label(configured_holiday), "Festivo")


class SiteOrderingTests(TestCase):
    def test_site_list_orders_codes_ascending(self):
        user = User.objects.create_user(username="admin_sites", password="secret")
        access = UserSiteAccess.objects.get(user=user)
        access.role = UserSiteAccess.Role.ADMIN
        access.save()
        Site.objects.create(code="010", name="Diez")
        Site.objects.create(code="002", name="Dos")
        Site.objects.create(code="001", name="Uno")

        self.client.login(username="admin_sites", password="secret")
        response = self.client.get(
            reverse("site-list"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        visible_codes = list(response.context["sites"].values_list("code", flat=True))
        self.assertEqual(visible_codes[:3], ["001", "002", "010"])
        self.assertIn(Site.PERSONAL_VARIO_CODE, visible_codes)

    def test_admin_sites_field_orders_codes_ascending(self):
        Site.objects.create(code="010", name="Diez")
        Site.objects.create(code="002", name="Dos")
        Site.objects.create(code="001", name="Uno")

        admin_instance = UserSiteAccessAdmin(UserSiteAccess, AdminSite())
        form_field = admin_instance.formfield_for_manytomany(
            UserSiteAccess._meta.get_field("sites"),
            request=None,
        )

        visible_codes = list(form_field.queryset.values_list("code", flat=True))
        self.assertEqual(visible_codes[:3], ["001", "002", "010"])
        self.assertIn(Site.PERSONAL_VARIO_CODE, visible_codes)


class ReportHubViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.site = Site.objects.create(code="007", name="JARDIN.I")
        self.user = User.objects.create_user(username="admin_reports", password="secret")
        access = UserSiteAccess.objects.get(user=self.user)
        access.role = UserSiteAccess.Role.ADMIN
        access.save()
        self.schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 5),
            status=WeeklySchedule.Status.DRAFT,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="5001",
            employee_name="Ana Inventario",
            job_role_name="AUXILIAR",
            day_1_inventory=True,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="5002",
            employee_name="Bruno Inventario",
            job_role_name="CAJERO",
            day_1_inventory=True,
            day_4_inventory=True,
        )

    def test_reports_can_export_inventory_excel(self):
        self.client.login(username="admin_reports", password="secret")

        response = self.client.post(
            reverse("reports"),
            {
                "week-site": str(self.site.pk),
                "week-week_start_date": "2026-07-05",
                "report_type": "inventory_excel",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A1"].value, "Fecha inventario")
        self.assertEqual(worksheet["C2"].value, "JARDIN.I")
        self.assertEqual(worksheet["D2"].value, "5001")
        self.assertEqual(worksheet["F2"].value, "AUXILIAR")

    def test_reports_can_export_weekly_balance_with_day_and_hour_columns(self):
        line = self.schedule.lines.order_by("employee_identifier").first()
        line.job_role_name = "AUXILIAR"
        line.accrued_day_balance = Decimal("2.00")
        line.accrued_hour_balance = Decimal("5.50")
        line.night_bonus_hours = Decimal("3.00")
        line.save(
            update_fields=[
                "job_role_name",
                "accrued_day_balance",
                "accrued_hour_balance",
                "night_bonus_hours",
                "updated_at",
            ]
        )

        self.client.login(username="admin_reports", password="secret")

        response = self.client.post(
            reverse("reports"),
            {
                "week-site": str(self.site.pk),
                "week-week_start_date": "2026-07-05",
                "report_type": "weekly_balance",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A1"].value, "Sede")
        self.assertEqual(worksheet["E1"].value, "Dias acumulados")
        self.assertEqual(worksheet["F1"].value, "Horas acumuladas")
        self.assertEqual(worksheet["G1"].value, "Recargos nocturnos")
        self.assertEqual(worksheet["A2"].value, "JARDIN.I")
        self.assertEqual(worksheet["E2"].value, 2)
        self.assertEqual(worksheet["F2"].value, 5.5)
        self.assertEqual(worksheet["G2"].value, 3)

    def test_reports_can_export_inventory_pdf(self):
        self.client.login(username="admin_reports", password="secret")

        response = self.client.post(
            reverse("reports"),
            {
                "week-site": str(self.site.pk),
                "week-week_start_date": "2026-07-05",
                "report_type": "inventory_pdf",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("planilla_inventario_20260705.pdf", response["Content-Disposition"])
        self.assertGreater(len(response.content), 1000)

    def test_reports_can_export_hourly_coverage_excel(self):
        lines = list(self.schedule.lines.order_by("employee_identifier"))
        lines[0].job_role_name = "AUXILIAR"
        lines[0].day_0_shift_1 = "06:00-10:00"
        lines[0].save()
        lines[1].job_role_name = "CAJERO"
        lines[1].day_0_shift_1 = "08:00-12:00"
        lines[1].save()

        self.client.login(username="admin_reports", password="secret")
        response = self.client.post(
            reverse("reports"),
            {
                "week-site": str(self.site.pk),
                "week-week_start_date": "2026-07-05",
                "report_type": "coverage_excel",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["Domingo"]
        self.assertEqual(worksheet["A2"].value, "Franja horaria")
        headers = [worksheet.cell(row=2, column=index).value for index in range(1, worksheet.max_column + 1)]
        auxiliar_column = headers.index("AUXILIAR") + 1
        cajero_column = headers.index("CAJERO") + 1
        self.assertEqual(worksheet.cell(row=3, column=auxiliar_column).value, 1)
        self.assertEqual(worksheet.cell(row=3, column=cajero_column).value, 0)
        self.assertEqual(worksheet.cell(row=5, column=auxiliar_column).value, 1)
        self.assertEqual(worksheet.cell(row=5, column=cajero_column).value, 1)


class UserAdminTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin_user = User.objects.create_superuser(
            username="admin_users",
            email="admin_users@example.com",
            password="secret",
        )

    def test_admin_user_add_page_hides_site_access_inline(self):
        self.client.login(username="admin_users", password="secret")

        response = self.client.get(
            reverse("admin:auth_user_add"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "site_access-TOTAL_FORMS", html=False)
        self.assertNotContains(response, "Acceso al portal")

    def test_admin_can_create_user_without_inline_management_form(self):
        self.client.login(username="admin_users", password="secret")

        response = self.client.post(
            reverse("admin:auth_user_add"),
            {
                "username": "nuevo_usuario",
                "password1": "ClaveSegura123!",
                "password2": "ClaveSegura123!",
                "_save": "Guardar",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(username="nuevo_usuario").exists())
        self.assertTrue(UserSiteAccess.objects.filter(user__username="nuevo_usuario").exists())
