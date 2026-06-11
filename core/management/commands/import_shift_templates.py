from __future__ import annotations

from datetime import datetime, time
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core.models import ShiftTemplate

NON_WORKED_LABELS = {
    "descanso",
    "incapacidad",
    "licencia",
    "prestamo",
    "traslado",
    "vacaciones",
    "volantes",
    "renuncia",
}


def time_to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    if isinstance(value, time):
        return (Decimal(value.hour) + (Decimal(value.minute) / Decimal("60"))).quantize(Decimal("0.01"))
    return Decimal("0.00")


def normalize_time_value(value):
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or ":" not in cleaned:
            return None
        return datetime.strptime(cleaned, "%H:%M").time()
    return None


def infer_range_from_label(label: str) -> tuple[time | None, time | None]:
    if "-" not in label:
        return None, None

    start_label, end_label = label.split("-", 1)
    start_time = datetime.strptime(start_label.strip(), "%H:%M").time()
    end_time = datetime.strptime(end_label.strip(), "%H:%M").time()
    return start_time, end_time


class Command(BaseCommand):
    help = "Importa o actualiza los turnos desde un archivo Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=r"C:\Users\POPULAR\Desktop\turnos.xlsx",
            help="Ruta absoluta al archivo Excel de turnos.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        workbook = load_workbook(file_path, data_only=True)
        if "TURNOS" not in workbook.sheetnames:
            raise CommandError("El archivo no contiene la hoja TURNOS.")

        worksheet = workbook["TURNOS"]
        created_count = 0
        updated_count = 0

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            code, start_time, end_time, label, total_hours, night_bonus = row[:6]
            normalized_label = (label or "").strip()
            if not normalized_label:
                continue

            start_time = normalize_time_value(start_time)
            end_time = normalize_time_value(end_time)

            if start_time is None and "-" in normalized_label:
                inferred_start, inferred_end = infer_range_from_label(normalized_label)
                start_time = start_time or inferred_start
                end_time = end_time or inferred_end

            duration_hours = time_to_decimal(total_hours)
            night_bonus_hours = time_to_decimal(night_bonus)
            counts_as_worked_time = (
                duration_hours > 0 and normalized_label.casefold() not in NON_WORKED_LABELS
            )

            shift, created = ShiftTemplate.objects.update_or_create(
                label=normalized_label,
                defaults={
                    "code": (code or "").strip() if code else "",
                    "start_time": start_time,
                    "end_time": end_time,
                    "duration_hours": duration_hours,
                    "night_bonus_hours": night_bonus_hours,
                    "display_order": row_number,
                    "counts_as_worked_time": counts_as_worked_time,
                    "is_active": True,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Turnos importados correctamente. Nuevos: {created_count}. Actualizados: {updated_count}."
            )
        )
