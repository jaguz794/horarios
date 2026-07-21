from io import BytesIO, StringIO
from datetime import date, time, timedelta
from decimal import Decimal
from unittest.mock import patch

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import ProgrammingError
from django.test import Client
from django.test import SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.models import Holiday, JobRole, OperationalStaffCache, ShiftTemplate, Site, SystemConfiguration, UserSiteAccess
from legacy.services import LegacyStaffLookupError, LegacyThirdPartyRecord, OperationalStaffingRecord
from schedules.forms import ScheduleLineForm, ScheduleLoadForm, build_shift_choices
from schedules.models import (
    EmployeeInitialBalance,
    EmployeeOvertimeRestriction,
    EmployeeScheduleBlacklist,
    ScheduleBalanceMovement,
    ScheduleLine,
    WeeklySchedule,
)
from schedules.services import (
    COMPANY_DAY_REPAYMENT_MODE,
    build_expected_week_plan,
    build_schedule_flat_file_headers,
    build_schedule_balance_audit_rows,
    copy_schedule_template,
    get_rest_shift_label,
    get_schedule_line_compact_alert_summary,
    get_schedule_line_balance_snapshot,
    import_employee_initial_balances,
    parse_shift_hours,
    recalculate_schedule_line,
    rebuild_balances_for_employees_from_week,
    round_minutes_to_interval,
    schedule_line_blocks_status_transition,
    sync_schedule_from_legacy,
)
from schedules.settlement_pdf import get_settlement_rows
from schedules.templatetags.schedule_tags import hours_int, non_negative_hours_int

User = get_user_model()


class ScheduleTemplateFilterTests(SimpleTestCase):
    def test_hours_int_preserves_half_hours(self):
        self.assertEqual(hours_int(Decimal("1.50")), "1.5")
        self.assertEqual(hours_int(Decimal("8.00")), "8")

    def test_non_negative_hours_int_hides_legacy_negative_values(self):
        self.assertEqual(non_negative_hours_int(Decimal("-3.50")), "0")
        self.assertEqual(non_negative_hours_int(Decimal("2.00")), "2")

    def test_data_upload_max_number_fields_is_large_enough(self):
        self.assertGreaterEqual(settings.DATA_UPLOAD_MAX_NUMBER_FIELDS, 12000)


class ScheduleRoundingRuleTests(SimpleTestCase):
    @staticmethod
    def minutes_from_clock(label: str) -> int:
        hours, minutes = label.split(":")
        return (int(hours) * 60) + int(minutes)

    def test_round_minutes_to_programming_interval_uses_half_hour_blocks(self):
        scenarios = [
            ("0:00", "0:00"),
            ("43:00", "43:00"),
            ("43:01", "43:00"),
            ("43:14", "43:00"),
            ("43:15", "43:30"),
            ("43:20", "43:30"),
            ("43:29", "43:30"),
            ("43:30", "43:30"),
            ("43:31", "43:30"),
            ("43:44", "43:30"),
            ("43:45", "44:00"),
            ("43:59", "44:00"),
        ]

        for raw_value, expected_value in scenarios:
            with self.subTest(raw=raw_value):
                rounded = round_minutes_to_interval(
                    self.minutes_from_clock(raw_value),
                    30,
                )
                self.assertEqual(int(rounded), self.minutes_from_clock(expected_value))


class ScheduleCalculationTests(TestCase):
    def setUp(self):
        config = SystemConfiguration.load()
        config.default_weekly_hours = Decimal("46.00")
        config.default_daily_max_hours = Decimal("9.00")
        config.save()

        self.site = Site.objects.create(code="007", name="JARDIN.I", legacy_name="JARDIN IBAGUE")
        self.schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 5),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ShiftTemplate.objects.create(
            code="T1",
            label="06:00-10:00",
            start_time=time(6, 0),
            end_time=time(10, 0),
            duration_hours=Decimal("4.00"),
            display_order=1,
        )
        ShiftTemplate.objects.create(
            code="T2",
            label="13:00-17:00",
            start_time=time(13, 0),
            end_time=time(17, 0),
            duration_hours=Decimal("4.00"),
            display_order=2,
        )
        ShiftTemplate.objects.create(
            code="T3",
            label="08:00-16:00",
            start_time=time(8, 0),
            end_time=time(16, 0),
            duration_hours=Decimal("8.00"),
            display_order=3,
        )
        ShiftTemplate.objects.create(
            code="TN",
            label="18:00-00:00",
            start_time=time(18, 0),
            end_time=time(0, 0),
            duration_hours=Decimal("6.00"),
            display_order=4,
        )

    def build_form_data(self, **overrides):
        data = {}
        for index in range(7):
            data[f"day_{index}_shift_1"] = ""
            data[f"day_{index}_shift_2"] = ""
            data[f"day_{index}_compensation_mode"] = ""
            data[f"day_{index}_compensation_hours"] = ""

        data.update(overrides)
        return data

    def test_parse_shift_hours_works_for_plain_range(self):
        self.assertEqual(parse_shift_hours("06:00-10:00"), Decimal("4.00"))

    def test_recalculate_line_sets_totals_and_variance(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="123",
            employee_name="Empleado Demo",
            weekly_target_hours=Decimal("8.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="06:00-10:00",
            day_0_shift_2="13:00-17:00",
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.day_0_hours, Decimal("8.00"))
        self.assertEqual(line.total_hours, Decimal("8.00"))
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.accrued_total_hours_balance, Decimal("1.33"))

    def test_recalculate_line_accumulates_prior_balance_and_current_payments(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="128",
            employee_name="Empleado Saldo",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("1.00"),
            accrued_hour_balance=Decimal("5.50"),
            accrued_total_hours_balance=Decimal("13.50"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="128",
            employee_name="Empleado Saldo",
            weekly_target_hours=Decimal("4.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.expected_weekly_hours, Decimal("4.00"))
        self.assertEqual(line.overtime_hours, Decimal("4.00"))
        self.assertEqual(line.payment_days_used, 1)
        self.assertEqual(line.payment_hours_used, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("9.50"))
        self.assertEqual(line.accrued_total_hours_balance, Decimal("10.17"))

    def test_recalculate_line_tracks_money_day_and_money_hour_payments_separately(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="128B",
            employee_name="Empleado Pago Dinero",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("3.00"),
            accrued_hour_balance=Decimal("6.00"),
            accrued_total_hours_balance=Decimal("30.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="128B",
            employee_name="Empleado Pago Dinero",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_compensation_mode=ScheduleLine.CompensationMode.PAY_MONEY_DAY,
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
            day_1_compensation_hours=Decimal("2.00"),
            day_1_shift_1="08:00-16:00",
            day_2_shift_1="08:00-16:00",
            day_3_shift_1="08:00-16:00",
            day_4_shift_1="08:00-16:00",
            day_5_shift_1="08:00-16:00",
            day_6_shift_1="18:00-00:00",
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.money_payment_days_used, 1)
        self.assertEqual(line.money_payment_hours_used, Decimal("2.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("2.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("4.00"))
        self.assertEqual(line.accrued_total_hours_balance, Decimal("19.34"))

    def test_form_hides_company_day_repayment_mode_for_new_entries(self):
        line = ScheduleLine(
            schedule=self.schedule,
            employee_identifier="128C",
            employee_name="Empleado Compensacion",
            weekly_target_hours=Decimal("48.00"),
            daily_max_hours=Decimal("8.00"),
        )

        form = ScheduleLineForm(instance=line, schedule=self.schedule)

        self.assertNotIn(
            COMPANY_DAY_REPAYMENT_MODE,
            dict(form.fields["day_0_compensation_mode"].choices),
        )

    def test_form_preserves_company_day_repayment_mode_if_already_selected(self):
        line = ScheduleLine(
            schedule=self.schedule,
            employee_identifier="128C-1",
            employee_name="Empleado Compensacion",
            weekly_target_hours=Decimal("48.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_compensation_mode=COMPANY_DAY_REPAYMENT_MODE,
        )

        form = ScheduleLineForm(instance=line, schedule=self.schedule)

        self.assertIn(
            COMPANY_DAY_REPAYMENT_MODE,
            dict(form.fields["day_0_compensation_mode"].choices),
        )

    def test_form_accepts_company_day_repayment_with_full_seventh_day(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="128D",
            employee_name="Empleado Compensacion",
            initial_day_balance=Decimal("-1.00"),
        )
        line = ScheduleLine(
            schedule=self.schedule,
            employee_identifier="128D",
            employee_name="Empleado Compensacion",
            weekly_target_hours=Decimal("48.00"),
            daily_max_hours=Decimal("8.00"),
        )
        data = self.build_form_data(day_0_compensation_mode=COMPANY_DAY_REPAYMENT_MODE)
        for index in range(7):
            data[f"day_{index}_shift_1"] = "08:00-16:00"

        form = ScheduleLineForm(data=data, instance=line, schedule=self.schedule)

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_rejects_company_day_repayment_without_additional_full_day(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="128E",
            employee_name="Empleado Compensacion",
            initial_day_balance=Decimal("-1.00"),
        )
        line = ScheduleLine(
            schedule=self.schedule,
            employee_identifier="128E",
            employee_name="Empleado Compensacion",
            weekly_target_hours=Decimal("48.00"),
            daily_max_hours=Decimal("8.00"),
        )
        data = self.build_form_data(day_0_compensation_mode=COMPANY_DAY_REPAYMENT_MODE)
        data["day_0_shift_1"] = "08:00-16:00"
        data["day_1_shift_1"] = "descanso"
        for index in range(2, 7):
            data[f"day_{index}_shift_1"] = "08:00-16:00"

        form = ScheduleLineForm(data=data, instance=line, schedule=self.schedule)

        self.assertFalse(form.is_valid())
        self.assertIn("dia completo adicional", form.errors["day_0_compensation_mode"][0].lower())

    def test_recalculate_line_tracks_night_bonus_hours(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="126",
            employee_name="Empleado Noche",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="18:00-00:00",
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.day_0_hours, Decimal("6.00"))
        self.assertEqual(line.night_bonus_hours, Decimal("5.00"))

    def test_recalculate_line_uses_initial_balance_when_no_prior_week_exists(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="140",
            employee_name="Empleado Inicial",
            initial_day_balance=Decimal("2.00"),
            initial_hour_balance=Decimal("3.50"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="140",
            employee_name="Empleado Inicial",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_1_shift_1="08:00-16:00",
            day_2_shift_1="08:00-16:00",
            day_3_shift_1="08:00-16:00",
            day_4_shift_1="08:00-16:00",
            day_5_shift_1="08:00-16:00",
            day_6_shift_1="18:00-00:00",
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.accrued_day_balance, Decimal("2.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("3.50"))
        self.assertEqual(line.accrued_total_hours_balance, Decimal("18.84"))

    def test_form_accepts_pay_day_using_same_week_special_day_generated_before(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141",
            employee_name="Empleado Semana",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-16:00",
                day_3_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_allows_pay_day_without_positive_day_balance_until_minus_two(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141B",
            employee_name="Empleado Pago Dia",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
                day_2_shift_1="09:00-16:00",
                day_3_shift_1="09:00-16:00",
                day_4_shift_1="09:00-16:00",
                day_5_shift_1="09:00-16:00",
                day_6_shift_1="09:00-16:00",
            ),
            instance=line,
            schedule=self.schedule,
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_rejects_plain_rest_day_when_company_day_limit_is_already_reached(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="141C",
            employee_name="Empleado Limite",
            initial_day_balance=Decimal("-2.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141C",
            employee_name="Empleado Limite",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="descanso",
                day_2_shift_1="09:00-16:00",
                day_3_shift_1="09:00-16:00",
                day_4_shift_1="09:00-16:00",
                day_5_shift_1="09:00-16:00",
                day_6_shift_1="09:00-16:00",
            ),
            instance=line,
            schedule=self.schedule,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("limite maximo", form.errors["day_1_shift_1"][0].lower())

    def test_form_rejects_pay_hours_without_available_balance_from_prior_weeks(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141A",
            employee_name="Empleado Horas Semana",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-16:00",
                day_2_shift_1="18:00-00:00",
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_HOURS,
                day_2_compensation_hours="1",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("saldo acumulado disponible", form.errors["day_2_compensation_hours"][0].lower())

    def test_form_rejects_weekly_overtime_over_medical_restriction(self):
        EmployeeOvertimeRestriction.objects.create(
            employee_identifier="141B",
            employee_name="Empleado Restringido",
            max_daily_overtime_hours=Decimal("8.00"),
            max_weekly_overtime_hours=Decimal("0.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141B",
            employee_name="Empleado Restringido",
            weekly_target_hours=Decimal("8.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-16:00",
                day_1_shift_1="08:00-16:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertTrue(
            any("restriccion medica" in error.lower() for error in form.non_field_errors()),
            form.errors,
        )

    def test_form_rejects_daily_overtime_over_medical_restriction(self):
        EmployeeOvertimeRestriction.objects.create(
            employee_identifier="141C",
            employee_name="Empleado Restringido Diario",
            max_daily_overtime_hours=Decimal("0.00"),
            max_weekly_overtime_hours=Decimal("20.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="141C",
            employee_name="Empleado Restringido Diario",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-16:00",
                day_0_shift_2="13:00-17:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertTrue(
            "restriccion medica" in form.errors["day_0_shift_2"][0].lower(),
            form.errors,
        )

    def test_copy_schedule_template_brings_previous_week_turns(self):
        source_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=source_schedule,
            employee_identifier="142",
            employee_name="Empleado Base",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="06:00-10:00",
            day_0_shift_2="13:00-17:00",
            day_1_shift_1="08:00-16:00",
        )
        target_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 14),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        target_line = ScheduleLine.objects.create(
            schedule=target_schedule,
            employee_identifier="142",
            employee_name="Empleado Base",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
        )

        copied_count, _ = copy_schedule_template(source_schedule, target_schedule)

        target_line.refresh_from_db()
        self.assertEqual(copied_count, 1)
        self.assertEqual(target_line.day_0_shift_1, "06:00-10:00")
        self.assertEqual(target_line.day_0_shift_2, "13:00-17:00")
        self.assertEqual(target_line.day_1_shift_1, "08:00-16:00")

    def test_schedule_load_form_rejects_non_sunday_start_date(self):
        admin_user = User.objects.create_user(username="admin_semana", password="secret")
        access = UserSiteAccess.objects.get(user=admin_user)
        access.role = UserSiteAccess.Role.ADMIN
        access.save()
        form = ScheduleLoadForm(
            data={
                "site": str(self.site.pk),
                "week_start_date": "2026-06-08",
                "copy_from_schedule": "",
            },
            user=admin_user,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("domingos", form.errors["week_start_date"][0].lower())

    def test_compact_alert_summary_groups_main_categories(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="133",
            employee_name="Empleado Resumen",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )

        recalculate_schedule_line(line)

        compact_summary = get_schedule_line_compact_alert_summary(line)
        self.assertIn("Resultado estimado:", compact_summary)
        self.assertIn("Bloquea revision/publicacion", compact_summary)
        self.assertIn("42 h", compact_summary)

    def test_compact_alert_summary_hides_allowed_one_hour_difference(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="134",
            employee_name="Empleado Margen",
            weekly_target_hours=Decimal("9.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
        )

        recalculate_schedule_line(line)

        compact_summary = get_schedule_line_compact_alert_summary(line)
        self.assertIn("Resultado estimado:", compact_summary)
        self.assertNotIn("Revisa jornada semanal", compact_summary)
        self.assertNotIn("Bloquea revision/publicacion", compact_summary)
        self.assertFalse(schedule_line_blocks_status_transition(line))

    def test_allowed_one_hour_excess_is_not_counted_as_warning(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="135",
            employee_name="Empleado Exceso Permitido",
            weekly_target_hours=Decimal("7.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
        )

        recalculate_schedule_line(line)

        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.OVERPLANNED)
        self.assertEqual(line.weekly_hour_difference, Decimal("1.00"))
        self.assertEqual(line.warnings_count, 0)
        self.assertIn("Estado: Valida con diferencia permitida.", line.validation_summary)

    def test_schedule_lines_are_ordered_by_role_name_then_employee(self):
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="201",
            employee_name="Zulu",
            job_role_name="ZAPATERO",
        )
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="202",
            employee_name="Ana",
            job_role_name="AUXILIAR",
        )

        ordered_roles = list(
            self.schedule.lines.values_list("job_role_name", "employee_name")
        )

        self.assertEqual(
            ordered_roles,
            [("AUXILIAR", "Ana"), ("ZAPATERO", "Zulu")],
        )

    def test_form_rejects_second_shift_when_first_is_continuous(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="124",
            employee_name="Empleado Demo 2",
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-16:00",
                day_0_shift_2="13:00-17:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("jornadas continuas", form.errors["day_0_shift_2"][0].lower())

    def test_form_allows_split_afternoon_shift_when_second_starts_after_first(self):
        ShiftTemplate.objects.create(
            code="T5",
            label="13:00-16:00",
            start_time=time(13, 0),
            end_time=time(16, 0),
            duration_hours=Decimal("3.00"),
            display_order=5,
        )
        ShiftTemplate.objects.create(
            code="T6",
            label="17:00-21:00",
            start_time=time(17, 0),
            end_time=time(21, 0),
            duration_hours=Decimal("4.00"),
            display_order=6,
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="124A",
            employee_name="Empleado Turno Tarde",
        )

        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="13:00-16:00",
                day_0_shift_2="17:00-21:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_allows_contractacion_as_non_worked_novelty(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="124C",
            employee_name="Empleado Contratacion",
        )

        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="contratacion",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=build_shift_choices(second_slot=False),
            secondary_shift_choices=build_shift_choices(second_slot=True),
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["day_1_shift_1"], "contratacion")

    def test_form_rejects_second_shift_when_it_starts_before_first_ends(self):
        ShiftTemplate.objects.create(
            code="T7",
            label="13:00-16:00",
            start_time=time(13, 0),
            end_time=time(16, 0),
            duration_hours=Decimal("3.00"),
            display_order=7,
        )
        ShiftTemplate.objects.create(
            code="T8",
            label="15:00-21:00",
            start_time=time(15, 0),
            end_time=time(21, 0),
            duration_hours=Decimal("6.00"),
            display_order=8,
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="124B",
            employee_name="Empleado Turno Cruzado",
        )

        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="13:00-16:00",
                day_0_shift_2="15:00-21:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("debe iniciar despues", form.errors["day_0_shift_2"][0].lower())

    def test_form_allows_pay_day_without_prior_day_balance_until_minus_two(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="130",
            employee_name="Empleado Demo 6",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
                day_1_shift_1="09:00-16:00",
                day_3_shift_1="09:00-16:00",
                day_4_shift_1="09:00-16:00",
                day_5_shift_1="09:00-16:00",
                day_6_shift_1="09:00-16:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_allows_pay_day_when_only_hour_balance_exists_until_minus_two(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="130A",
            employee_name="Empleado Solo Horas",
            accrued_hour_balance=Decimal("8.00"),
            accrued_total_hours_balance=Decimal("8.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="130A",
            employee_name="Empleado Solo Horas",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
                day_1_shift_1="09:00-16:00",
                day_3_shift_1="09:00-16:00",
                day_4_shift_1="09:00-16:00",
                day_5_shift_1="09:00-16:00",
                day_6_shift_1="09:00-16:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_rejects_pay_day_on_sunday(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="130AA",
            employee_name="Empleado Domingo",
            accrued_day_balance=Decimal("2.00"),
            accrued_total_hours_balance=Decimal("16.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="130AA",
            employee_name="Empleado Domingo",
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(day_0_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("domingo o festivo", form.errors["day_0_compensation_mode"][0].lower())

    def test_form_rejects_money_day_when_only_hour_balance_exists(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="130B",
            employee_name="Empleado Dinero Dia",
            accrued_hour_balance=Decimal("8.00"),
            accrued_total_hours_balance=Decimal("8.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="130B",
            employee_name="Empleado Dinero Dia",
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_MONEY_DAY),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
            allow_money_payment=True,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("pago en dinero por dia", form.errors["day_2_compensation_mode"][0].lower())

    def test_form_accepts_money_day_on_holiday(self):
        holiday_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 28),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 21),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="130BB",
            employee_name="Empleado Festivo",
            accrued_day_balance=Decimal("2.00"),
            accrued_total_hours_balance=Decimal("16.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=holiday_schedule,
            employee_identifier="130BB",
            employee_name="Empleado Festivo",
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_MONEY_DAY),
            instance=line,
            schedule=holiday_schedule,
            shift_choices=[],
            secondary_shift_choices=[],
            allow_money_payment=True,
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_accepts_money_day_with_prior_day_balance(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="130C",
            employee_name="Empleado Dinero Dia Valido",
            accrued_day_balance=Decimal("2.00"),
            accrued_total_hours_balance=Decimal("16.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="130C",
            employee_name="Empleado Dinero Dia Valido",
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_MONEY_DAY),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
            allow_money_payment=True,
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_pay_day_converts_day_to_rest_automatically(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="131",
            employee_name="Empleado Demo 7",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("1.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="131",
            employee_name="Empleado Demo 7",
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_2_shift_1="08:00-16:00",
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["day_2_shift_1"], get_rest_shift_label())
        self.assertEqual(form.cleaned_data["day_2_shift_2"], "")

    def test_form_rejects_advance_day_when_company_day_limit_is_already_reached(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="131A",
            employee_name="Empleado Tope Adelantado",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("-2.00"),
            advance_rest_pending_balance=Decimal("2.00"),
            accrued_total_hours_balance=Decimal("-16.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="131A",
            employee_name="Empleado Tope Adelantado",
        )
        form = ScheduleLineForm(
            data=self.build_form_data(day_2_compensation_mode=ScheduleLine.CompensationMode.ADVANCE_DAY),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("limite maximo", form.errors["day_2_compensation_mode"][0].lower())

    def test_form_rejects_second_advance_day_when_balance_would_drop_below_minus_two(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="131B",
            employee_name="Empleado Adelantado Parcial",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("-1.00"),
            advance_rest_pending_balance=Decimal("1.00"),
            accrued_total_hours_balance=Decimal("-8.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="131B",
            employee_name="Empleado Adelantado Parcial",
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_2_compensation_mode=ScheduleLine.CompensationMode.ADVANCE_DAY,
                day_3_compensation_mode=ScheduleLine.CompensationMode.ADVANCE_DAY,
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("limite maximo", form.errors["day_3_compensation_mode"][0].lower())

    def test_form_rejects_pay_hours_over_prior_available_balance(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="132",
            employee_name="Empleado Demo 8",
            accrued_hour_balance=Decimal("1.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="132",
            employee_name="Empleado Demo 8",
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="18:00-00:00",
                day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_HOURS,
                day_1_compensation_hours="1.5",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("saldo acumulado disponible hasta ese dia", form.errors["day_1_compensation_hours"][0].lower())

    def test_form_rejects_pay_hours_above_needed_for_daily_journey(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="134",
            employee_name="Empleado Demo 9",
            daily_max_hours=Decimal("8.00"),
            accrued_hour_balance=Decimal("5.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="134",
            employee_name="Empleado Demo 9",
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="06:00-10:00",
                day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_HOURS,
                day_1_compensation_hours="5",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("completar la jornada", form.errors["day_1_compensation_hours"][0].lower())

    def test_form_allows_pay_hours_until_daily_max_not_weekly_prorated_average(self):
        prior_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 5, 31),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=prior_schedule,
            employee_identifier="134B",
            employee_name="Empleado Demo 9B",
            weekly_target_hours=Decimal("44.00"),
            daily_max_hours=Decimal("8.00"),
            accrued_hour_balance=Decimal("2.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="134B",
            employee_name="Empleado Demo 9B",
            weekly_target_hours=Decimal("44.00"),
            daily_max_hours=Decimal("8.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="18:00-00:00",
                day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_HOURS,
                day_1_compensation_hours="2",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_form_rejects_second_shift_when_first_shift_already_meets_daily_journey(self):
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="135",
            employee_name="Empleado Demo 10",
            weekly_target_hours=Decimal("24.00"),
            daily_max_hours=Decimal("4.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_1_shift_1="06:00-10:00",
                day_1_shift_2="13:00-17:00",
            ),
            instance=line,
            schedule=self.schedule,
            shift_choices=[],
            secondary_shift_choices=[],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("cumple la jornada", form.errors["day_1_shift_2"][0].lower())

    def test_same_week_balance_is_carried_by_employee_between_sites(self):
        first_site = Site.objects.create(code="009", name="RIOJA")
        second_site = Site.objects.create(code="001", name="ALAMOS")
        first_schedule = WeeklySchedule.objects.create(
            site=first_site,
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        second_schedule = WeeklySchedule.objects.create(
            site=second_site,
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="2000",
            employee_name="Empleado Trasladado",
            initial_day_balance=Decimal("6.00"),
        )
        first_line = ScheduleLine.objects.create(
            schedule=first_schedule,
            employee_identifier="2000",
            employee_name="Empleado Trasladado",
            weekly_target_hours=Decimal("36.00"),
            daily_max_hours=Decimal("8.00"),
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_3_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_4_shift_1="18:00-00:00",
            day_5_shift_1="18:00-00:00",
            day_6_shift_1="18:00-00:00",
        )
        second_line = ScheduleLine.objects.create(
            schedule=second_schedule,
            employee_identifier="2000",
            employee_name="Empleado Trasladado",
            weekly_target_hours=Decimal("36.00"),
            daily_max_hours=Decimal("8.00"),
        )

        rebuild_balances_for_employees_from_week(first_schedule.week_start_date, ["2000"])
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.accrued_day_balance, Decimal("3.00"))
        self.assertEqual(second_line.accrued_day_balance, Decimal("3.00"))
        self.assertEqual(second_line.accrued_hour_balance, Decimal("0.00"))

    def test_same_week_balance_uses_real_day_order_even_if_later_site_schedule_was_created_first(self):
        later_site_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="014", name="JARDIN.N"),
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        earlier_site_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="015", name="RIOJA"),
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="2001",
            employee_name="Empleado Orden Real",
            initial_day_balance=Decimal("6.00"),
        )
        later_line = ScheduleLine.objects.create(
            schedule=later_site_schedule,
            employee_identifier="2001",
            employee_name="Empleado Orden Real",
            daily_max_hours=Decimal("8.00"),
            day_4_shift_1="08:00-16:00",
            day_5_shift_1="08:00-16:00",
            day_6_shift_1="08:00-16:00",
        )
        earlier_line = ScheduleLine.objects.create(
            schedule=earlier_site_schedule,
            employee_identifier="2001",
            employee_name="Empleado Orden Real",
            daily_max_hours=Decimal("8.00"),
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_3_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )

        rebuild_balances_for_employees_from_week(earlier_site_schedule.week_start_date, ["2001"])
        earlier_line.refresh_from_db()
        later_line.refresh_from_db()

        self.assertEqual(earlier_line.accrued_day_balance, Decimal("3.00"))
        self.assertEqual(later_line.accrued_day_balance, Decimal("3.00"))

    def test_same_week_manual_adjustment_is_carried_to_following_site_balance(self):
        later_site_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="016", name="JARDIN.N"),
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        earlier_site_schedule = WeeklySchedule.objects.create(
            site=Site.objects.create(code="017", name="RIOJA"),
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="2002",
            employee_name="Empleado Ajuste",
            initial_day_balance=Decimal("6.00"),
        )
        later_line = ScheduleLine.objects.create(
            schedule=later_site_schedule,
            employee_identifier="2002",
            employee_name="Empleado Ajuste",
            daily_max_hours=Decimal("8.00"),
            day_4_shift_1="08:00-16:00",
            day_5_shift_1="08:00-16:00",
        )
        earlier_line = ScheduleLine.objects.create(
            schedule=earlier_site_schedule,
            employee_identifier="2002",
            employee_name="Empleado Ajuste",
            daily_max_hours=Decimal("8.00"),
            day_1_shift_1="08:00-16:00",
            day_2_shift_1="08:00-16:00",
            day_3_shift_1="08:00-16:00",
            manual_day_adjustment=Decimal("-3.00"),
        )

        rebuild_balances_for_employees_from_week(earlier_site_schedule.week_start_date, ["2002"])
        earlier_line.refresh_from_db()
        later_line.refresh_from_db()

        self.assertEqual(earlier_line.accrued_day_balance, Decimal("4.00"))
        self.assertEqual(later_line.accrued_day_balance, Decimal("4.00"))

    @patch("schedules.services.fetch_active_staff_for_site", return_value=[])
    def test_sync_schedule_uses_schedule_week_start_for_legacy_period(self, mock_fetch_staff):
        created_count, updated_count = sync_schedule_from_legacy(self.schedule)

        self.assertEqual(created_count, 0)
        self.assertEqual(updated_count, 0)
        mock_fetch_staff.assert_called_once_with(
            self.site.code,
            week_start_date=self.schedule.week_start_date,
        )

    @patch("schedules.services.fetch_active_staff_for_site")
    def test_sync_schedule_excludes_blacklisted_employee(self, mock_fetch_staff):
        EmployeeScheduleBlacklist.objects.create(
            employee_identifier="36178712",
            employee_name="Persona Bloqueada",
        )
        mock_fetch_staff.return_value = [
            OperationalStaffingRecord(
                employee_id="36178712",
                employee_name="Persona Bloqueada",
                site_code=self.site.code,
                department_code="A1",
                department_name="ABASTOS",
                role_code="AUX",
                role_name="AUXILIAR",
            ),
            OperationalStaffingRecord(
                employee_id="99887766",
                employee_name="Persona Permitida",
                site_code=self.site.code,
                department_code="B1",
                department_name="CARNES",
                role_code="CARN",
                role_name="CARNICERO",
            ),
        ]

        created_count, updated_count = sync_schedule_from_legacy(self.schedule)

        self.assertEqual(created_count, 1)
        self.assertEqual(updated_count, 0)
        self.assertFalse(
            ScheduleLine.objects.filter(schedule=self.schedule, employee_identifier="36178712").exists()
        )
        self.assertTrue(
            ScheduleLine.objects.filter(schedule=self.schedule, employee_identifier="99887766").exists()
        )

    @patch("schedules.services.fetch_active_staff_for_site")
    def test_sync_schedule_preserves_existing_line_workload_snapshot_after_role_change(self, mock_fetch_staff):
        job_role = JobRole.objects.create(
            code="AUX",
            name="AUXILIAR",
            weekly_target_hours=Decimal("44.00"),
            daily_max_hours=Decimal("9.00"),
            base_work_days=6,
        )
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="5001",
            employee_name="Empleado Jornada",
            job_role_code="AUX",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("44.00"),
            daily_max_hours=Decimal("9.00"),
            base_work_days=6,
        )
        job_role.weekly_target_hours = Decimal("42.00")
        job_role.save(update_fields=["weekly_target_hours", "updated_at"])
        mock_fetch_staff.return_value = [
            OperationalStaffingRecord(
                employee_id="5001",
                employee_name="Empleado Jornada Actualizado",
                site_code=self.site.code,
                department_code="A1",
                department_name="ABASTOS",
                role_code="AUX",
                role_name="AUXILIAR",
            ),
        ]

        created_count, updated_count = sync_schedule_from_legacy(self.schedule)

        line = ScheduleLine.objects.get(schedule=self.schedule, employee_identifier="5001")
        self.assertEqual(created_count, 0)
        self.assertEqual(updated_count, 1)
        self.assertEqual(line.employee_name, "Empleado Jornada Actualizado")
        self.assertEqual(line.weekly_target_hours, Decimal("44.00"))
        self.assertEqual(line.daily_max_hours, Decimal("9.00"))
        self.assertEqual(line.base_work_days, 6)

    @patch("schedules.services.fetch_active_staff_for_site")
    def test_sync_schedule_uses_current_role_workload_for_new_schedule_lines(self, mock_fetch_staff):
        JobRole.objects.create(
            code="AUX",
            name="AUXILIAR",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("9.00"),
            base_work_days=6,
        )
        next_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 19),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        mock_fetch_staff.return_value = [
            OperationalStaffingRecord(
                employee_id="5002",
                employee_name="Empleado Nuevo",
                site_code=self.site.code,
                department_code="A1",
                department_name="ABASTOS",
                role_code="AUX",
                role_name="AUXILIAR",
            ),
        ]

        created_count, updated_count = sync_schedule_from_legacy(next_schedule)

        line = ScheduleLine.objects.get(schedule=next_schedule, employee_identifier="5002")
        self.assertEqual(created_count, 1)
        self.assertEqual(updated_count, 0)
        self.assertEqual(line.weekly_target_hours, Decimal("42.00"))
        self.assertEqual(line.daily_max_hours, Decimal("9.00"))
        self.assertEqual(line.base_work_days, 6)

    def test_sync_schedule_loads_blacklisted_staff_into_personal_vario(self):
        personal_vario, _ = Site.objects.get_or_create(
            code=Site.PERSONAL_VARIO_CODE,
            defaults={
                "name": Site.PERSONAL_VARIO_NAME,
                "admin_only": True,
                "is_active": True,
            },
        )
        OperationalStaffCache.objects.create(
            site_code="007",
            employee_identifier="36178712",
            employee_name="Persona Bloqueada",
            department_code="TR",
            department_name="TRANSPORTE",
            role_code="COND",
            role_name="CONDUCTOR",
        )
        EmployeeScheduleBlacklist.objects.create(
            employee_identifier="36178712",
            employee_name="Persona Bloqueada",
        )
        schedule = WeeklySchedule.objects.create(
            site=personal_vario,
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )

        created_count, updated_count = sync_schedule_from_legacy(schedule)

        self.assertEqual(created_count, 1)
        self.assertEqual(updated_count, 0)
        created_line = ScheduleLine.objects.get(schedule=schedule, employee_identifier="36178712")
        self.assertEqual(created_line.employee_name, "Persona Bloqueada")
        self.assertEqual(created_line.job_role_name, "CONDUCTOR")
        self.assertEqual(created_line.department_name, "TRANSPORTE")


class ProportionalWeeklyBalanceTests(TestCase):
    def setUp(self):
        config = SystemConfiguration.load()
        config.default_weekly_hours = Decimal("42.00")
        config.default_daily_max_hours = Decimal("9.00")
        config.save()
        Holiday.objects.all().delete()
        base_holidays_patch = patch("schedules.calendar_utils.get_base_colombian_holidays", return_value=set())
        self.addCleanup(base_holidays_patch.stop)
        base_holidays_patch.start()
        self.site = Site.objects.create(code="042", name="PRUEBA")

    def ensure_holiday(self, holiday_date: date, name: str = "Festivo prueba"):
        Holiday.objects.get_or_create(holiday_date=holiday_date, defaults={"name": name})

    def build_form_data(self, **overrides):
        data = {}
        for index in range(7):
            data[f"day_{index}_shift_1"] = ""
            data[f"day_{index}_shift_2"] = ""
            data[f"day_{index}_compensation_mode"] = ""
            data[f"day_{index}_compensation_hours"] = ""

        data.update(overrides)
        return data

    def build_line(
        self,
        *,
        week_start: date,
        employee_identifier: str,
        weekly_target_hours: Decimal,
        shift_map: dict[int, str] | None = None,
        second_shift_map: dict[int, str] | None = None,
        compensation_map: dict[int, str] | None = None,
        compensation_hours_map: dict[int, Decimal | str] | None = None,
        daily_max_hours: Decimal = Decimal("9.00"),
    ) -> ScheduleLine:
        schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=week_start,
            first_day_index=SystemConfiguration.SUNDAY,
        )
        payload = {
            "schedule": schedule,
            "employee_identifier": employee_identifier,
            "employee_name": f"Empleado {employee_identifier}",
            "job_role_name": "AUXILIAR",
            "weekly_target_hours": weekly_target_hours,
            "daily_max_hours": daily_max_hours,
        }
        for index, label in (shift_map or {}).items():
            payload[f"day_{index}_shift_1"] = label
        for index, label in (second_shift_map or {}).items():
            payload[f"day_{index}_shift_2"] = label
        for index, mode in (compensation_map or {}).items():
            payload[f"day_{index}_compensation_mode"] = mode
        for index, hours in (compensation_hours_map or {}).items():
            payload[f"day_{index}_compensation_hours"] = Decimal(str(hours))
        return ScheduleLine.objects.create(**payload)

    def rebuild_employee(self, week_start: date, employee_identifier: str):
        rebuild_balances_for_employees_from_week(week_start, [employee_identifier])

    def test_42_hour_week_without_holidays_balances_to_zero(self):
        line = self.build_line(
            week_start=date(2026, 8, 2),
            employee_identifier="4201",
            weekly_target_hours=Decimal("42.00"),
            shift_map={1: "09:00-16:00", 2: "09:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 6)
        self.assertEqual(line.expected_weekly_hours, Decimal("42.00"))
        self.assertEqual(line.total_hours, Decimal("42.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))

    def test_42_hour_week_with_one_holiday_balances_to_zero(self):
        self.ensure_holiday(date(2026, 8, 11), "Festivo martes")
        line = self.build_line(
            week_start=date(2026, 8, 9),
            employee_identifier="4202",
            weekly_target_hours=Decimal("42.00"),
            shift_map={1: "09:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 5)
        self.assertEqual(line.expected_weekly_hours, Decimal("35.00"))
        self.assertEqual(line.total_hours, Decimal("35.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))

    def test_contractacion_days_reduce_weekly_journey_without_balance_movement(self):
        line = self.build_line(
            week_start=date(2026, 8, 16),
            employee_identifier="4202C",
            weekly_target_hours=Decimal("42.00"),
            shift_map={
                0: "contratacion",
                1: "contratacion",
                2: "contratacion",
                3: "09:00-16:00",
                4: "09:00-16:00",
                5: "09:00-16:00",
                6: "09:00-16:00",
            },
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 4)
        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("28.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))
        self.assertEqual(ScheduleBalanceMovement.objects.filter(schedule=line.schedule).count(), 0)

    def test_paid_hours_count_for_weekly_journey_without_generating_new_overtime(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4210",
            employee_name="Empleado 4210",
            initial_hour_balance=Decimal("5.00"),
        )
        line = self.build_line(
            week_start=date(2026, 8, 16),
            employee_identifier="4210",
            weekly_target_hours=Decimal("44.00"),
            shift_map={
                0: "14:00-21:00",
                1: "14:00-21:00",
                2: "07:00-14:00",
                3: "09:00-16:00",
                4: "14:00-21:00",
                5: "descanso",
                6: "10:00-18:00",
            },
            compensation_map={
                2: ScheduleLine.CompensationMode.PAY_HOURS,
                4: ScheduleLine.CompensationMode.PAY_HOURS,
                5: ScheduleLine.CompensationMode.PAY_DAY,
            },
            compensation_hours_map={2: "1.00", 4: "1.00"},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.payment_hours_used, Decimal("2.00"))
        self.assertEqual(line.total_hours, Decimal("45.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("1.00"))
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("3.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.OVERPLANNED)

    def test_only_paid_hours_can_cover_adjusted_weekly_journey(self):
        self.ensure_holiday(date(2026, 8, 17), "Festivo lunes")
        EmployeeInitialBalance.objects.create(
            employee_identifier="4211",
            employee_name="Empleado 4211",
            initial_day_balance=Decimal("1.00"),
            initial_hour_balance=Decimal("28.00"),
        )
        line = self.build_line(
            week_start=date(2026, 8, 16),
            employee_identifier="4211",
            weekly_target_hours=Decimal("42.00"),
            shift_map={0: "descanso", 1: "festivo", 6: "descanso"},
            compensation_map={
                2: ScheduleLine.CompensationMode.PAY_HOURS,
                3: ScheduleLine.CompensationMode.PAY_HOURS,
                4: ScheduleLine.CompensationMode.PAY_HOURS,
                5: ScheduleLine.CompensationMode.PAY_HOURS,
                6: ScheduleLine.CompensationMode.PAY_DAY,
            },
            compensation_hours_map={2: "7.00", 3: "7.00", 4: "7.00", 5: "7.00"},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.payment_hours_used, Decimal("28.00"))
        self.assertEqual(line.expected_work_days, 4)
        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("28.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.VALID)

    def test_sunday_and_holiday_generate_days_and_weekly_rest_pay_day_consumes_one(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4301",
            weekly_target_hours=Decimal("43.00"),
            shift_map={
                0: "08:00-15:00",
                1: "08:00-16:00",
                2: "descanso",
                3: "08:00-15:00",
                4: "08:00-15:00",
                5: "08:00-15:00",
                6: "08:00-16:00",
            },
            compensation_map={2: ScheduleLine.CompensationMode.PAY_DAY},
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        movements = list(
            line.balance_movements.filter(is_reversal=False, is_reversed=False).order_by("movement_date", "pk")
        )

        self.assertEqual(line.special_days_generated, 2)
        self.assertEqual(line.payment_days_used, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(line.expected_weekly_hours, Decimal("43.00"))
        self.assertEqual(line.total_hours, Decimal("44.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("1.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.OVERPLANNED)
        self.assertIn("1 dia(s) generado(s) por domingo trabajado", line.validation_summary)
        self.assertIn("1 dia(s) generado(s) por festivo trabajado", line.validation_summary)
        self.assertIn("1 dia(s) consumido(s) mediante Pago dia", line.validation_summary)
        self.assertIn("Saldo neto dias: 1 dia(s) a favor del trabajador.", line.validation_summary)

        self.assertEqual(
            [movement.movement_type for movement in movements],
            [
                ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
                ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
                ScheduleBalanceMovement.MovementType.PAY_DAY,
                ScheduleBalanceMovement.MovementType.OVERTIME,
            ],
        )
        self.assertEqual(movements[0].description, "Domingo laborado")
        self.assertEqual(movements[1].description, "Festivo laborado")
        self.assertEqual(movements[2].quantity_days, Decimal("-1.00"))

        compact_summary = get_schedule_line_compact_alert_summary(line)
        self.assertIn("Resultado estimado: 1 dia(s) a favor del trabajador y 1.00 h de excedente.", compact_summary)
        self.assertIn("1 dia(s) generado(s) por domingo trabajado", compact_summary)
        self.assertIn("1 dia(s) generado(s) por festivo trabajado", compact_summary)
        self.assertIn("1 dia(s) consumido(s) mediante Pago dia", compact_summary)

    def test_same_week_special_days_and_pay_day_keep_positive_starting_balance(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        EmployeeInitialBalance.objects.create(
            employee_identifier="4302",
            employee_name="Empleado 4302",
            initial_day_balance=Decimal("1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4302",
            weekly_target_hours=Decimal("43.00"),
            shift_map={
                0: "08:00-15:00",
                1: "08:00-16:00",
                2: "descanso",
                3: "08:00-15:00",
                4: "08:00-15:00",
                5: "08:00-15:00",
                6: "08:00-16:00",
            },
            compensation_map={2: ScheduleLine.CompensationMode.PAY_DAY},
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.special_days_generated, 2)
        self.assertEqual(line.payment_days_used, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("2.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("1.00"))

    def test_bound_form_display_metrics_use_recalculated_positive_starting_balance(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 12),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="4304",
            employee_name="Empleado 4304",
            initial_day_balance=Decimal("1.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=schedule,
            employee_identifier="4304",
            employee_name="Empleado 4304",
            weekly_target_hours=Decimal("43.00"),
            daily_max_hours=Decimal("9.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-15:00",
                day_1_shift_1="08:00-16:00",
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
                day_3_shift_1="08:00-15:00",
                day_4_shift_1="08:00-15:00",
                day_5_shift_1="08:00-15:00",
                day_6_shift_1="08:00-16:00",
            ),
            instance=line,
            schedule=schedule,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.balance_snapshot["prior_day_balance"], Decimal("1.00"))
        self.assertEqual(form.display_day_balance, Decimal("2.00"))
        self.assertEqual(form.display_hour_balance, Decimal("1.00"))
        self.assertIn("Resultado estimado: 2 dia(s) a favor del trabajador", form.compact_alert_summary)

    def test_bound_form_display_metrics_show_one_day_when_no_historical_balance_exists(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 12),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        line = ScheduleLine.objects.create(
            schedule=schedule,
            employee_identifier="4305",
            employee_name="Empleado 4305",
            weekly_target_hours=Decimal("43.00"),
            daily_max_hours=Decimal("9.00"),
        )
        form = ScheduleLineForm(
            data=self.build_form_data(
                day_0_shift_1="08:00-15:00",
                day_1_shift_1="08:00-16:00",
                day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
                day_3_shift_1="08:00-15:00",
                day_4_shift_1="08:00-15:00",
                day_5_shift_1="08:00-15:00",
                day_6_shift_1="08:00-16:00",
            ),
            instance=line,
            schedule=schedule,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.balance_snapshot["prior_day_balance"], Decimal("0.00"))
        self.assertEqual(form.display_day_balance, Decimal("1.00"))
        self.assertEqual(form.display_hour_balance, Decimal("1.00"))
        self.assertIn("Resultado estimado: 1 dia(s) a favor del trabajador", form.compact_alert_summary)

    def test_shifted_weekly_rest_without_pay_day_keeps_both_generated_days(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4303",
            weekly_target_hours=Decimal("43.00"),
            shift_map={
                0: "08:00-15:00",
                1: "08:00-16:00",
                2: "descanso",
                3: "08:00-15:00",
                4: "08:00-15:00",
                5: "08:00-15:00",
                6: "08:00-16:00",
            },
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.special_days_generated, 2)
        self.assertEqual(line.payment_days_used, 0)
        self.assertEqual(line.accrued_day_balance, Decimal("2.00"))
        self.assertEqual(line.expected_weekly_hours, Decimal("43.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("1.00"))

    def test_non_worked_holiday_with_worked_sunday_does_not_create_company_day_debt(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4306",
            weekly_target_hours=Decimal("50.00"),
            shift_map={
                0: "07:00-15:00",
                1: "descanso",
                2: "07:00-13:00",
                3: "07:00-13:00",
                4: "07:00-13:00",
                5: "07:00-15:00",
                6: "07:00-15:00",
            },
            second_shift_map={
                2: "14:00-17:00",
                3: "14:00-17:00",
                4: "14:00-17:00",
            },
            daily_max_hours=Decimal("10.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        expected_plan = build_expected_week_plan(line)
        monday_plan = next(day for day in expected_plan["day_plans"] if day["index"] == 1)

        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.payment_days_used, 0)
        self.assertEqual(line.advance_rest_days_used, 0)
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(monday_plan["expected_reason"], "descanso_obligatorio")
        self.assertTrue(monday_plan["is_non_worked_holiday"])
        self.assertFalse(monday_plan["is_additional_rest_day"])

    def test_blank_non_worked_holiday_with_worked_sunday_can_be_weekly_rest(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4306B",
            weekly_target_hours=Decimal("50.00"),
            shift_map={
                0: "07:00-15:00",
                2: "07:00-13:00",
                3: "07:00-13:00",
                4: "07:00-13:00",
                5: "07:00-15:00",
                6: "07:00-15:00",
            },
            second_shift_map={
                2: "14:00-17:00",
                3: "14:00-17:00",
                4: "14:00-17:00",
            },
            daily_max_hours=Decimal("10.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        expected_plan = build_expected_week_plan(line)
        monday_plan = next(day for day in expected_plan["day_plans"] if day["index"] == 1)

        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(line.expected_work_days, 6)
        self.assertEqual(line.expected_weekly_hours, Decimal("50.00"))
        self.assertEqual(expected_plan["mandatory_rest_index"], 1)
        self.assertEqual(monday_plan["expected_reason"], "descanso_obligatorio")
        self.assertTrue(monday_plan["is_non_worked_holiday"])
        self.assertFalse(monday_plan["is_additional_rest_day"])

    def test_worked_holiday_and_pay_day_from_zero_never_creates_negative_day_balance(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4307",
            weekly_target_hours=Decimal("46.00"),
            shift_map={
                0: "descanso",
                1: "07:00-12:00",
                2: "10:00-14:00",
                3: "10:00-14:00",
                4: "10:00-14:00",
                5: "descanso",
                6: "09:00-14:00",
            },
            second_shift_map={
                1: "17:00-21:00",
                2: "16:00-21:00",
                3: "16:00-21:00",
                4: "16:00-21:00",
                6: "16:00-21:00",
            },
            compensation_map={5: ScheduleLine.CompensationMode.PAY_DAY},
            daily_max_hours=Decimal("10.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.payment_days_used, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(
            list(
                line.balance_movements.filter(is_reversal=False, is_reversed=False)
                .order_by("movement_date", "pk")
                .values_list("movement_type", "quantity_days")
            ),
            [
                (ScheduleBalanceMovement.MovementType.SPECIAL_DAY, Decimal("1.00")),
                (ScheduleBalanceMovement.MovementType.PAY_DAY, Decimal("-1.00")),
            ],
        )

    def test_rebuild_reverses_legacy_duplicate_special_day_movements_without_idempotency_key(self):
        self.ensure_holiday(date(2026, 7, 13), "Festivo lunes")
        line = self.build_line(
            week_start=date(2026, 7, 12),
            employee_identifier="4303_DUP",
            weekly_target_hours=Decimal("43.00"),
            shift_map={
                0: "08:00-15:00",
                1: "08:00-16:00",
                2: "descanso",
                3: "08:00-15:00",
                4: "08:00-15:00",
                5: "08:00-15:00",
                6: "08:00-16:00",
            },
            compensation_map={2: ScheduleLine.CompensationMode.PAY_DAY},
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        sunday_movement = line.balance_movements.filter(
            is_reversal=False,
            is_reversed=False,
            movement_type=ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
            movement_date=date(2026, 7, 12),
        ).first()
        holiday_movement = line.balance_movements.filter(
            is_reversal=False,
            is_reversed=False,
            movement_type=ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
            movement_date=date(2026, 7, 13),
        ).first()
        self.assertIsNotNone(sunday_movement)
        self.assertIsNotNone(holiday_movement)

        for original in (sunday_movement, holiday_movement):
            ScheduleBalanceMovement.objects.create(
                schedule=line.schedule,
                line=line,
                site=line.schedule.site,
                employee_identifier=line.employee_identifier,
                employee_name=line.employee_name,
                job_role_name=line.job_role_name,
                movement_date=original.movement_date,
                movement_type=original.movement_type,
                quantity_days=original.quantity_days,
                quantity_hours=original.quantity_hours,
                equivalent_hours=original.equivalent_hours,
                balance_before_days=original.balance_before_days,
                balance_after_days=original.balance_after_days,
                movement_origin="legacy",
                idempotency_key="",
                description=original.description,
            )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        active_special_days = list(
            line.balance_movements.filter(
                is_reversal=False,
                is_reversed=False,
                movement_type=ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
            ).order_by("movement_date", "pk")
        )
        reversed_legacy = list(
            line.balance_movements.filter(
                movement_origin="legacy",
                is_reversal=False,
                is_reversed=True,
            ).order_by("pk")
        )
        legacy_reversals = list(
            line.balance_movements.filter(
                is_reversal=True,
                reversed_movement__in=reversed_legacy,
            ).order_by("pk")
        )

        self.assertEqual(
            [(movement.movement_date, movement.description) for movement in active_special_days],
            [
                (date(2026, 7, 12), "Domingo laborado"),
                (date(2026, 7, 13), "Festivo laborado"),
            ],
        )
        self.assertEqual(len(reversed_legacy), 2)
        self.assertEqual(len(legacy_reversals), 2)
        self.assertTrue(all(reversal.idempotency_key for reversal in legacy_reversals))
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))

    def test_compensatory_days_reduce_expected_hours_proportionally(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4203",
            employee_name="Empleado 4203",
            initial_day_balance=Decimal("2.00"),
        )
        line = self.build_line(
            week_start=date(2026, 8, 16),
            employee_identifier="4203",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={1: ScheduleLine.CompensationMode.PAY_DAY, 2: ScheduleLine.CompensationMode.PAY_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 4)
        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("28.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))

    def test_additional_rest_days_without_pay_day_reduce_balance_and_journey(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4203B",
            employee_name="Empleado 4203B",
            initial_day_balance=Decimal("3.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 1),
            employee_identifier="4203B",
            weekly_target_hours=Decimal("42.00"),
            shift_map={
                1: "descanso",
                2: "descanso",
                3: "09:00-16:00",
                4: "09:00-16:00",
                5: "09:00-16:00",
                6: "09:00-16:00",
            },
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 4)
        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("28.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(line.advance_rest_pending_balance, Decimal("0.00"))

    def test_additional_rest_day_without_balance_creates_company_day_debt(self):
        line = self.build_line(
            week_start=date(2026, 11, 8),
            employee_identifier="4203C",
            weekly_target_hours=Decimal("42.00"),
            shift_map={
                1: "descanso",
                2: "09:00-16:00",
                3: "09:00-16:00",
                4: "09:00-16:00",
                5: "09:00-16:00",
                6: "09:00-16:00",
            },
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 5)
        self.assertEqual(line.expected_weekly_hours, Decimal("35.00"))
        self.assertEqual(line.total_hours, Decimal("35.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))
        self.assertEqual(line.advance_rest_pending_balance, Decimal("1.00"))

    def test_additional_rest_day_extends_negative_balance_until_minus_two(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4203D",
            employee_name="Empleado 4203D",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 15),
            employee_identifier="4203D",
            weekly_target_hours=Decimal("42.00"),
            shift_map={
                1: "descanso",
                2: "09:00-16:00",
                3: "09:00-16:00",
                4: "09:00-16:00",
                5: "09:00-16:00",
                6: "09:00-16:00",
            },
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 5)
        self.assertEqual(line.expected_weekly_hours, Decimal("35.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-2.00"))

    def test_underworked_week_does_not_generate_negative_hour_balance(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4204",
            employee_name="Empleado 4204",
            initial_day_balance=Decimal("2.00"),
        )
        line = self.build_line(
            week_start=date(2026, 8, 23),
            employee_identifier="4204",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-12:00"},
            compensation_map={1: ScheduleLine.CompensationMode.PAY_DAY, 2: ScheduleLine.CompensationMode.PAY_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("24.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("-4.00"))
        self.assertEqual(line.accrued_hour_balance, Decimal("0.00"))
        self.assertNotIn("No cumple las horas esperadas de la semana", line.validation_summary)

    def test_future_extra_hours_accumulate_without_crossing_negative_hour_balance(self):
        first_line = self.build_line(
            week_start=date(2026, 8, 30),
            employee_identifier="4205",
            weekly_target_hours=Decimal("42.00"),
            shift_map={1: "09:00-16:00", 2: "09:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-12:00"},
        )
        second_line = self.build_line(
            week_start=date(2026, 9, 6),
            employee_identifier="4205",
            weekly_target_hours=Decimal("42.00"),
            shift_map={1: "08:00-16:00", 2: "08:00-16:00", 3: "08:00-16:00", 4: "08:00-16:00", 5: "08:00-16:00", 6: "08:00-16:00"},
        )

        self.rebuild_employee(first_line.schedule.week_start_date, first_line.employee_identifier)
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.accrued_hour_balance, Decimal("0.00"))
        self.assertEqual(second_line.weekly_hour_difference, Decimal("6.00"))
        self.assertEqual(second_line.accrued_hour_balance, Decimal("6.00"))

    def test_44_hour_week_uses_proportional_daily_base(self):
        line = self.build_line(
            week_start=date(2026, 9, 13),
            employee_identifier="4401",
            weekly_target_hours=Decimal("44.00"),
            shift_map={1: "08:00-16:00", 2: "08:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        snapshot = get_schedule_line_balance_snapshot(line)

        self.assertEqual(snapshot["day_reference_hours"], Decimal("7.33"))
        self.assertEqual(line.expected_weekly_hours, Decimal("44.00"))
        self.assertEqual(line.total_hours, Decimal("44.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))

    def test_52_hour_week_uses_proportional_daily_base(self):
        line = self.build_line(
            week_start=date(2026, 9, 20),
            employee_identifier="5201",
            weekly_target_hours=Decimal("52.00"),
            shift_map={1: "08:00-17:00", 2: "08:00-17:00", 3: "08:00-17:00", 4: "08:00-17:00", 5: "08:00-16:00", 6: "08:00-16:00"},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        snapshot = get_schedule_line_balance_snapshot(line)

        self.assertEqual(snapshot["day_reference_hours"], Decimal("8.67"))
        self.assertEqual(line.expected_weekly_hours, Decimal("52.00"))
        self.assertEqual(line.total_hours, Decimal("52.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))

    def test_52_hour_week_rounds_adjusted_journey_to_nearest_half_hour(self):
        scenarios = [
            (0, "52:00", Decimal("52.00")),
            (1, "43:20", Decimal("43.50")),
            (2, "34:40", Decimal("34.50")),
            (3, "26:00", Decimal("26.00")),
            (4, "17:20", Decimal("17.50")),
            (5, "8:40", Decimal("8.50")),
            (6, "0:00", Decimal("0.00")),
        ]

        for rest_days, expected_exact_clock, expected_programmable_hours in scenarios:
            with self.subTest(rest_days=rest_days):
                shift_map = {
                    index: "08:00-16:00"
                    for index in range(rest_days + 1, 7)
                }
                shift_map.update({index: "descanso" for index in range(1, rest_days + 1)})
                line = self.build_line(
                    week_start=date(2026, 9, 27) + timedelta(days=7 * rest_days),
                    employee_identifier=f"5201-{rest_days}",
                    weekly_target_hours=Decimal("52.00"),
                    shift_map=shift_map,
                )

                expected_plan = build_expected_week_plan(line)

                self.assertEqual(
                    int(expected_plan["expected_weekly_exact_minutes"]),
                    ScheduleRoundingRuleTests.minutes_from_clock(expected_exact_clock),
                )
                self.assertEqual(expected_plan["expected_weekly_hours"], expected_programmable_hours)

    def test_44_hour_week_rounds_adjusted_journey_to_nearest_half_hour(self):
        scenarios = [
            (1, "36:40", Decimal("36.50")),
            (2, "29:20", Decimal("29.50")),
            (3, "22:00", Decimal("22.00")),
            (5, "7:20", Decimal("7.50")),
        ]

        for rest_days, expected_exact_clock, expected_programmable_hours in scenarios:
            with self.subTest(rest_days=rest_days):
                shift_map = {
                    index: "08:00-16:00"
                    for index in range(rest_days + 1, 7)
                }
                shift_map.update({index: "descanso" for index in range(1, rest_days + 1)})
                line = self.build_line(
                    week_start=date(2026, 10, 18) + timedelta(days=7 * rest_days),
                    employee_identifier=f"4401-{rest_days}",
                    weekly_target_hours=Decimal("44.00"),
                    shift_map=shift_map,
                )

                expected_plan = build_expected_week_plan(line)

                self.assertEqual(
                    int(expected_plan["expected_weekly_exact_minutes"]),
                    ScheduleRoundingRuleTests.minutes_from_clock(expected_exact_clock),
                )
                self.assertEqual(expected_plan["expected_weekly_hours"], expected_programmable_hours)

    def test_42_hour_week_keeps_exact_adjusted_journey_without_rounding_noise(self):
        scenarios = [
            (1, "35:00", Decimal("35.00")),
            (2, "28:00", Decimal("28.00")),
            (3, "21:00", Decimal("21.00")),
        ]

        for rest_days, expected_exact_clock, expected_programmable_hours in scenarios:
            with self.subTest(rest_days=rest_days):
                shift_map = {
                    index: "09:00-16:00"
                    for index in range(rest_days + 1, 7)
                }
                shift_map.update({index: "descanso" for index in range(1, rest_days + 1)})
                line = self.build_line(
                    week_start=date(2026, 11, 29) + timedelta(days=7 * rest_days),
                    employee_identifier=f"4201-{rest_days}",
                    weekly_target_hours=Decimal("42.00"),
                    shift_map=shift_map,
                )

                expected_plan = build_expected_week_plan(line)

                self.assertEqual(
                    int(expected_plan["expected_weekly_exact_minutes"]),
                    ScheduleRoundingRuleTests.minutes_from_clock(expected_exact_clock),
                )
                self.assertEqual(expected_plan["expected_weekly_hours"], expected_programmable_hours)

    def test_programmable_weekly_journey_validates_against_rounded_result(self):
        line = self.build_line(
            week_start=date(2026, 12, 13),
            employee_identifier="5202",
            weekly_target_hours=Decimal("52.00"),
            shift_map={
                1: "descanso",
                2: "08:00-16:00",
                3: "08:00-16:30",
                4: "08:00-17:00",
                5: "08:00-17:00",
                6: "08:00-17:00",
            },
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("43.50"))
        self.assertEqual(line.total_hours, Decimal("43.50"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.VALID)
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))

    def test_two_reducers_round_final_weekly_result_only(self):
        line = self.build_line(
            week_start=date(2026, 12, 20),
            employee_identifier="5203",
            weekly_target_hours=Decimal("52.00"),
            shift_map={
                1: "descanso",
                2: "descanso",
                3: "08:00-16:30",
                4: "08:00-16:30",
                5: "08:00-16:30",
                6: "08:00-17:00",
            },
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        expected_plan = build_expected_week_plan(line)

        self.assertEqual(
            int(expected_plan["expected_weekly_exact_minutes"]),
            ScheduleRoundingRuleTests.minutes_from_clock("34:40"),
        )
        self.assertEqual(line.expected_weekly_hours, Decimal("34.50"))
        self.assertNotEqual(line.expected_weekly_hours, Decimal("35.00"))
        self.assertEqual(line.total_hours, Decimal("34.50"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.VALID)

    def test_rounded_programmable_week_can_still_be_incomplete_by_thirty_minutes(self):
        line = self.build_line(
            week_start=date(2026, 12, 27),
            employee_identifier="5204",
            weekly_target_hours=Decimal("52.00"),
            shift_map={
                1: "descanso",
                2: "08:00-16:00",
                3: "08:00-16:30",
                4: "08:00-17:00",
                5: "08:00-17:00",
                6: "08:00-16:30",
            },
            daily_max_hours=Decimal("9.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("43.50"))
        self.assertEqual(line.total_hours, Decimal("43.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("-0.50"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.INCOMPLETE)

    def test_legacy_advance_rest_days_create_company_day_balance_until_minus_two(self):
        line = self.build_line(
            week_start=date(2026, 9, 27),
            employee_identifier="4208",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={1: ScheduleLine.CompensationMode.ADVANCE_DAY, 2: ScheduleLine.CompensationMode.ADVANCE_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-2.00"))
        self.assertEqual(line.advance_rest_pending_balance, Decimal("2.00"))

    def test_legacy_advance_rest_days_stop_at_minus_two_and_raise_alert(self):
        line = self.build_line(
            week_start=date(2026, 10, 4),
            employee_identifier="4208A",
            weekly_target_hours=Decimal("42.00"),
            shift_map={4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={
                1: ScheduleLine.CompensationMode.ADVANCE_DAY,
                2: ScheduleLine.CompensationMode.ADVANCE_DAY,
                3: ScheduleLine.CompensationMode.ADVANCE_DAY,
            },
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("21.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-2.00"))
        self.assertEqual(line.advance_rest_pending_balance, Decimal("2.00"))
        self.assertIn("limite maximo de 2 dias adelantados", line.validation_summary.lower())

    def test_generated_day_offsets_company_day_debt_first(self):
        first_line = self.build_line(
            week_start=date(2026, 10, 4),
            employee_identifier="4209",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={1: ScheduleLine.CompensationMode.ADVANCE_DAY, 2: ScheduleLine.CompensationMode.ADVANCE_DAY},
        )
        second_line = self.build_line(
            week_start=date(2026, 10, 11),
            employee_identifier="4209",
            weekly_target_hours=Decimal("42.00"),
            shift_map={0: "09:00-16:00", 1: "descanso", 2: "09:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
        )

        self.rebuild_employee(first_line.schedule.week_start_date, first_line.employee_identifier)
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.advance_rest_pending_balance, Decimal("2.00"))
        self.assertEqual(second_line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(second_line.advance_rest_pending_balance, Decimal("1.00"))

    def test_full_seventh_day_can_repay_company_day_without_new_special_day_or_overtime(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209R1",
            employee_name="Empleado 4209R1",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209R1",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            compensation_map={0: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 6)
        self.assertEqual(line.expected_weekly_hours, Decimal("48.00"))
        self.assertEqual(line.total_hours, Decimal("56.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(line.special_days_generated, 0)
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.VALID)

    def test_full_seventh_day_compensates_company_day_automatically_without_manual_selection(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO1",
            employee_name="Empleado 4209AUTO1",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209AUTO1",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("48.00"))
        self.assertEqual(line.total_hours, Decimal("56.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.special_days_generated, 0)
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))
        self.assertIn("compensacion automatica aplicada", line.validation_summary.lower())

    def test_partial_repayment_day_does_not_clear_company_day_debt(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209R2",
            employee_name="Empleado 4209R2",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 25),
            employee_identifier="4209R2",
            weekly_target_hours=Decimal("48.00"),
            shift_map={
                0: "08:00-12:00",
                1: "08:00-16:00",
                2: "08:00-16:00",
                3: "08:00-16:00",
                4: "08:00-16:00",
                5: "08:00-16:00",
                6: "08:00-16:00",
            },
            compensation_map={0: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.overtime_hours, Decimal("4.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.INCONSISTENT)
        self.assertIn("no corresponde a un dia completo", line.validation_summary.lower())

    def test_partial_seventh_day_without_manual_selection_does_not_compensate_company_day(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO2",
            employee_name="Empleado 4209AUTO2",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 25),
            employee_identifier="4209AUTO2",
            weekly_target_hours=Decimal("48.00"),
            shift_map={
                0: "08:00-12:00",
                1: "08:00-16:00",
                2: "08:00-16:00",
                3: "08:00-16:00",
                4: "08:00-16:00",
                5: "08:00-16:00",
                6: "08:00-16:00",
            },
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.special_days_generated, 0)
        self.assertEqual(line.overtime_hours, Decimal("4.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.OVERPLANNED)
        self.assertNotIn("compensacion automatica aplicada", line.validation_summary.lower())

    def test_company_day_repayment_and_holiday_generation_remain_independent(self):
        self.ensure_holiday(date(2026, 10, 19), "Festivo lunes")
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209R3",
            employee_name="Empleado 4209R3",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209R3",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            compensation_map={0: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))

    def test_automatic_company_day_repayment_and_holiday_generation_remain_independent(self):
        self.ensure_holiday(date(2026, 10, 19), "Festivo lunes")
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO3",
            employee_name="Empleado 4209AUTO3",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209AUTO3",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))

    def test_company_day_repayment_requires_additional_full_day(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209R4",
            employee_name="Empleado 4209R4",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 1),
            employee_identifier="4209R4",
            weekly_target_hours=Decimal("48.00"),
            shift_map={
                0: "08:00-16:00",
                1: "descanso",
                2: "08:00-16:00",
                3: "08:00-16:00",
                4: "08:00-16:00",
                5: "08:00-16:00",
                6: "08:00-16:00",
            },
            compensation_map={0: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("48.00"))
        self.assertEqual(line.total_hours, Decimal("48.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.validation_status, ScheduleLine.ValidationStatus.INCONSISTENT)
        self.assertIn("dia completo adicional", line.validation_summary.lower())

    def test_six_complete_days_without_manual_selection_do_not_compensate_company_day(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO4",
            employee_name="Empleado 4209AUTO4",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 1),
            employee_identifier="4209AUTO4",
            weekly_target_hours=Decimal("48.00"),
            shift_map={
                1: "08:00-16:00",
                2: "08:00-16:00",
                3: "08:00-16:00",
                4: "08:00-16:00",
                5: "08:00-16:00",
                6: "08:00-16:00",
            },
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.total_hours, Decimal("48.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.special_days_generated, 0)
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertNotIn("compensacion automatica aplicada", line.validation_summary.lower())

    def test_editing_compensation_day_back_to_rest_restores_company_day_debt(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209R5",
            employee_name="Empleado 4209R5",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 8),
            employee_identifier="4209R5",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            compensation_map={0: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))

        line.day_0_shift_1 = "descanso"
        line.day_0_compensation_mode = ""
        line.save(update_fields=["day_0_shift_1", "day_0_compensation_mode", "updated_at"])
        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.total_hours, Decimal("48.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))

    def test_editing_automatic_compensation_day_back_to_rest_restores_company_day_debt(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO5",
            employee_name="Empleado 4209AUTO5",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 11, 15),
            employee_identifier="4209AUTO5",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))

        line.day_0_shift_1 = "descanso"
        line.save(update_fields=["day_0_shift_1", "updated_at"])
        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.total_hours, Decimal("48.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))

    def test_automatic_company_day_repayment_respects_valid_manual_selection(self):
        self.ensure_holiday(date(2026, 10, 19), "Festivo lunes")
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO6",
            employee_name="Empleado 4209AUTO6",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209AUTO6",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            compensation_map={1: COMPANY_DAY_REPAYMENT_MODE},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(line.special_days_generated, 1)
        self.assertEqual(line.overtime_hours, Decimal("0.00"))
        self.assertNotIn("compensacion automatica aplicada", line.validation_summary.lower())

    def test_automatic_company_day_repayment_uses_only_prior_negative_balance(self):
        line = self.build_line(
            week_start=date(2026, 11, 22),
            employee_identifier="4209AUTO7",
            weekly_target_hours=Decimal("48.00"),
            shift_map={
                2: "08:00-16:00",
                3: "08:00-16:00",
                4: "08:00-16:00",
                5: "08:00-16:00",
                6: "08:00-16:00",
            },
            compensation_map={1: ScheduleLine.CompensationMode.ADVANCE_DAY},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_weekly_hours, Decimal("40.00"))
        self.assertEqual(line.total_hours, Decimal("40.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.special_days_generated, 0)
        self.assertEqual(line.overtime_hours, Decimal("0.00"))

    def test_automatic_company_day_repayment_requires_two_weeks_to_clear_minus_two_balance(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO8",
            employee_name="Empleado 4209AUTO8",
            initial_day_balance=Decimal("-2.00"),
        )
        first_line = self.build_line(
            week_start=date(2026, 11, 29),
            employee_identifier="4209AUTO8",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )
        second_line = self.build_line(
            week_start=date(2026, 12, 6),
            employee_identifier="4209AUTO8",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(first_line.schedule.week_start_date, first_line.employee_identifier)
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(second_line.accrued_day_balance, Decimal("0.00"))

    def test_automatic_company_day_repayment_excludes_only_one_full_day_when_extra_hours_exist(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO9",
            employee_name="Empleado 4209AUTO9",
            initial_day_balance=Decimal("-1.00"),
        )
        line = self.build_line(
            week_start=date(2026, 12, 13),
            employee_identifier="4209AUTO9",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            second_shift_map={0: "17:00-19:00"},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.total_hours, Decimal("58.00"))
        self.assertEqual(line.overtime_hours, Decimal("2.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("0.00"))
        self.assertIn("horas excluidas por compensacion de deuda: 8.00 h.", line.validation_summary.lower())

    def test_automatic_company_day_repayment_does_not_run_with_zero_or_positive_balance(self):
        zero_line = self.build_line(
            week_start=date(2026, 12, 20),
            employee_identifier="4209AUTO10",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )
        EmployeeInitialBalance.objects.create(
            employee_identifier="4209AUTO11",
            employee_name="Empleado 4209AUTO11",
            initial_day_balance=Decimal("1.00"),
        )
        positive_line = self.build_line(
            week_start=date(2026, 12, 27),
            employee_identifier="4209AUTO11",
            weekly_target_hours=Decimal("48.00"),
            shift_map={index: "08:00-16:00" for index in range(7)},
            daily_max_hours=Decimal("8.00"),
        )

        self.rebuild_employee(zero_line.schedule.week_start_date, zero_line.employee_identifier)
        self.rebuild_employee(positive_line.schedule.week_start_date, positive_line.employee_identifier)
        zero_line.refresh_from_db()
        positive_line.refresh_from_db()

        self.assertEqual(zero_line.special_days_generated, 1)
        self.assertEqual(zero_line.accrued_day_balance, Decimal("1.00"))
        self.assertNotIn("compensacion automatica aplicada", zero_line.validation_summary.lower())
        self.assertEqual(positive_line.special_days_generated, 1)
        self.assertEqual(positive_line.accrued_day_balance, Decimal("2.00"))
        self.assertNotIn("compensacion automatica aplicada", positive_line.validation_summary.lower())

    def test_shifted_weekly_rest_with_sunday_work_does_not_reduce_expected_hours_twice(self):
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209A",
            weekly_target_hours=Decimal("50.00"),
            shift_map={
                0: "07:00-12:00",
                1: "07:00-12:00",
                2: "07:00-12:00",
                4: "07:00-12:00",
                5: "07:00-12:00",
                6: "07:00-12:00",
            },
            second_shift_map={
                0: "13:00-17:00",
                1: "13:00-17:00",
                2: "13:00-17:00",
                4: "13:00-17:00",
                5: "13:00-17:00",
                6: "13:00-17:00",
            },
            compensation_map={3: ScheduleLine.CompensationMode.PAY_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 6)
        self.assertEqual(line.expected_weekly_hours, Decimal("50.00"))
        self.assertEqual(line.total_hours, Decimal("54.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("4.00"))

    def test_holiday_and_shifted_weekly_rest_only_reduce_expected_hours_once_each(self):
        self.ensure_holiday(date(2026, 10, 20), "Festivo martes")
        line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4209B",
            weekly_target_hours=Decimal("50.00"),
            shift_map={
                0: "07:00-12:00",
                1: "07:00-12:00",
                2: "07:00-12:00",
                4: "07:00-12:00",
                5: "07:00-12:00",
            },
            second_shift_map={
                0: "13:00-17:00",
                1: "13:00-17:00",
                2: "13:00-17:00",
                4: "13:00-17:00",
                5: "13:00-17:00",
            },
            compensation_map={3: ScheduleLine.CompensationMode.PAY_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 6)
        self.assertEqual(line.expected_weekly_hours, Decimal("50.00"))
        self.assertEqual(line.total_hours, Decimal("45.00"))
        self.assertEqual(line.weekly_hour_difference, Decimal("-5.00"))

    def test_two_generated_days_can_compensate_previous_company_day_debt(self):
        self.ensure_holiday(date(2026, 10, 21), "Festivo miercoles")
        first_line = self.build_line(
            week_start=date(2026, 10, 11),
            employee_identifier="4210",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={1: ScheduleLine.CompensationMode.ADVANCE_DAY},
        )
        second_line = self.build_line(
            week_start=date(2026, 10, 18),
            employee_identifier="4210",
            weekly_target_hours=Decimal("42.00"),
            shift_map={0: "09:00-16:00", 1: "descanso", 2: "09:00-16:00", 3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
        )

        self.rebuild_employee(first_line.schedule.week_start_date, first_line.employee_identifier)
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(second_line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(second_line.advance_rest_pending_balance, Decimal("0.00"))

    def test_balances_do_not_cross_between_different_workers(self):
        EmployeeInitialBalance.objects.create(employee_identifier="4211A", employee_name="Empleado A", initial_day_balance=Decimal("2.00"))
        EmployeeInitialBalance.objects.create(employee_identifier="4211B", employee_name="Empleado B", initial_hour_balance=Decimal("3.00"))
        schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 10, 25),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        first_line = ScheduleLine.objects.create(
            schedule=schedule,
            employee_identifier="4211A",
            employee_name="Empleado 4211A",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("9.00"),
            day_1_shift_1="09:00-16:00",
            day_2_shift_1="09:00-16:00",
            day_3_shift_1="09:00-16:00",
            day_4_shift_1="09:00-16:00",
            day_5_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )
        second_line = ScheduleLine.objects.create(
            schedule=schedule,
            employee_identifier="4211B",
            employee_name="Empleado 4211B",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("42.00"),
            daily_max_hours=Decimal("9.00"),
            day_1_shift_1="09:00-16:00",
            day_2_shift_1="09:00-16:00",
            day_3_shift_1="09:00-16:00",
            day_4_shift_1="09:00-16:00",
            day_5_shift_1="09:00-12:00",
        )

        self.rebuild_employee(first_line.schedule.week_start_date, "4211A")
        self.rebuild_employee(second_line.schedule.week_start_date, "4211B")
        first_line.refresh_from_db()
        second_line.refresh_from_db()

        self.assertEqual(first_line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(second_line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(second_line.accrued_hour_balance, Decimal("3.00"))

    def test_holiday_and_legacy_advance_rest_reduce_journey_and_day_balance(self):
        self.ensure_holiday(date(2026, 11, 3), "Festivo martes")
        line = self.build_line(
            week_start=date(2026, 11, 1),
            employee_identifier="4212",
            weekly_target_hours=Decimal("42.00"),
            shift_map={3: "09:00-16:00", 4: "09:00-16:00", 5: "09:00-16:00", 6: "09:00-16:00"},
            compensation_map={1: ScheduleLine.CompensationMode.ADVANCE_DAY},
        )

        self.rebuild_employee(line.schedule.week_start_date, line.employee_identifier)
        line.refresh_from_db()

        self.assertEqual(line.expected_work_days, 4)
        self.assertEqual(line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(line.total_hours, Decimal("28.00"))
        self.assertEqual(line.accrued_day_balance, Decimal("-1.00"))
        self.assertEqual(line.advance_rest_pending_balance, Decimal("1.00"))


class ScheduleDeleteViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.site = Site.objects.create(code="007", name="JARDIN.I")
        self.personal_vario, _ = Site.objects.get_or_create(
            code=Site.PERSONAL_VARIO_CODE,
            defaults={
                "name": Site.PERSONAL_VARIO_NAME,
                "admin_only": True,
                "is_active": True,
            },
        )
        self.job_role = JobRole.objects.create(
            code="CARN",
            name="Carnicero",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("9.00"),
        )
        self.schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 7),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        self.line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="123",
            employee_name="Empleado Demo",
            weekly_target_hours=Decimal("4.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
        )
        recalculate_schedule_line(self.line)
        self.line.save()
        self.user = User.objects.create_user(username="operador_delete", password="secret")
        access = UserSiteAccess.objects.get(user=self.user)
        access.sites.add(self.site)
        self.admin_user = User.objects.create_user(username="admin_delete", password="secret")
        admin_access = UserSiteAccess.objects.get(user=self.admin_user)
        admin_access.role = UserSiteAccess.Role.ADMIN
        admin_access.save()

    def build_schedule_form_payload(self, *, status=None, notes="", line=None, extra=None):
        active_line = line or self.line
        payload = {
            "status": status or self.schedule.status or WeeklySchedule.Status.DRAFT,
            "notes": notes,
            "lines-TOTAL_FORMS": "1",
            "lines-INITIAL_FORMS": "1",
            "lines-MIN_NUM_FORMS": "0",
            "lines-MAX_NUM_FORMS": "1000",
            "lines-0-id": str(active_line.pk),
        }
        for index in range(7):
            payload[f"lines-0-day_{index}_shift_1"] = getattr(active_line, f"day_{index}_shift_1", "") or ""
            payload[f"lines-0-day_{index}_shift_2"] = getattr(active_line, f"day_{index}_shift_2", "") or ""
            payload[f"lines-0-day_{index}_compensation_mode"] = getattr(
                active_line,
                f"day_{index}_compensation_mode",
                "",
            ) or ""
            compensation_hours = getattr(active_line, f"day_{index}_compensation_hours", "") or ""
            payload[f"lines-0-day_{index}_compensation_hours"] = str(compensation_hours) if compensation_hours else ""
        if extra:
            payload.update(extra)
        return payload

    def test_site_user_cannot_delete_schedule(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:delete", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 403)
        self.assertTrue(WeeklySchedule.objects.filter(pk=self.schedule.pk).exists())
        self.assertEqual(ScheduleLine.objects.count(), 1)

    def test_site_user_does_not_see_delete_action(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:list"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "plain-action-button")
        self.assertNotContains(response, "Eliminar horario")

    def test_admin_can_delete_schedule(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:delete", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("schedules:list"))
        self.assertFalse(WeeklySchedule.objects.filter(pk=self.schedule.pk).exists())
        self.assertEqual(ScheduleLine.objects.count(), 0)

    def test_admin_delete_rebuilds_future_day_balance_after_rest_payment_removed(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="7001",
            employee_name="Empleado Saldo",
            initial_day_balance=Decimal("4.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="7001",
            employee_name="Empleado Saldo",
            daily_max_hours=Decimal("8.00"),
            day_1_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )
        future_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 14),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        future_line = ScheduleLine.objects.create(
            schedule=future_schedule,
            employee_identifier="7001",
            employee_name="Empleado Saldo",
            daily_max_hours=Decimal("8.00"),
        )
        rebuild_balances_for_employees_from_week(self.schedule.week_start_date, ["7001"])
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("2.00"))

        self.client.login(username="admin_delete", password="secret")
        response = self.client.post(
            reverse("schedules:delete", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("4.00"))
        self.assertFalse(ScheduleLine.objects.filter(pk=line.pk).exists())

    def test_admin_delete_rebuilds_future_day_balance_after_special_day_removed(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="7002",
            employee_name="Empleado Domingo",
            initial_day_balance=Decimal("4.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="7002",
            employee_name="Empleado Domingo",
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
        )
        future_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 14),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        future_line = ScheduleLine.objects.create(
            schedule=future_schedule,
            employee_identifier="7002",
            employee_name="Empleado Domingo",
            daily_max_hours=Decimal("8.00"),
        )
        rebuild_balances_for_employees_from_week(self.schedule.week_start_date, ["7002"])
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("5.00"))

        self.client.login(username="admin_delete", password="secret")
        response = self.client.post(
            reverse("schedules:delete", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("4.00"))
        self.assertFalse(ScheduleLine.objects.filter(pk=line.pk).exists())

    def test_admin_delete_schedule_with_reversal_movements_rebuilds_future_balance(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="7003",
            employee_name="Empleado Con Reversos",
            initial_day_balance=Decimal("0.00"),
        )
        line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="7003",
            employee_name="Empleado Con Reversos",
            daily_max_hours=Decimal("8.00"),
            accrued_day_balance=Decimal("1.00"),
        )
        original_movement = ScheduleBalanceMovement.objects.create(
            schedule=self.schedule,
            line=line,
            site=self.site,
            employee_identifier="7003",
            employee_name="Empleado Con Reversos",
            movement_date=self.schedule.week_start_date,
            movement_type=ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
            quantity_days=Decimal("1.00"),
            balance_before_days=Decimal("0.00"),
            balance_after_days=Decimal("1.00"),
            idempotency_key="delete-reversal-original",
            description="Domingo laborado",
        )
        ScheduleBalanceMovement.objects.create(
            schedule=self.schedule,
            line=line,
            site=self.site,
            employee_identifier="7003",
            employee_name="Empleado Con Reversos",
            movement_date=self.schedule.week_end_date,
            movement_type=ScheduleBalanceMovement.MovementType.REVERSAL,
            quantity_days=Decimal("-1.00"),
            balance_before_days=Decimal("1.00"),
            balance_after_days=Decimal("0.00"),
            idempotency_key="delete-reversal-original:reversal",
            is_reversal=True,
            reversed_movement=original_movement,
            description="Reverso de domingo o festivo laborado",
        )
        future_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 14),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        future_line = ScheduleLine.objects.create(
            schedule=future_schedule,
            employee_identifier="7003",
            employee_name="Empleado Con Reversos",
            daily_max_hours=Decimal("8.00"),
        )
        rebuild_balances_for_employees_from_week(future_schedule.week_start_date, ["7003"])
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("1.00"))

        self.client.login(username="admin_delete", password="secret")
        response = self.client.post(
            reverse("schedules:delete", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(WeeklySchedule.objects.filter(pk=self.schedule.pk).exists())
        self.assertFalse(ScheduleBalanceMovement.objects.filter(schedule=self.schedule).exists())
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("0.00"))

    def test_admin_sees_delete_action(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:list"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "plain-action-button")
        self.assertContains(response, "Excel")
        self.assertContains(response, "Cargar saldos iniciales")

    @patch("schedules.forms.lookup_third_party_by_identifier", return_value=None)
    def test_site_user_can_add_manual_schedule_line(self, _mock_lookup):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "manual_add_submit": "1",
                "manual-lookup_attempts": "2",
                "manual-employee_identifier": "987654321",
                "manual-employee_name": "Trabajador Manual",
                "manual-job_role": str(self.job_role.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("schedules:edit", kwargs={"pk": self.schedule.pk}))
        manual_line = ScheduleLine.objects.get(schedule=self.schedule, employee_identifier="987654321")
        self.assertEqual(manual_line.employee_document_type, "")
        self.assertEqual(manual_line.employee_name, "Trabajador Manual")
        self.assertEqual(manual_line.job_role_name, self.job_role.name)
        self.assertEqual(manual_line.department_name, "")

    def test_schedule_edit_collapses_manual_add_and_hides_area_field(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "manual-add-disclosure", html=False)
        self.assertContains(response, "Agregar persona manualmente")
        self.assertNotContains(response, 'name="manual-department"', html=False)
        self.assertNotContains(response, 'name="manual-employee_document_type"', html=False)
        self.assertContains(response, "manual-add-submit-button", html=False)

    def test_schedule_edit_hides_area_column_and_shows_remove_action(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "<th>Area</th>", html=False)
        self.assertContains(response, "Retirar")

    def test_schedule_edit_shows_role_filter(self):
        ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="456",
            employee_name="Empleado Cargo",
            job_role_name="Carnicero",
        )
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Filtrar por cargo")
        self.assertContains(response, '<option value="Carnicero">Carnicero</option>', html=False)

    def test_schedule_edit_shows_inventory_checkbox(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-inventory-checkbox="true"', html=False)
        self.assertContains(response, "Inventario")

    def test_schedule_edit_persists_inventory_selection_in_database(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.DRAFT,
                notes="Inventario semanal",
                extra={
                    "lines-0-day_1_inventory": "on",
                    "lines-0-manual_day_adjustment": "",
                    "lines-0-manual_hour_adjustment": "",
                },
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.line.refresh_from_db()
        self.assertTrue(self.line.day_1_inventory)
        self.assertFalse(self.line.day_0_inventory)
        self.assertEqual(self.line.inventory_days_total(), 1)
        self.assertIn("08/06/2026", self.line.inventory_days_summary())

    def test_schedule_edit_autosaves_without_redirect(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.DRAFT,
                notes="Autoguardado silencioso",
                extra={"lines-0-day_1_inventory": "on"},
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_X_SCHEDULE_AUTOSAVE="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)
        self.line.refresh_from_db()
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.notes, "Autoguardado silencioso")
        self.assertTrue(self.line.day_1_inventory)

    def test_site_user_can_remove_schedule_line(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "remove_line_id": str(self.line.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("schedules:edit", kwargs={"pk": self.schedule.pk}))
        self.assertFalse(ScheduleLine.objects.filter(pk=self.line.pk).exists())

    @patch(
        "schedules.views.lookup_third_party_by_identifier",
        return_value=LegacyThirdPartyRecord(
            employee_id="555444333",
            employee_name="Carlos Demo",
            role_name="Carnicero",
        ),
    )
    def test_manual_lookup_by_name_prefills_identifier_name_and_role(self, _mock_lookup):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "manual_lookup_submit": "1",
                "manual-lookup_attempts": "0",
                "manual-employee_identifier": "Carlos Demo",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="555444333"', html=False)
        self.assertContains(response, 'value="Carlos Demo"', html=False)
        self.assertContains(response, f'value="{self.job_role.pk}" selected', html=False)

    def test_manual_schedule_line_rejects_duplicate_document(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "manual_add_submit": "1",
                "manual-employee_identifier": self.line.employee_identifier,
                "manual-employee_name": "Otro Nombre",
                "manual-job_role": str(self.job_role.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ya existe en este horario")
        self.assertEqual(ScheduleLine.objects.filter(schedule=self.schedule).count(), 1)

    @patch("schedules.forms.lookup_third_party_by_identifier", return_value=None)
    def test_manual_schedule_line_rejects_blacklisted_document(self, _mock_lookup):
        EmployeeScheduleBlacklist.objects.create(
            employee_identifier="36178712",
            employee_name="Persona Bloqueada",
        )
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "manual_add_submit": "1",
                "manual-lookup_attempts": "2",
                "manual-employee_identifier": "36178712",
                "manual-employee_name": "Persona Bloqueada",
                "manual-job_role": str(self.job_role.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "lista negra")
        self.assertFalse(
            ScheduleLine.objects.filter(schedule=self.schedule, employee_identifier="36178712").exists()
        )

    def test_site_user_does_not_see_night_hours_column(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Rec. noct.")
        self.assertContains(response, 'data-show-night-hours="false"', html=False)

    def test_admin_sees_night_hours_column(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rec. noct.")
        self.assertContains(response, 'data-show-night-hours="true"', html=False)

    def test_schedule_alerts_column_has_live_and_persisted_summaries(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-live-summary', html=False)
        self.assertContains(response, 'aria-live="polite"', html=False)
        self.assertContains(response, 'data-persisted-summary', html=False)

    def test_site_user_sees_compact_alert_summary(self):
        self.line.day_0_shift_1 = ""
        self.line.weekly_target_hours = Decimal("42.00")
        recalculate_schedule_line(self.line)
        self.line.save()
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resultado estimado:")
        self.assertContains(response, "Bloquea revision/publicacion")
        self.assertNotContains(response, "Estado:")

    def test_admin_sees_detailed_alert_summary(self):
        self.line.day_0_shift_1 = ""
        self.line.weekly_target_hours = Decimal("42.00")
        recalculate_schedule_line(self.line)
        self.line.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloquea revision/publicacion")
        self.assertContains(response, "Jornada ajustada: 35.00 h.")

    def test_allowed_one_hour_difference_is_not_rendered_as_warning(self):
        self.line.weekly_target_hours = Decimal("9.00")
        recalculate_schedule_line(self.line)
        self.line.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Estado: Valida con diferencia permitida.")
        self.assertNotContains(response, "Bloquea revision/publicacion")
        self.assertNotContains(response, "alerts-cell--blocking")

    def test_blocking_hour_difference_is_rendered_in_red(self):
        self.line.weekly_target_hours = Decimal("9.50")
        recalculate_schedule_line(self.line)
        self.line.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloquea revision/publicacion")
        self.assertContains(response, "diferencia de 1.5 h")
        self.assertContains(response, "alerts-cell--blocking")

    def test_published_schedule_hides_save_button_and_shows_closed_notice(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "quedo cerrado para edicion")
        self.assertNotContains(response, "Guardar cambios")
        self.assertContains(response, "Habilitar edicion")

    def test_published_schedule_shows_view_action_in_list(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:list"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ver")
        self.assertNotContains(response, "Editar")

    def test_published_schedule_rejects_post_updates(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "status": WeeklySchedule.Status.PUBLISHED,
                "notes": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "1",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-id": str(self.line.pk),
                "lines-0-day_0_shift_1": self.line.day_0_shift_1,
                "lines-0-day_0_shift_2": self.line.day_0_shift_2,
                "lines-0-day_0_compensation_mode": "",
                "lines-0-day_0_compensation_hours": "",
                "lines-0-day_1_shift_1": "",
                "lines-0-day_1_shift_2": "",
                "lines-0-day_1_compensation_mode": "",
                "lines-0-day_1_compensation_hours": "",
                "lines-0-day_2_shift_1": "",
                "lines-0-day_2_shift_2": "",
                "lines-0-day_2_compensation_mode": "",
                "lines-0-day_2_compensation_hours": "",
                "lines-0-day_3_shift_1": "",
                "lines-0-day_3_shift_2": "",
                "lines-0-day_3_compensation_mode": "",
                "lines-0-day_3_compensation_hours": "",
                "lines-0-day_4_shift_1": "",
                "lines-0-day_4_shift_2": "",
                "lines-0-day_4_compensation_mode": "",
                "lines-0-day_4_compensation_hours": "",
                "lines-0-day_5_shift_1": "",
                "lines-0-day_5_shift_2": "",
                "lines-0-day_5_compensation_mode": "",
                "lines-0-day_5_compensation_hours": "",
                "lines-0-day_6_shift_1": "",
                "lines-0-day_6_shift_2": "",
                "lines-0-day_6_compensation_mode": "",
                "lines-0-day_6_compensation_hours": "",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("schedules:edit", kwargs={"pk": self.schedule.pk}))

    def test_published_schedule_rejects_manual_add(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "manual_add_submit": "1",
                "manual-employee_identifier": "111222333",
                "manual-employee_name": "Manual Cerrado",
                "manual-job_role": str(self.job_role.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("schedules:edit", kwargs={"pk": self.schedule.pk}))
        self.assertFalse(
            ScheduleLine.objects.filter(schedule=self.schedule, employee_identifier="111222333").exists()
        )

    def test_published_schedule_rejects_remove_line(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "remove_line_id": str(self.line.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ScheduleLine.objects.filter(pk=self.line.pk).exists())

    def test_admin_can_unlock_published_schedule_and_save_it_closed_again(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        unlock_response = self.client.post(
            reverse("schedules:unlock", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(unlock_response.status_code, 302)
        self.schedule.refresh_from_db()
        self.assertTrue(self.schedule.admin_edit_enabled)
        self.assertFalse(self.schedule.is_closed)

        save_response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "status": WeeklySchedule.Status.PUBLISHED,
                "notes": "Revisado por admin",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "1",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-id": str(self.line.pk),
                "lines-0-day_0_shift_1": self.line.day_0_shift_1,
                "lines-0-day_0_shift_2": self.line.day_0_shift_2,
                "lines-0-day_0_compensation_mode": "",
                "lines-0-day_0_compensation_hours": "",
                "lines-0-day_1_shift_1": "",
                "lines-0-day_1_shift_2": "",
                "lines-0-day_1_compensation_mode": "",
                "lines-0-day_1_compensation_hours": "",
                "lines-0-day_1_inventory": "",
                "lines-0-day_2_shift_1": "",
                "lines-0-day_2_shift_2": "",
                "lines-0-day_2_compensation_mode": "",
                "lines-0-day_2_compensation_hours": "",
                "lines-0-day_2_inventory": "",
                "lines-0-day_3_shift_1": "",
                "lines-0-day_3_shift_2": "",
                "lines-0-day_3_compensation_mode": "",
                "lines-0-day_3_compensation_hours": "",
                "lines-0-day_3_inventory": "",
                "lines-0-day_4_shift_1": "",
                "lines-0-day_4_shift_2": "",
                "lines-0-day_4_compensation_mode": "",
                "lines-0-day_4_compensation_hours": "",
                "lines-0-day_4_inventory": "",
                "lines-0-day_5_shift_1": "",
                "lines-0-day_5_shift_2": "",
                "lines-0-day_5_compensation_mode": "",
                "lines-0-day_5_compensation_hours": "",
                "lines-0-day_5_inventory": "",
                "lines-0-day_6_shift_1": "",
                "lines-0-day_6_shift_2": "",
                "lines-0-day_6_compensation_mode": "",
                "lines-0-day_6_compensation_hours": "",
                "lines-0-day_6_inventory": "",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(save_response.status_code, 302)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.notes, "Revisado por admin")
        self.assertFalse(self.schedule.admin_edit_enabled)
        self.assertTrue(self.schedule.is_closed)

    def test_admin_can_unlock_published_schedule_and_save_line_changes(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier=self.line.employee_identifier,
            employee_name=self.line.employee_name,
            initial_day_balance=Decimal("1.00"),
        )
        self.line.weekly_target_hours = Decimal("42.00")
        self.line.daily_max_hours = Decimal("8.00")
        self.line.save(update_fields=["weekly_target_hours", "daily_max_hours", "updated_at"])
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        unlock_response = self.client.post(
            reverse("schedules:unlock", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(unlock_response.status_code, 302)
        save_response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.PUBLISHED,
                notes="Horario republicado",
                extra={
                    "lines-0-day_0_shift_1": "",
                    "lines-0-day_1_shift_1": "",
                    "lines-0-day_2_shift_1": "descanso",
                    "lines-0-day_3_shift_1": "09:00-16:00",
                    "lines-0-day_4_shift_1": "09:00-16:00",
                    "lines-0-day_5_shift_1": "09:00-16:00",
                    "lines-0-day_6_shift_1": "09:00-16:00",
                },
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(save_response.status_code, 302)
        self.schedule.refresh_from_db()
        self.line.refresh_from_db()
        self.assertEqual(self.schedule.notes, "Horario republicado")
        self.assertFalse(self.schedule.admin_edit_enabled)
        self.assertTrue(self.schedule.is_closed)
        self.assertEqual(self.line.expected_weekly_hours, Decimal("28.00"))
        self.assertEqual(self.line.weekly_hour_difference, Decimal("0.00"))
        self.assertEqual(self.line.accrued_day_balance, Decimal("0.00"))

    def test_admin_can_unlock_published_schedule_and_remove_line(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        unlock_response = self.client.post(
            reverse("schedules:unlock", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(unlock_response.status_code, 302)
        remove_response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            {
                "remove_line_id": str(self.line.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(remove_response.status_code, 302)
        self.assertFalse(ScheduleLine.objects.filter(pk=self.line.pk).exists())

    def test_autosave_keeps_reopened_published_schedule_editable(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.admin_edit_enabled = True
        self.schedule.save()
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.PUBLISHED,
                notes="Autoguardado en horario reabierto",
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_X_SCHEDULE_AUTOSAVE="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.notes, "Autoguardado en horario reabierto")
        self.assertTrue(self.schedule.admin_edit_enabled)
        self.assertFalse(self.schedule.is_closed)

    def test_site_user_cannot_unlock_published_schedule(self):
        self.schedule.status = WeeklySchedule.Status.PUBLISHED
        self.schedule.save()
        self.client.login(username="operador_delete", password="secret")

        response = self.client.post(
            reverse("schedules:unlock", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 403)

    @patch("schedules.forms.lookup_third_party_by_identifier", return_value=None)
    def test_personal_vario_manual_add_requires_blacklisted_identifier(self, _mock_lookup):
        schedule = WeeklySchedule.objects.create(
            site=self.personal_vario,
            week_start_date=date(2026, 6, 14),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": schedule.pk}),
            {
                "manual_add_submit": "1",
                "manual-lookup_attempts": "2",
                "manual-employee_identifier": "999888777",
                "manual-employee_name": "No Bloqueado",
                "manual-job_role": str(self.job_role.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "lista negra")
        self.assertFalse(
            ScheduleLine.objects.filter(schedule=schedule, employee_identifier="999888777").exists()
        )

    def test_schedule_excel_download_returns_spreadsheet(self):
        self.client.login(username="operador_delete", password="secret")

        response = self.client.get(
            reverse("schedules:excel-download", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A6"].value, "Cedula")
        self.assertEqual(worksheet["D6"].value, "Domingo\n07/06/2026")
        self.assertEqual(worksheet["D7"].value, "Turno 1")
        self.assertEqual(worksheet["E7"].value, "Turno 2")
        self.assertEqual(worksheet["F7"].value, "Horas")
        self.assertEqual(worksheet["A8"].value, self.line.employee_identifier)

    def test_settlement_rows_include_employees_with_zero_balance(self):
        self.line.day_0_shift_1 = ""
        recalculate_schedule_line(self.line)
        self.line.save()
        rows = get_settlement_rows(self.schedule)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].employee_identifier, self.line.employee_identifier)
        self.assertEqual(rows[0].accrued_days, Decimal("0.00"))
        self.assertEqual(rows[0].accrued_hours, Decimal("0.00"))

    def test_admin_can_download_schedule_flat_file_template(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.get(
            reverse("schedules:flatfile-template", kwargs={"pk": self.schedule.pk}),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        expected_headers = build_schedule_flat_file_headers(self.schedule)
        self.assertEqual(worksheet["A1"].value, expected_headers[0])
        self.assertEqual(worksheet["D1"].value, expected_headers[3])
        self.assertEqual(worksheet["A2"].value, self.line.employee_identifier)

    def test_admin_can_upload_schedule_flat_file(self):
        self.client.login(username="admin_delete", password="secret")
        workbook = Workbook()
        worksheet = workbook.active
        headers = build_schedule_flat_file_headers(self.schedule)
        worksheet.append(headers)
        row = {header: "" for header in headers}
        row["Cedula"] = self.line.employee_identifier
        row["Empleado"] = self.line.employee_name
        row["Cargo"] = self.line.job_role_name
        row["Domingo turno 1"] = "08:00-16:00"
        row["Domingo inventario"] = "Si"
        row["Lunes turno 1"] = "06:00-10:00"
        row["Lunes turno 2"] = "13:00-17:00"
        row["Ajuste dias"] = 1
        row["Ajuste horas"] = 2
        worksheet.append([row[header] for header in headers])
        upload_buffer = BytesIO()
        workbook.save(upload_buffer)
        upload_buffer.seek(0)
        upload = SimpleUploadedFile(
            "horario_plano.xlsx",
            upload_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("schedules:flatfile-upload", kwargs={"pk": self.schedule.pk}),
            {
                "schedulefile-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.line.refresh_from_db()
        self.assertTrue(self.line.day_0_inventory)
        self.assertEqual(self.line.day_1_shift_1, "06:00-10:00")
        self.assertEqual(self.line.day_1_shift_2, "13:00-17:00")
        self.assertEqual(self.line.manual_day_adjustment, Decimal("1.00"))
        self.assertEqual(self.line.manual_hour_adjustment, Decimal("2.00"))

    def test_admin_can_upload_schedule_flat_file_with_blank_manual_adjustments(self):
        self.client.login(username="admin_delete", password="secret")
        workbook = Workbook()
        worksheet = workbook.active
        headers = build_schedule_flat_file_headers(self.schedule)
        worksheet.append(headers)
        row = {header: "" for header in headers}
        row["Cedula"] = self.line.employee_identifier
        row["Empleado"] = self.line.employee_name
        row["Cargo"] = self.line.job_role_name
        row["Domingo turno 1"] = "08:00-16:00"
        row["Lunes turno 1"] = "06:00-10:00"
        row["Lunes turno 2"] = "13:00-17:00"
        worksheet.append([row[header] for header in headers])
        upload_buffer = BytesIO()
        workbook.save(upload_buffer)
        upload_buffer.seek(0)
        upload = SimpleUploadedFile(
            "horario_plano_sin_ajustes.xlsx",
            upload_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("schedules:flatfile-upload", kwargs={"pk": self.schedule.pk}),
            {
                "schedulefile-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.line.refresh_from_db()
        self.assertEqual(self.line.manual_day_adjustment, Decimal("0.00"))
        self.assertEqual(self.line.manual_hour_adjustment, Decimal("0.00"))

    def test_flat_file_upload_preserves_existing_line_workload_snapshot_after_role_change(self):
        self.line.job_role_code = self.job_role.code
        self.line.job_role_name = self.job_role.name
        self.line.weekly_target_hours = Decimal("44.00")
        self.line.daily_max_hours = Decimal("9.00")
        self.line.base_work_days = 6
        self.line.save(
            update_fields=[
                "job_role_code",
                "job_role_name",
                "weekly_target_hours",
                "daily_max_hours",
                "base_work_days",
                "updated_at",
            ]
        )
        self.job_role.weekly_target_hours = Decimal("42.00")
        self.job_role.save(update_fields=["weekly_target_hours", "updated_at"])

        self.client.login(username="admin_delete", password="secret")
        workbook = Workbook()
        worksheet = workbook.active
        headers = build_schedule_flat_file_headers(self.schedule)
        worksheet.append(headers)
        row = {header: "" for header in headers}
        row["Cedula"] = self.line.employee_identifier
        row["Empleado"] = self.line.employee_name
        row["Cargo"] = self.job_role.name
        row["Domingo turno 1"] = "08:00-16:00"
        worksheet.append([row[header] for header in headers])
        upload_buffer = BytesIO()
        workbook.save(upload_buffer)
        upload_buffer.seek(0)
        upload = SimpleUploadedFile(
            "horario_plano_jornada_congelada.xlsx",
            upload_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("schedules:flatfile-upload", kwargs={"pk": self.schedule.pk}),
            {
                "schedulefile-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.line.refresh_from_db()
        self.assertEqual(self.line.job_role_name, self.job_role.name)
        self.assertEqual(self.line.weekly_target_hours, Decimal("44.00"))
        self.assertEqual(self.line.daily_max_hours, Decimal("9.00"))
        self.assertEqual(self.line.base_work_days, 6)

    @patch("schedules.views.import_schedule_flat_file", side_effect=ProgrammingError("columna faltante"))
    def test_admin_sees_friendly_message_when_flat_file_upload_requires_pending_migrations(self, mocked_import):
        self.client.login(username="admin_delete", password="secret")
        upload = SimpleUploadedFile(
            "horario_plano.xlsx",
            b"dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("schedules:flatfile-upload", kwargs={"pk": self.schedule.pk}),
            {
                "schedulefile-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(mocked_import.called)
        messages_list = [str(message) for message in response.context["messages"]]
        self.assertTrue(
            any("base de datos del portal no esta actualizada" in message.lower() for message in messages_list)
        )

    @patch("schedules.views.import_schedule_flat_file", side_effect=RuntimeError("fallo inesperado"))
    def test_admin_sees_friendly_message_when_flat_file_upload_raises_unexpected_error(self, mocked_import):
        self.client.login(username="admin_delete", password="secret")
        upload = SimpleUploadedFile(
            "horario_plano.xlsx",
            b"dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("schedules:flatfile-upload", kwargs={"pk": self.schedule.pk}),
            {
                "schedulefile-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(mocked_import.called)
        messages_list = [str(message) for message in response.context["messages"]]
        self.assertTrue(
            any("error inesperado" in message.lower() for message in messages_list)
        )

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_review_transition_allows_incomplete_difference_up_to_one_hour(self):
        self.line.weekly_target_hours = Decimal("9.00")
        self.line.save(update_fields=["weekly_target_hours"])
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.REVIEW,
                notes="Pasa a revision con faltante menor",
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.status, WeeklySchedule.Status.REVIEW)
        self.line.refresh_from_db()
        self.assertEqual(self.line.validation_status, ScheduleLine.ValidationStatus.INCOMPLETE)
        self.assertEqual(self.line.weekly_hour_difference, Decimal("-1.00"))

    @patch("schedules.views.generate_and_store_schedule_settlement")
    def test_published_transition_allows_incomplete_difference_up_to_one_hour(self, mocked_settlement):
        self.line.weekly_target_hours = Decimal("9.00")
        self.line.save(update_fields=["weekly_target_hours"])
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.PUBLISHED,
                notes="Publica con faltante menor",
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.status, WeeklySchedule.Status.PUBLISHED)
        self.line.refresh_from_db()
        self.assertEqual(self.line.validation_status, ScheduleLine.ValidationStatus.INCOMPLETE)
        self.assertEqual(self.line.weekly_hour_difference, Decimal("-1.00"))
        self.assertTrue(mocked_settlement.called)

    def test_review_transition_blocks_incomplete_difference_over_one_hour(self):
        self.line.weekly_target_hours = Decimal("9.50")
        self.line.save(update_fields=["weekly_target_hours"])
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.REVIEW,
                notes="Intenta revision con faltante mayor",
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.status, WeeklySchedule.Status.DRAFT)
        messages_list = [str(message) for message in response.context["messages"]]
        self.assertTrue(any("faltante mayor a 1 hora" in message.lower() for message in messages_list))

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_review_transition_sends_hourly_coverage_email(self):
        self.client.login(username="admin_delete", password="secret")

        response = self.client.post(
            reverse("schedules:edit", kwargs={"pk": self.schedule.pk}),
            self.build_schedule_form_payload(
                status=WeeklySchedule.Status.REVIEW,
                notes="Pasa a revision",
            ),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.status, WeeklySchedule.Status.REVIEW)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("Horario en revision", mail.outbox[0].subject)
        self.assertEqual(len(mail.outbox[0].attachments), 1)
        self.assertIn("cobertura_horaria_", mail.outbox[0].attachments[0][0])


class ScheduleLoadViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.site = Site.objects.create(code="007", name="JARDIN.I")
        self.admin_user = User.objects.create_user(username="admin_load", password="secret")
        admin_access = UserSiteAccess.objects.get(user=self.admin_user)
        admin_access.role = UserSiteAccess.Role.ADMIN
        admin_access.save()

    @patch("schedules.services.fetch_active_staff_for_site")
    def test_schedule_load_rebuilds_future_balance_automatically(self, mock_fetch_staff):
        EmployeeInitialBalance.objects.create(
            employee_identifier="7100",
            employee_name="Empleado Creacion",
            initial_day_balance=Decimal("4.00"),
        )
        template_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 19),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        ScheduleLine.objects.create(
            schedule=template_schedule,
            employee_identifier="7100",
            employee_name="Empleado Creacion",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
        )
        future_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 12),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        future_line = ScheduleLine.objects.create(
            schedule=future_schedule,
            employee_identifier="7100",
            employee_name="Empleado Creacion",
            job_role_name="AUXILIAR",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
        )
        rebuild_balances_for_employees_from_week(future_schedule.week_start_date, ["7100"])
        future_line.refresh_from_db()
        self.assertEqual(future_line.accrued_day_balance, Decimal("4.00"))

        mock_fetch_staff.return_value = [
            OperationalStaffingRecord(
                employee_id="7100",
                employee_name="Empleado Creacion",
                site_code=self.site.code,
                department_code="A1",
                department_name="CARNES",
                role_code="AUX",
                role_name="AUXILIAR",
            ),
        ]

        self.client.login(username="admin_load", password="secret")
        response = self.client.post(
            reverse("schedules:load"),
            {
                "site": str(self.site.pk),
                "week_start_date": "2026-07-05",
                "copy_from_schedule": str(template_schedule.pk),
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        created_schedule = WeeklySchedule.objects.get(site=self.site, week_start_date=date(2026, 7, 5))
        created_line = ScheduleLine.objects.get(schedule=created_schedule, employee_identifier="7100")
        future_line.refresh_from_db()

        self.assertEqual(created_line.day_0_shift_1, "08:00-16:00")
        self.assertEqual(created_line.accrued_day_balance, Decimal("5.00"))
        self.assertEqual(future_line.accrued_day_balance, Decimal("5.00"))

    @patch(
        "schedules.views.sync_schedule_from_legacy",
        side_effect=LegacyStaffLookupError("No fue posible consultar el personal desde la base de datos."),
    )
    def test_schedule_load_shows_legacy_error_without_creating_partial_schedule(self, _mock_sync):
        self.client.login(username="admin_load", password="secret")

        response = self.client.post(
            reverse("schedules:load"),
            {
                "site": str(self.site.pk),
                "week_start_date": "2026-07-05",
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No fue posible consultar el personal desde la base de datos.")
        self.assertFalse(
            WeeklySchedule.objects.filter(site=self.site, week_start_date=date(2026, 7, 5)).exists()
        )


class InitialBalanceUploadViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.site = Site.objects.create(code="007", name="JARDIN.I")
        self.schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 5),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        self.line = ScheduleLine.objects.create(
            schedule=self.schedule,
            employee_identifier="999",
            employee_name="Empleado Base",
            weekly_target_hours=Decimal("36.00"),
            daily_max_hours=Decimal("8.00"),
            day_1_shift_1="18:00-00:00",
            day_2_shift_1="18:00-00:00",
            day_3_shift_1="18:00-00:00",
            day_4_shift_1="18:00-00:00",
            day_5_shift_1="18:00-00:00",
            day_6_shift_1="18:00-00:00",
        )
        self.admin_user = User.objects.create_user(username="admin_balances", password="secret")
        admin_access = UserSiteAccess.objects.get(user=self.admin_user)
        admin_access.role = UserSiteAccess.Role.ADMIN
        admin_access.save()
        self.site_user = User.objects.create_user(username="site_balances", password="secret")
        site_access = UserSiteAccess.objects.get(user=self.site_user)
        site_access.sites.add(self.site)

    def build_upload_file(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Cedula", "Nombres y apellidos", "Dias extras", "Horas extras"])
        worksheet.append(["999", "Empleado Base", 1, 2.5])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def build_csv_upload_file(self):
        content = (
            "Cedula,Nombres y apellidos,Dias extras,Horas extras\n"
            "999,Empleado Base,1,2.5\n"
        )
        return SimpleUploadedFile(
            "saldos_iniciales.csv",
            content.encode("utf-8"),
            content_type="text/csv",
        )

    def test_site_user_cannot_access_initial_balance_loader(self):
        self.client.login(username="site_balances", password="secret")

        response = self.client.get(
            reverse("schedules:initial-balances"),
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 403)

    def test_admin_can_upload_initial_balances_and_rebuild_history(self):
        recalculate_schedule_line(self.line)
        self.line.save()
        self.assertEqual(self.line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(self.line.accrued_hour_balance, Decimal("0.00"))

        self.client.login(username="admin_balances", password="secret")
        upload = self.build_upload_file()
        upload.name = "saldos_iniciales.xlsx"

        response = self.client.post(
            reverse("schedules:initial-balances"),
            {
                "balances-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        balance = EmployeeInitialBalance.objects.get(employee_identifier="999")
        self.assertEqual(balance.initial_day_balance, Decimal("1.00"))
        self.assertEqual(balance.initial_hour_balance, Decimal("2.50"))

        self.line.refresh_from_db()
        self.assertEqual(self.line.accrued_day_balance, Decimal("1.00"))
        self.assertEqual(self.line.accrued_hour_balance, Decimal("2.50"))

    def test_admin_can_upload_initial_balances_from_csv(self):
        recalculate_schedule_line(self.line)
        self.line.save()

        self.client.login(username="admin_balances", password="secret")
        upload = self.build_csv_upload_file()

        response = self.client.post(
            reverse("schedules:initial-balances"),
            {
                "balances-file": upload,
            },
            SERVER_NAME="127.0.0.1",
            SERVER_PORT="8000",
        )

        self.assertEqual(response.status_code, 302)
        balance = EmployeeInitialBalance.objects.get(employee_identifier="999")
        self.assertEqual(balance.initial_day_balance, Decimal("1.00"))
        self.assertEqual(balance.initial_hour_balance, Decimal("2.50"))


class InitialBalanceAuditTests(TestCase):
    def setUp(self):
        config = SystemConfiguration.load()
        config.default_weekly_hours = Decimal("46.00")
        config.default_daily_max_hours = Decimal("8.00")
        config.save()

        self.site = Site.objects.create(code="007", name="JARDIN.I")
        self.first_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 6, 28),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        self.second_schedule = WeeklySchedule.objects.create(
            site=self.site,
            week_start_date=date(2026, 7, 5),
            first_day_index=SystemConfiguration.SUNDAY,
        )
        self.first_line = ScheduleLine.objects.create(
            schedule=self.first_schedule,
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            job_role_name="AUXILIAR DE CARNES",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
            day_0_shift_1="08:00-16:00",
            day_1_shift_1="08:00-16:00",
            day_2_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
            day_3_compensation_mode=ScheduleLine.CompensationMode.PAY_DAY,
        )
        self.second_line = ScheduleLine.objects.create(
            schedule=self.second_schedule,
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            job_role_name="AUXILIAR DE CARNES",
            weekly_target_hours=Decimal("46.00"),
            daily_max_hours=Decimal("8.00"),
        )

    def test_initial_balance_update_rebuilds_existing_weeks_and_keeps_audit_in_sync(self):
        balance = EmployeeInitialBalance.objects.create(
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            initial_day_balance=Decimal("9.00"),
            initial_hour_balance=Decimal("0.00"),
        )

        self.first_line.refresh_from_db()
        self.second_line.refresh_from_db()
        self.assertEqual(self.first_line.accrued_day_balance, Decimal("9.00"))
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("9.00"))

        audit_row = build_schedule_balance_audit_rows(["9100"])[0]
        self.assertEqual(audit_row["audited_day_balance"], Decimal("9.00"))
        self.assertEqual(audit_row["stored_day_balance"], Decimal("9.00"))
        self.assertFalse(audit_row["has_difference"])

        balance.initial_day_balance = Decimal("7.00")
        balance.save()

        self.first_line.refresh_from_db()
        self.second_line.refresh_from_db()
        self.assertEqual(self.first_line.accrued_day_balance, Decimal("7.00"))
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("7.00"))

        audit_row = build_schedule_balance_audit_rows(["9100"])[0]
        self.assertEqual(audit_row["audited_day_balance"], Decimal("7.00"))
        self.assertEqual(audit_row["stored_day_balance"], Decimal("7.00"))
        self.assertFalse(audit_row["has_difference"])

    def test_initial_balance_delete_rebuilds_existing_weeks(self):
        balance = EmployeeInitialBalance.objects.create(
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            initial_day_balance=Decimal("9.00"),
        )
        self.first_line.refresh_from_db()
        self.second_line.refresh_from_db()
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("9.00"))

        balance.delete()

        self.first_line.refresh_from_db()
        self.second_line.refresh_from_db()
        self.assertEqual(self.first_line.accrued_day_balance, Decimal("0.00"))
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("0.00"))

        audit_row = build_schedule_balance_audit_rows(["9100"])[0]
        self.assertEqual(audit_row["audited_day_balance"], Decimal("0.00"))
        self.assertEqual(audit_row["stored_day_balance"], Decimal("0.00"))
        self.assertFalse(audit_row["has_difference"])

    @patch("schedules.services.rebuild_balances_for_employees_from_week")
    def test_import_employee_initial_balances_rebuilds_history_once(self, rebuild_mock):
        upload = SimpleUploadedFile(
            "saldos_iniciales.csv",
            "Cedula,Nombres y apellidos,Dias extras,Horas extras\n9100,Empleado Auditoria,9,0\n".encode("utf-8"),
            content_type="text/csv",
        )

        result = import_employee_initial_balances(upload)

        self.assertEqual(result["processed_count"], 1)
        rebuild_mock.assert_called_once()

    def test_audit_command_detects_and_fixes_mismatch(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            initial_day_balance=Decimal("9.00"),
        )
        self.first_line.refresh_from_db()
        self.second_line.refresh_from_db()
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("9.00"))

        EmployeeInitialBalance.objects.filter(employee_identifier="9100").update(
            initial_day_balance=Decimal("8.00")
        )

        output = StringIO()
        call_command("audit_schedule_balances", stdout=output)
        command_output = output.getvalue()
        self.assertIn("9100", command_output)
        self.assertIn("dif_dias=1.00", command_output)

        fixed_output = StringIO()
        call_command("audit_schedule_balances", "--fix", stdout=fixed_output)
        self.second_line.refresh_from_db()
        self.assertEqual(self.second_line.accrued_day_balance, Decimal("8.00"))

        audit_row = build_schedule_balance_audit_rows(["9100"])[0]
        self.assertFalse(audit_row["has_difference"])
        self.assertIn("Recalculo aplicado correctamente", fixed_output.getvalue())

    def test_audit_ignores_reversed_legacy_movements(self):
        EmployeeInitialBalance.objects.create(
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            initial_day_balance=Decimal("1.00"),
        )
        self.second_line.accrued_day_balance = Decimal("-1.00")
        self.second_line.save(update_fields=["accrued_day_balance"])
        ScheduleBalanceMovement.objects.create(
            schedule=self.second_schedule,
            line=self.second_line,
            site=self.site,
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            job_role_name="AUXILIAR DE CARNES",
            movement_date=date(2026, 7, 5),
            movement_type=ScheduleBalanceMovement.MovementType.PAY_DAY,
            quantity_days=Decimal("-2.00"),
            equivalent_hours=Decimal("-16.00"),
            balance_before_days=Decimal("1.00"),
            balance_after_days=Decimal("-1.00"),
            idempotency_key="activo-pay-day",
            description="Pago con descanso",
        )
        ScheduleBalanceMovement.objects.create(
            schedule=self.first_schedule,
            line=self.first_line,
            site=self.site,
            employee_identifier="9100",
            employee_name="Empleado Auditoria",
            job_role_name="AUXILIAR DE CARNES",
            movement_date=date(2026, 6, 28),
            movement_type=ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
            quantity_days=Decimal("1.00"),
            equivalent_hours=Decimal("8.00"),
            balance_before_days=Decimal("0.00"),
            balance_after_days=Decimal("1.00"),
            movement_origin="legacy",
            is_reversed=True,
            description="Movimiento historico revertido",
        )

        audit_row = build_schedule_balance_audit_rows(["9100"])[0]

        self.assertEqual(audit_row["audited_day_balance"], Decimal("-1.00"))
        self.assertEqual(audit_row["stored_day_balance"], Decimal("-1.00"))
        self.assertFalse(audit_row["has_difference"])
