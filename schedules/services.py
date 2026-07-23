from __future__ import annotations

import csv
import hashlib
from contextlib import contextmanager
from contextvars import ContextVar
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_FLOOR, ROUND_HALF_UP
from io import StringIO
import re
import unicodedata

from core.models import (
    DEFAULT_PROGRAMMING_INTERVAL_MINUTES,
    JobRole,
    OperationalStaffCache,
    ShiftTemplate,
    Site,
    SystemConfiguration,
)
from django.db import transaction
from django.db.models import DecimalField, Sum, Value
from django.db.models.functions import Coalesce
from django.db.utils import OperationalError, ProgrammingError
from legacy.services import OperationalStaffingRecord, fetch_active_staff_for_site
from openpyxl import load_workbook
from schedules.calendar_utils import get_special_day_label
from schedules.models import (
    EmployeeInitialBalance,
    EmployeeScheduleBlacklist,
    EmployeeOvertimeRestriction,
    ScheduleBalanceMovement,
    ScheduleLine,
    WeeklySchedule,
)

NON_WORKED_SHIFT_LABELS = {
    "",
    "-",
    "contratacion",
    "descanso",
    "festivo",
    "incapacidad",
    "inasistencia",
    "prestamo",
    "traslado",
    "vacaciones",
    "volante",
    "volantes",
    "renuncia",
    "licencia",
}
REST_SHIFT_LABEL = "descanso"
LOAN_SHIFT_LABELS = {"prestamo"}
SHIFT_PATTERN = re.compile(r"^(?P<start>\d{1,2}:\d{2})-(?P<end>\d{1,2}:\d{2})$")
TWO_DECIMALS = Decimal("0.01")

MONEY_HOUR_COMPENSATION_MODES = {
    ScheduleLine.CompensationMode.PAY_MONEY,
    ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
}
MONEY_DAY_COMPENSATION_MODES = {
    ScheduleLine.CompensationMode.PAY_MONEY_DAY,
}
ADVANCE_DAY_COMPENSATION_MODES = {
    ScheduleLine.CompensationMode.ADVANCE_DAY,
}
COMPANY_DAY_REPAYMENT_MODE = "repay_company_day"
COMPANY_DAY_REPAYMENT_LABEL = "Compensa dia empresa"
COMPANY_DAY_REPAYMENT_MOVEMENT = "company_day_repayment"
COMPANY_DAY_REPAYMENT_AUTO_DESCRIPTION = "Compensacion automatica de dia a favor de la empresa"
COMPANY_DAY_REPAYMENT_MANUAL_DESCRIPTION = "Compensacion de dia a favor de la empresa"
COMPANY_DAY_REPAYMENT_MOVEMENT_LABEL = "Dia a favor de la empresa compensado"
MAX_ADVANCE_REST_PENDING_DAYS = Decimal("2.00")
ALLOWED_INCOMPLETE_STATUS_DIFFERENCE_HOURS = Decimal("1.00")
ADVANCE_REST_LIMIT_ERROR_MESSAGE = (
    f"El trabajador ya alcanzo el limite maximo de {int(MAX_ADVANCE_REST_PENDING_DAYS)} dias adelantados "
    "a favor de la empresa."
)
_INITIAL_BALANCE_REBUILD_SUPPRESSION = ContextVar("initial_balance_rebuild_suppression", default=0)

REST_SHIFT_ALIASES = {
    REST_SHIFT_LABEL,
}
LEAVE_SHIFT_LABELS = {
    "contratacion",
    "festivo",
    "incapacidad",
    "traslado",
    "vacaciones",
    "volante",
    "volantes",
    "renuncia",
    "licencia",
}
ABSENCE_SHIFT_LABELS = {"inasistencia"}

SCHEDULE_BALANCE_MOVEMENT_LABELS = {
    ScheduleBalanceMovement.MovementType.OVERTIME: "Hora extra generada",
    ScheduleBalanceMovement.MovementType.HOUR_DEFICIT: "Horas faltantes",
    ScheduleBalanceMovement.MovementType.SPECIAL_DAY: "Domingo o festivo laborado",
    ScheduleBalanceMovement.MovementType.ADDITIONAL_REST: "Descanso adicional disfrutado",
    ScheduleBalanceMovement.MovementType.MANUAL_DAY: "Ajuste manual de dias",
    ScheduleBalanceMovement.MovementType.MANUAL_HOUR: "Ajuste manual de horas",
    ScheduleBalanceMovement.MovementType.PAY_DAY: "Pago con descanso",
    ScheduleBalanceMovement.MovementType.ADVANCE_DAY: "Descanso adelantado",
    ScheduleBalanceMovement.MovementType.PAY_HOURS: "Pago con horas",
    ScheduleBalanceMovement.MovementType.PAY_MONEY_DAY: "Pago en dinero por dia",
    ScheduleBalanceMovement.MovementType.PAY_MONEY_HOURS: "Pago en dinero por horas",
    ScheduleBalanceMovement.MovementType.PAY_MONEY: "Pago en dinero",
    ScheduleBalanceMovement.MovementType.REVERSAL: "Reverso",
    COMPANY_DAY_REPAYMENT_MOVEMENT: COMPANY_DAY_REPAYMENT_MOVEMENT_LABEL,
}


@contextmanager
def suppress_initial_balance_rebuild():
    token = _INITIAL_BALANCE_REBUILD_SUPPRESSION.set(_INITIAL_BALANCE_REBUILD_SUPPRESSION.get() + 1)
    try:
        yield
    finally:
        _INITIAL_BALANCE_REBUILD_SUPPRESSION.reset(token)


def initial_balance_rebuild_is_suppressed() -> bool:
    return _INITIAL_BALANCE_REBUILD_SUPPRESSION.get() > 0


def normalize_shift_label(value: str) -> str:
    return (value or "").strip()


def normalize_shift_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return normalized.encode("ascii", "ignore").decode("ascii").strip().casefold()


def decimal_hours(value: float | Decimal | int | str) -> Decimal:
    return Decimal(str(value or "0")).quantize(TWO_DECIMALS)


def decimal_hours_to_exact_minutes(value: float | Decimal | int | str) -> Decimal:
    return Decimal(str(value or "0")) * Decimal("60")


def minutes_to_decimal_hours(value: float | Decimal | int | str) -> Decimal:
    return (Decimal(str(value or "0")) / Decimal("60")).quantize(TWO_DECIMALS)


def format_minutes_duration(value: float | Decimal | int | str) -> str:
    normalized = Decimal(str(value or "0"))
    sign = "-" if normalized < Decimal("0") else ""
    absolute_minutes = abs(normalized).to_integral_value(rounding=ROUND_HALF_UP)
    hours, minutes = divmod(int(absolute_minutes), 60)
    return f"{sign}{hours}:{minutes:02d}"


def get_programming_interval_minutes(config: SystemConfiguration | None = None) -> int:
    config = config or SystemConfiguration.load()
    configured_value = (
        getattr(config, "programming_interval_minutes", None)
        or DEFAULT_PROGRAMMING_INTERVAL_MINUTES
    )
    try:
        normalized_value = int(configured_value)
    except (TypeError, ValueError):
        normalized_value = DEFAULT_PROGRAMMING_INTERVAL_MINUTES
    return max(1, normalized_value)


def round_minutes_to_interval(
    minutes: float | Decimal | int | str,
    interval_minutes: int | Decimal,
) -> Decimal:
    normalized_minutes = max(Decimal(str(minutes or "0")), Decimal("0"))
    normalized_interval = max(Decimal(str(interval_minutes or "0")), Decimal("1"))
    quotient = (normalized_minutes / normalized_interval).to_integral_value(rounding=ROUND_FLOOR)
    remainder = normalized_minutes - (quotient * normalized_interval)
    if remainder * Decimal("2") >= normalized_interval:
        quotient += 1
    return quotient * normalized_interval


def get_line_base_work_days(
    line: ScheduleLine,
    config: SystemConfiguration | None = None,
) -> int:
    config = config or SystemConfiguration.load()
    configured_value = getattr(line, "base_work_days", None) or getattr(config, "default_base_work_days", 6) or 6
    try:
        normalized_value = int(configured_value)
    except (TypeError, ValueError):
        normalized_value = 6
    return max(1, min(normalized_value, 7))


def get_line_day_reference_minutes_exact(
    line: ScheduleLine,
    config: SystemConfiguration | None = None,
) -> Decimal:
    config = config or SystemConfiguration.load()
    weekly_target_hours = decimal_hours(line.weekly_target_hours or config.default_weekly_hours or "0")
    base_work_days = Decimal(str(get_line_base_work_days(line, config=config)))
    if weekly_target_hours > Decimal("0.00"):
        return decimal_hours_to_exact_minutes(weekly_target_hours) / base_work_days
    return decimal_hours_to_exact_minutes(line.daily_max_hours or config.default_daily_max_hours or Decimal("0.00"))


def format_signed_day_balance(balance: Decimal) -> str:
    normalized = decimal_hours(balance)
    rendered = format(abs(normalized), "f").rstrip("0").rstrip(".") or "0"
    if normalized > Decimal("0.00"):
        return f"{rendered} dia(s) a favor del trabajador"
    if normalized < Decimal("0.00"):
        return f"{rendered} dia(s) a favor de la empresa"
    return "Sin dias pendientes"


def get_schedule_balance_movement_label(movement_type: str) -> str:
    return SCHEDULE_BALANCE_MOVEMENT_LABELS.get(str(movement_type or "").strip(), str(movement_type or "").strip())


def format_weekly_difference_compact(difference_hours: Decimal) -> str:
    normalized = decimal_hours(difference_hours)
    if normalized > Decimal("0.00"):
        return f"{normalized} h de excedente"
    if normalized < Decimal("0.00"):
        return f"{abs(normalized)} h pendientes"
    return "sin diferencia horaria"


def format_hours_for_message(value: Decimal) -> str:
    rendered = format(decimal_hours(value).normalize(), "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def is_non_blocking_hour_difference(status: str, difference_hours: Decimal) -> bool:
    if status not in {
        ScheduleLine.ValidationStatus.INCOMPLETE,
        ScheduleLine.ValidationStatus.OVERPLANNED,
    }:
        return False
    return abs(decimal_hours(difference_hours)) <= ALLOWED_INCOMPLETE_STATUS_DIFFERENCE_HOURS


def schedule_line_blocks_status_transition(line: ScheduleLine) -> bool:
    if line.validation_status == ScheduleLine.ValidationStatus.INCOMPLETE:
        return abs(decimal_hours(line.weekly_hour_difference)) > ALLOWED_INCOMPLETE_STATUS_DIFFERENCE_HOURS
    return line.validation_status in {
        ScheduleLine.ValidationStatus.IMPOSSIBLE,
        ScheduleLine.ValidationStatus.INCONSISTENT,
    }


def get_schedule_line_status_blocker_message(line: ScheduleLine) -> str:
    difference_hours = decimal_hours(line.weekly_hour_difference)
    if (
        line.validation_status == ScheduleLine.ValidationStatus.INCOMPLETE
        and abs(difference_hours) > ALLOWED_INCOMPLETE_STATUS_DIFFERENCE_HOURS
    ):
        return (
            "Bloquea revision/publicacion: tiene una diferencia de "
            f"{format_hours_for_message(abs(difference_hours))} h. "
            "Modifica la programacion o agrega horas pagas y guarda."
        )
    if line.validation_status == ScheduleLine.ValidationStatus.IMPOSSIBLE:
        return (
            "Bloquea revision/publicacion: la jornada ajustada no cabe en la capacidad disponible. "
            "Revisa turnos, novedades o parametrizacion del cargo y guarda."
        )
    if line.validation_status == ScheduleLine.ValidationStatus.INCONSISTENT:
        return (
            "Bloquea revision/publicacion: hay una configuracion inconsistente. "
            "Revisa pagos, descansos, saldos o turnos y guarda."
        )
    return ""


def summarize_schedule_day_movements(
    day_breakdown: list[dict[str, Decimal | date | str | int]],
    day_states: dict[int, dict[str, Decimal | str | bool]],
) -> dict[str, int]:
    summary = {
        "generated_sunday_days": 0,
        "generated_holiday_days": 0,
        "paid_days": 0,
        "advance_days": 0,
        "additional_rest_days": 0,
        "absence_days": 0,
        "company_day_repayments": 0,
    }
    for day_info in day_breakdown:
        index = int(day_info["index"])
        day_date = day_info["date"]
        special_label = str(day_info.get("special_label") or "")
        day_state = day_states.get(index, {})
        if bool(day_state.get("generated_day")) and special_label:
            if getattr(day_date, "weekday", lambda: -1)() == 6:
                summary["generated_sunday_days"] += 1
            else:
                summary["generated_holiday_days"] += 1

        movement_type = str(day_state.get("day_movement_type") or "")
        applied_day_delta = decimal_hours(day_state.get("applied_day_delta", Decimal("0.00")) or "0")
        if movement_type == ScheduleBalanceMovement.MovementType.PAY_DAY and applied_day_delta < Decimal("0.00"):
            summary["paid_days"] += 1
        elif movement_type == ScheduleBalanceMovement.MovementType.ADVANCE_DAY and applied_day_delta < Decimal("0.00"):
            summary["advance_days"] += 1
        elif movement_type == ScheduleBalanceMovement.MovementType.ADDITIONAL_REST and applied_day_delta < Decimal("0.00"):
            if str(day_state.get("day_movement_description") or "").strip().casefold() == "inasistencia":
                summary["absence_days"] += 1
            else:
                summary["additional_rest_days"] += 1
        elif movement_type == COMPANY_DAY_REPAYMENT_MOVEMENT and applied_day_delta > Decimal("0.00"):
            summary["company_day_repayments"] += 1
    return summary


def normalize_full_day_threshold(
    worked_hours: Decimal,
    reference_hours: Decimal,
) -> Decimal:
    normalized_worked_hours = decimal_hours(worked_hours)
    normalized_reference_hours = decimal_hours(reference_hours)
    if normalized_reference_hours <= Decimal("0.00"):
        return normalized_worked_hours
    return max(normalized_worked_hours, normalized_reference_hours)


def get_complete_work_day_threshold_hours(
    entry: dict[str, Decimal | int | str | bool],
    *,
    day_reference_hours: Decimal,
) -> Decimal:
    expected_hours = decimal_hours(entry.get("expected_hours", Decimal("0.00")) or "0")
    configured_day_hours = decimal_hours(entry.get("daily_target_hours", Decimal("0.00")) or "0")
    reference_hours = max(expected_hours, decimal_hours(day_reference_hours))
    if configured_day_hours > Decimal("0.00"):
        reference_hours = min(reference_hours if reference_hours > Decimal("0.00") else configured_day_hours, configured_day_hours)
    if reference_hours <= Decimal("0.00"):
        return Decimal("0.00")

    reference_minutes = decimal_hours_to_exact_minutes(reference_hours)
    whole_hours = (reference_minutes / Decimal("60")).to_integral_value(rounding=ROUND_FLOOR)
    fractional_minutes = reference_minutes - (whole_hours * Decimal("60"))
    if fractional_minutes == Decimal("30"):
        threshold_minutes = reference_minutes
    elif fractional_minutes == Decimal("0"):
        threshold_minutes = whole_hours * Decimal("60")
    else:
        threshold_minutes = whole_hours * Decimal("60")

    if threshold_minutes <= Decimal("0"):
        threshold_minutes = reference_minutes
    return minutes_to_decimal_hours(threshold_minutes).quantize(TWO_DECIMALS)


def is_complete_work_day_entry(
    entry: dict[str, Decimal | int | str | bool],
    *,
    day_reference_hours: Decimal,
) -> bool:
    worked_hours = decimal_hours(entry.get("worked_hours", Decimal("0.00")) or "0")
    if worked_hours <= Decimal("0.00"):
        return False
    if str(entry.get("expected_reason") or "") == "fuera_de_rango":
        return False
    if bool(entry.get("is_leave_day")) or bool(entry.get("is_non_worked_holiday")):
        return False
    if str(entry.get("mode") or "") in {
        ScheduleLine.CompensationMode.PAY_DAY,
        *ADVANCE_DAY_COMPENSATION_MODES,
    }:
        return False

    full_day_threshold = get_complete_work_day_threshold_hours(
        entry,
        day_reference_hours=day_reference_hours,
    )
    if full_day_threshold <= Decimal("0.00"):
        return False
    return worked_hours >= full_day_threshold


def get_company_day_repayment_exclusion_hours(
    entry: dict[str, Decimal | int | str | bool],
    *,
    day_reference_hours: Decimal,
) -> Decimal:
    worked_hours = decimal_hours(entry.get("worked_hours", Decimal("0.00")) or "0")
    if worked_hours <= Decimal("0.00"):
        return Decimal("0.00")

    expected_hours = decimal_hours(entry.get("expected_hours", Decimal("0.00")) or "0")
    configured_day_hours = decimal_hours(entry.get("daily_target_hours", Decimal("0.00")) or "0")
    full_day_hours = max(expected_hours, configured_day_hours, decimal_hours(day_reference_hours))
    if full_day_hours <= Decimal("0.00"):
        return worked_hours
    return min(worked_hours, full_day_hours).quantize(TWO_DECIMALS)


def get_company_day_repayment_auto_priority(
    entry: dict[str, Decimal | int | str | bool],
) -> tuple[int, int]:
    index = int(entry.get("index", 0))
    if bool(entry.get("is_mandatory_rest_day", False)):
        return (0, index)
    if index == 0:
        return (1, index)
    if bool(entry.get("special_generated", False)) or bool(entry.get("is_holiday", False)):
        return (2, index)
    return (3, index)


def build_company_day_repayment_plan(
    compensation_entries: list[dict[str, Decimal | int | str | bool]],
    *,
    starting_day_balance: Decimal,
    day_reference_hours: Decimal,
    base_work_days: int,
) -> dict[str, object]:
    normalized_starting_balance = decimal_hours(starting_day_balance)
    pending_company_days = int(
        abs(min(normalized_starting_balance, Decimal("0.00"))).to_integral_value(rounding=ROUND_FLOOR)
    )
    full_work_day_indexes = [
        int(entry.get("index", 0))
        for entry in compensation_entries
        if is_complete_work_day_entry(entry, day_reference_hours=day_reference_hours)
    ]
    additional_full_work_days = max(len(full_work_day_indexes) - max(int(base_work_days or 0), 0), 0)
    repayment_capacity = min(additional_full_work_days, pending_company_days)
    marked_repayment_indexes = [
        int(entry.get("index", 0))
        for entry in compensation_entries
        if str(entry.get("mode") or "") == COMPANY_DAY_REPAYMENT_MODE
    ]

    valid_repayment_indexes: list[int] = []
    automatic_repayment_indexes: list[int] = []
    invalid_repayment_reasons: dict[int, str] = {}
    excluded_hours_by_index: dict[int, Decimal] = {}
    candidate_entries: list[dict[str, Decimal | int | str | bool]] = []

    has_manual_selection = bool(marked_repayment_indexes)
    for entry in compensation_entries:
        index = int(entry.get("index", 0))
        is_manual_selection = str(entry.get("mode") or "") == COMPANY_DAY_REPAYMENT_MODE
        if has_manual_selection and not is_manual_selection:
            continue
        if pending_company_days <= 0:
            if is_manual_selection:
                invalid_repayment_reasons[index] = "no_negative_balance"
            continue
        if not is_complete_work_day_entry(entry, day_reference_hours=day_reference_hours):
            if is_manual_selection:
                invalid_repayment_reasons[index] = "partial_day"
            continue
        candidate_entries.append(entry)

    if has_manual_selection:
        selected_entries = sorted(candidate_entries, key=lambda item: int(item.get("index", 0)))
    else:
        selected_entries = sorted(candidate_entries, key=get_company_day_repayment_auto_priority)

    for position, entry in enumerate(selected_entries):
        index = int(entry.get("index", 0))
        if position >= repayment_capacity:
            if has_manual_selection:
                invalid_repayment_reasons[index] = "no_additional_full_day"
            continue
        valid_repayment_indexes.append(index)
        if not has_manual_selection:
            automatic_repayment_indexes.append(index)
        excluded_hours_by_index[index] = get_company_day_repayment_exclusion_hours(
            entry,
            day_reference_hours=day_reference_hours,
        )

    return {
        "starting_day_balance": normalized_starting_balance.quantize(TWO_DECIMALS),
        "pending_company_days": pending_company_days,
        "full_work_day_indexes": full_work_day_indexes,
        "complete_work_days": len(full_work_day_indexes),
        "additional_full_work_days": additional_full_work_days,
        "repayment_capacity": repayment_capacity,
        "marked_repayment_indexes": marked_repayment_indexes,
        "valid_repayment_indexes": valid_repayment_indexes,
        "automatic_repayment_indexes": automatic_repayment_indexes,
        "invalid_repayment_reasons": invalid_repayment_reasons,
        "excluded_hours_by_index": excluded_hours_by_index,
    }


def get_company_day_repayment_error_message(reason: str) -> str:
    normalized_reason = str(reason or "").strip()
    if normalized_reason == "no_negative_balance":
        return "Solo puedes compensar un dia a favor de la empresa si el trabajador tiene saldo negativo previo."
    if normalized_reason == "partial_day":
        return (
            "La jornada no puede utilizarse para compensar el descanso adelantado porque no corresponde a un dia completo."
        )
    if normalized_reason == "no_additional_full_day":
        return (
            "La compensacion requiere un dia completo adicional respecto de los dias base del cargo."
        )
    return "No fue posible aplicar la compensacion del dia a favor de la empresa."


def compute_night_hours(start_time, end_time, night_shift_start) -> Decimal:
    if not start_time or not end_time:
        return Decimal("0.00")

    anchor_date = date.today()
    start_value = datetime.combine(anchor_date, start_time)
    end_value = datetime.combine(anchor_date, end_time)
    if end_value <= start_value:
        end_value += timedelta(days=1)

    night_start = datetime.combine(anchor_date, night_shift_start)
    if start_value >= night_start:
        return decimal_hours((end_value - start_value).total_seconds() / 3600)
    if end_value <= night_start:
        return Decimal("0.00")
    return decimal_hours((end_value - night_start).total_seconds() / 3600)


def resolve_shift_metrics(
    label: str,
    config: SystemConfiguration | None = None,
    shift_templates: dict[str, ShiftTemplate] | None = None,
) -> tuple[Decimal, Decimal]:
    normalized = normalize_shift_label(label)
    if normalize_shift_key(normalized) in NON_WORKED_SHIFT_LABELS:
        return Decimal("0.00"), Decimal("0.00")

    config = config or SystemConfiguration.load()
    template = None
    if shift_templates is not None:
        template = shift_templates.get(normalized)
    elif normalized:
        template = ShiftTemplate.objects.filter(label=normalized, is_active=True).first()

    if template:
        if not template.counts_as_worked_time:
            return Decimal("0.00"), Decimal("0.00")
        total_hours = decimal_hours(template.duration_hours)
        computed_night_hours = compute_night_hours(
            template.start_time,
            template.end_time,
            config.night_shift_start,
        )
        return total_hours, computed_night_hours

    match = SHIFT_PATTERN.match(normalized)
    if not match:
        return Decimal("0.00"), Decimal("0.00")

    start_time = datetime.strptime(match.group("start"), "%H:%M").time()
    end_time = datetime.strptime(match.group("end"), "%H:%M").time()
    start_value = datetime.combine(date.today(), start_time)
    end_value = datetime.combine(date.today(), end_time)
    if end_value <= start_value:
        end_value += timedelta(days=1)
    total_seconds = (end_value - start_value).total_seconds()
    total_hours = decimal_hours(total_seconds / 3600)
    night_hours = compute_night_hours(start_time, end_time, config.night_shift_start)
    return total_hours, night_hours


def parse_shift_hours(label: str) -> Decimal:
    total_hours, _ = resolve_shift_metrics(label)
    return total_hours


def parse_shift_night_hours(label: str) -> Decimal:
    _, night_hours = resolve_shift_metrics(label)
    return night_hours


def build_shift_metrics_catalog(config: SystemConfiguration | None = None) -> dict[str, dict[str, str | bool]]:
    config = config or SystemConfiguration.load()
    templates = {
        shift.label: shift
        for shift in ShiftTemplate.objects.filter(is_active=True)
    }
    catalog: dict[str, dict[str, str | bool]] = {}
    for label, template in templates.items():
        total_hours, night_hours = resolve_shift_metrics(label, config=config, shift_templates=templates)
        catalog[label] = {
            "hours": str(total_hours),
            "night_hours": str(night_hours),
            "counts_as_worked_time": template.counts_as_worked_time,
        }
    return catalog


def get_rest_shift_label() -> str:
    template = ShiftTemplate.objects.filter(label__iexact=REST_SHIFT_LABEL, is_active=True).first()
    return template.label if template else REST_SHIFT_LABEL


def get_line_day_reference_hours(
    line: ScheduleLine,
    config: SystemConfiguration | None = None,
) -> Decimal:
    return minutes_to_decimal_hours(get_line_day_reference_minutes_exact(line, config=config))


def get_shift_non_work_category(
    label: str,
    *,
    shift_templates: dict[str, ShiftTemplate] | None = None,
) -> str:
    normalized = normalize_shift_label(label)
    if not normalized:
        return ""

    lowered = normalize_shift_key(normalized)
    if lowered in REST_SHIFT_ALIASES:
        return "rest"
    if lowered in LOAN_SHIFT_LABELS:
        return "loan"
    if lowered in ABSENCE_SHIFT_LABELS:
        return "absence"
    if lowered in LEAVE_SHIFT_LABELS:
        return "leave"

    template = None
    if shift_templates is not None:
        template = shift_templates.get(normalized)
    if template is not None and not template.counts_as_worked_time:
        if lowered in REST_SHIFT_ALIASES:
            return "rest"
        if lowered in LOAN_SHIFT_LABELS:
            return "loan"
        if lowered in ABSENCE_SHIFT_LABELS:
            return "absence"
        return "leave"

    return ""


def get_active_overtime_restriction(employee_identifier: str) -> EmployeeOvertimeRestriction | None:
    cleaned_identifier = (employee_identifier or "").strip()
    if not cleaned_identifier:
        return None
    try:
        return (
            EmployeeOvertimeRestriction.objects.filter(
                employee_identifier=cleaned_identifier,
                is_active=True,
            )
            .only(
                "employee_identifier",
                "employee_name",
                "max_daily_overtime_hours",
                "max_weekly_overtime_hours",
            )
            .first()
        )
    except (ProgrammingError, OperationalError):
        return None


def get_blacklisted_employee_identifiers(
    employee_identifiers: list[str] | set[str] | tuple[str, ...] | None = None,
) -> set[str]:
    cleaned_identifiers = {
        (employee_identifier or "").strip()
        for employee_identifier in (employee_identifiers or [])
        if (employee_identifier or "").strip()
    }
    if employee_identifiers is not None and not cleaned_identifiers:
        return set()

    try:
        queryset = EmployeeScheduleBlacklist.objects.filter(is_active=True)
        if cleaned_identifiers:
            queryset = queryset.filter(employee_identifier__in=cleaned_identifiers)
        return set(queryset.values_list("employee_identifier", flat=True))
    except (ProgrammingError, OperationalError):
        return set()


def is_employee_blacklisted(employee_identifier: str) -> bool:
    cleaned_identifier = (employee_identifier or "").strip()
    if not cleaned_identifier:
        return False
    return cleaned_identifier in get_blacklisted_employee_identifiers([cleaned_identifier])


def schedule_accepts_blacklisted_staff(schedule: WeeklySchedule | None) -> bool:
    site = getattr(schedule, "site", None)
    return bool(site and site.is_personal_vario)


def build_personal_vario_staff_records() -> list[OperationalStaffingRecord]:
    blacklist_entries = list(
        EmployeeScheduleBlacklist.objects.filter(is_active=True).order_by("employee_name", "employee_identifier")
    )
    if not blacklist_entries:
        return []

    identifiers = [entry.employee_identifier for entry in blacklist_entries if (entry.employee_identifier or "").strip()]
    cache_rows = list(
        OperationalStaffCache.objects.filter(employee_identifier__in=identifiers, is_active=True).order_by(
            "employee_identifier",
            "role_name",
            "employee_name",
        )
    )
    cache_map: dict[str, OperationalStaffCache] = {}
    for row in cache_rows:
        cache_map.setdefault((row.employee_identifier or "").strip(), row)

    records = []
    for entry in blacklist_entries:
        employee_identifier = (entry.employee_identifier or "").strip()
        if not employee_identifier:
            continue
        cached_row = cache_map.get(employee_identifier)
        employee_name = (
            (entry.employee_name or "").strip()
            or ((cached_row.employee_name or "").strip() if cached_row else "")
            or employee_identifier
        )
        role_code = ((cached_row.role_code or "").strip() if cached_row else "")
        role_name = ((cached_row.role_name or "").strip() if cached_row else "") or "PERSONAL VARIO"
        department_code = ((cached_row.department_code or "").strip() if cached_row else "")
        department_name = ((cached_row.department_name or "").strip() if cached_row else "")
        records.append(
            OperationalStaffingRecord(
                employee_id=employee_identifier,
                employee_name=employee_name,
                site_code=Site.PERSONAL_VARIO_CODE,
                department_code=department_code,
                department_name=department_name,
                role_code=role_code,
                role_name=role_name,
            )
        )

    return sorted(
        records,
        key=lambda item: (
            (item.role_name or "").casefold(),
            (item.employee_name or "").casefold(),
            (item.employee_id or "").casefold(),
        ),
    )


def get_daily_overtime_hours(worked_hours: Decimal, day_reference_hours: Decimal) -> Decimal:
    return max(
        decimal_hours(worked_hours) - decimal_hours(day_reference_hours),
        Decimal("0.00"),
    ).quantize(TWO_DECIMALS)


def get_selected_shift_templates(line: ScheduleLine) -> dict[str, ShiftTemplate]:
    selected_labels = {
        normalize_shift_label(getattr(line, f"day_{index}_shift_{slot}"))
        for index in range(7)
        for slot in (1, 2)
        if normalize_shift_label(getattr(line, f"day_{index}_shift_{slot}"))
    }
    return {
        shift.label: shift
        for shift in ShiftTemplate.objects.filter(label__in=selected_labels, is_active=True)
    }


def build_line_day_breakdown(
    line: ScheduleLine,
    config: SystemConfiguration | None = None,
    shift_templates: dict[str, ShiftTemplate] | None = None,
) -> list[dict[str, Decimal | date | str | int]]:
    schedule = getattr(line, "schedule", None)
    if schedule is None or not schedule.week_start_date:
        return []

    config = config or SystemConfiguration.load()
    shift_templates = shift_templates or get_selected_shift_templates(line)
    breakdown: list[dict[str, Decimal | date | str | int]] = []

    for index in range(7):
        day_date = schedule.week_start_date + timedelta(days=index)
        shift_1_label = getattr(line, f"day_{index}_shift_1", "") or ""
        shift_2_label = getattr(line, f"day_{index}_shift_2", "") or ""
        shift_1_hours, shift_1_night = resolve_shift_metrics(
            shift_1_label,
            config=config,
            shift_templates=shift_templates,
        )
        shift_2_hours, shift_2_night = resolve_shift_metrics(
            shift_2_label,
            config=config,
            shift_templates=shift_templates,
        )
        compensation_mode = getattr(line, f"day_{index}_compensation_mode", "") or ""
        compensation_hours = decimal_hours(getattr(line, f"day_{index}_compensation_hours", Decimal("0.00")) or "0")
        breakdown.append(
            {
                "index": index,
                "date": day_date,
                "shift_1_label": shift_1_label,
                "shift_2_label": shift_2_label,
                "worked_hours": (shift_1_hours + shift_2_hours).quantize(TWO_DECIMALS),
                "night_hours": (shift_1_night + shift_2_night).quantize(TWO_DECIMALS),
                "special_label": get_special_day_label(day_date),
                "compensation_mode": compensation_mode,
                "compensation_hours": compensation_hours,
            }
        )

    return breakdown


def get_weekly_rest_day_index(
    line: ScheduleLine,
    day_breakdown: list[dict[str, Decimal | date | str | int]],
    *,
    shift_templates: dict[str, ShiftTemplate] | None = None,
) -> int:
    sunday_info = next((item for item in day_breakdown if int(item["index"]) == 0), None)
    sunday_worked = decimal_hours((sunday_info or {}).get("worked_hours", Decimal("0.00"))) > Decimal("0.00")
    if not sunday_worked:
        return 0

    candidate_indexes: list[int] = []
    for day_info in day_breakdown:
        index = int(day_info["index"])
        if index == 0:
            continue
        worked_hours = decimal_hours(day_info["worked_hours"])
        if worked_hours > Decimal("0.00"):
            continue
        special_label = str(day_info.get("special_label") or "")
        is_non_worked_holiday = index != 0 and "Festivo" in special_label
        shift_categories = {
            get_shift_non_work_category(str(day_info.get("shift_1_label") or ""), shift_templates=shift_templates),
            get_shift_non_work_category(str(day_info.get("shift_2_label") or ""), shift_templates=shift_templates),
        }
        compensation_mode = str(day_info.get("compensation_mode") or "")
        if is_non_worked_holiday or "rest" in shift_categories or compensation_mode in {
            ScheduleLine.CompensationMode.PAY_DAY,
            *ADVANCE_DAY_COMPENSATION_MODES,
        }:
            candidate_indexes.append(index)

    if candidate_indexes:
        return min(candidate_indexes)

    return 0


def build_expected_week_plan(
    line: ScheduleLine,
    *,
    day_breakdown: list[dict[str, Decimal | date | str | int]] | None = None,
    config: SystemConfiguration | None = None,
    shift_templates: dict[str, ShiftTemplate] | None = None,
    scope_indexes: set[int] | None = None,
) -> dict[str, object]:
    config = config or SystemConfiguration.load()
    day_breakdown = day_breakdown or build_line_day_breakdown(line, config=config, shift_templates=shift_templates)
    shift_templates = shift_templates or get_selected_shift_templates(line)
    day_reference_hours = get_line_day_reference_hours(line, config=config)
    day_reference_exact_minutes = get_line_day_reference_minutes_exact(line, config=config)
    weekly_target_hours = decimal_hours(line.weekly_target_hours or config.default_weekly_hours or "0")
    weekly_target_exact_minutes = decimal_hours_to_exact_minutes(weekly_target_hours)
    base_work_days = get_line_base_work_days(line, config=config)
    programming_interval_minutes = get_programming_interval_minutes(config=config)
    mandatory_rest_index = get_weekly_rest_day_index(line, day_breakdown, shift_templates=shift_templates)
    scope_indexes = scope_indexes if scope_indexes is not None else get_schedule_line_scope_indexes(line)
    external_loan_hours_by_index = get_schedule_line_external_loan_hours_by_index(line, config=config)
    day_plans: list[dict[str, object]] = []
    pending_expected_indexes: list[int] = []
    reducer_indexes: list[int] = []

    for day_info in day_breakdown:
        index = int(day_info["index"])
        day_date = day_info["date"]
        compensation_mode = str(day_info.get("compensation_mode") or "")
        special_label = str(day_info.get("special_label") or "")
        worked_hours = decimal_hours(day_info["worked_hours"])
        is_holiday = day_date.weekday() != 6 and "Festivo" in special_label
        is_non_worked_holiday = is_holiday and worked_hours == Decimal("0.00")
        shift_categories = {
            get_shift_non_work_category(str(day_info.get("shift_1_label") or ""), shift_templates=shift_templates),
            get_shift_non_work_category(str(day_info.get("shift_2_label") or ""), shift_templates=shift_templates),
        }
        external_loan_hours = decimal_hours(external_loan_hours_by_index.get(index, Decimal("0.00")))
        is_loan_day = worked_hours == Decimal("0.00") and "loan" in shift_categories
        is_unlinked_loan_day = is_loan_day and external_loan_hours == Decimal("0.00")
        is_absence_day = (
            worked_hours == Decimal("0.00")
            and "absence" in shift_categories
            and compensation_mode not in {ScheduleLine.CompensationMode.PAY_DAY, *ADVANCE_DAY_COMPENSATION_MODES}
        )
        is_leave_day = (
            worked_hours == Decimal("0.00")
            and "leave" in shift_categories
            and not is_non_worked_holiday
            and compensation_mode not in {ScheduleLine.CompensationMode.PAY_DAY, *ADVANCE_DAY_COMPENSATION_MODES}
        )
        is_additional_rest_day = (
            worked_hours == Decimal("0.00")
            and "rest" in shift_categories
            and compensation_mode not in {ScheduleLine.CompensationMode.PAY_DAY, *ADVANCE_DAY_COMPENSATION_MODES}
            and index != mandatory_rest_index
            and not is_non_worked_holiday
        )

        expected_hours = day_reference_hours
        expected_reason = "laborable"
        if index not in scope_indexes:
            expected_hours = Decimal("0.00")
            expected_reason = "fuera_de_rango"
        elif index == mandatory_rest_index:
            expected_hours = Decimal("0.00")
            expected_reason = "descanso_obligatorio"
        elif compensation_mode == ScheduleLine.CompensationMode.PAY_DAY:
            expected_hours = Decimal("0.00")
            expected_reason = "descanso_compensatorio"
        elif compensation_mode in ADVANCE_DAY_COMPENSATION_MODES:
            expected_hours = Decimal("0.00")
            expected_reason = "descanso_adelantado"
        elif is_additional_rest_day:
            expected_hours = Decimal("0.00")
            expected_reason = "descanso_adicional"
        elif is_non_worked_holiday:
            expected_hours = Decimal("0.00")
            expected_reason = "festivo_no_trabajado"
        elif is_unlinked_loan_day:
            expected_hours = Decimal("0.00")
            expected_reason = "prestamo_sin_destino"
        elif is_absence_day:
            expected_hours = Decimal("0.00")
            expected_reason = "inasistencia"
        elif is_leave_day:
            expected_hours = Decimal("0.00")
            expected_reason = "novedad_no_laborable"

        if expected_hours > Decimal("0.00"):
            pending_expected_indexes.append(index)
        elif expected_reason in {
            "descanso_compensatorio",
            "descanso_adelantado",
            "descanso_adicional",
            "festivo_no_trabajado",
            "novedad_no_laborable",
            "prestamo_sin_destino",
            "inasistencia",
        }:
            reducer_indexes.append(index)

        day_plans.append(
            {
                "index": index,
                "date": day_date,
                "expected_hours": decimal_hours(expected_hours),
                "expected_reason": expected_reason,
                "is_holiday": is_holiday,
                "is_non_worked_holiday": is_non_worked_holiday,
                "is_leave_day": is_leave_day,
                "is_loan_day": is_loan_day,
                "is_unlinked_loan_day": is_unlinked_loan_day,
                "external_loan_hours": external_loan_hours,
                "is_absence_day": is_absence_day,
                "is_mandatory_rest_day": index == mandatory_rest_index,
                "is_advance_rest_day": compensation_mode in ADVANCE_DAY_COMPENSATION_MODES,
                "is_compensatory_rest_day": compensation_mode == ScheduleLine.CompensationMode.PAY_DAY,
                "is_additional_rest_day": is_additional_rest_day,
                "worked_hours": worked_hours,
            }
        )

    expected_work_days = len(pending_expected_indexes)
    if expected_work_days > 0:
        if weekly_target_hours > Decimal("0.00"):
            expected_weekly_exact_minutes = (
                weekly_target_exact_minutes * Decimal(str(expected_work_days))
            ) / Decimal(str(base_work_days))
        else:
            expected_weekly_exact_minutes = day_reference_exact_minutes * Decimal(str(expected_work_days))

        expected_weekly_programmable_minutes = round_minutes_to_interval(
            expected_weekly_exact_minutes,
            programming_interval_minutes,
        )
        expected_weekly_hours = minutes_to_decimal_hours(expected_weekly_programmable_minutes)

        total_cents = int((expected_weekly_hours * Decimal("100")).quantize(Decimal("1")))
        base_cents, remainder_cents = divmod(total_cents, expected_work_days)
        expected_hours_by_index = {
            index: Decimal(base_cents + (1 if position < remainder_cents else 0)) / Decimal("100")
            for position, index in enumerate(pending_expected_indexes)
        }
        for day_plan in day_plans:
            day_plan["expected_hours"] = expected_hours_by_index.get(int(day_plan["index"]), Decimal("0.00"))
    else:
        expected_weekly_exact_minutes = Decimal("0")
        expected_weekly_programmable_minutes = Decimal("0")
        expected_weekly_hours = Decimal("0.00")

    return {
        "mandatory_rest_index": mandatory_rest_index,
        "base_work_days": base_work_days,
        "programming_interval_minutes": programming_interval_minutes,
        "day_reference_hours": day_reference_hours,
        "day_reference_exact_minutes": day_reference_exact_minutes,
        "expected_work_days": expected_work_days,
        "expected_weekly_exact_minutes": expected_weekly_exact_minutes,
        "expected_weekly_programmable_minutes": expected_weekly_programmable_minutes,
        "rounding_adjustment_minutes": expected_weekly_programmable_minutes - expected_weekly_exact_minutes,
        "expected_weekly_hours": expected_weekly_hours,
        "reducer_indexes": reducer_indexes,
        "day_plans": day_plans,
    }


def build_compensation_entries(
    line: ScheduleLine,
    *,
    day_breakdown: list[dict[str, Decimal | date | str | int]] | None = None,
    expected_plan: dict[str, object] | None = None,
    config: SystemConfiguration | None = None,
    shift_templates: dict[str, ShiftTemplate] | None = None,
) -> list[dict[str, Decimal | int | str | bool]]:
    config = config or SystemConfiguration.load()
    day_breakdown = day_breakdown or build_line_day_breakdown(line, config=config, shift_templates=shift_templates)
    expected_plan = expected_plan or build_expected_week_plan(
        line,
        day_breakdown=day_breakdown,
        config=config,
        shift_templates=shift_templates,
    )
    expected_plan_by_index = {
        int(day_plan["index"]): day_plan
        for day_plan in expected_plan["day_plans"]
    }
    daily_target_hours = decimal_hours(line.daily_max_hours or config.default_daily_max_hours or Decimal("0.00"))

    return [
        {
            "index": int(day_info["index"]),
            "mode": str(day_info["compensation_mode"] or ""),
            "hours": decimal_hours(day_info["compensation_hours"]),
            "worked_hours": decimal_hours(day_info["worked_hours"]),
            "daily_target_hours": daily_target_hours,
            "expected_hours": decimal_hours(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("expected_hours", Decimal("0.00"))
            ),
            "expected_reason": str(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("expected_reason", "") or ""
            ),
            "is_mandatory_rest_day": bool(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("is_mandatory_rest_day", False)
            ),
            "is_holiday": bool(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("is_holiday", False)
            ),
            "is_non_worked_holiday": bool(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("is_non_worked_holiday", False)
            ),
            "is_leave_day": bool(expected_plan_by_index.get(int(day_info["index"]), {}).get("is_leave_day", False)),
            "is_absence_day": bool(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("is_absence_day", False)
            ),
            "is_additional_rest_day": bool(
                expected_plan_by_index.get(int(day_info["index"]), {}).get("is_additional_rest_day", False)
            ),
            "special_generated": bool(
                decimal_hours(day_info["worked_hours"]) > Decimal("0.00") and day_info["special_label"]
            ),
        }
        for day_info in day_breakdown
    ]


def schedule_day_is_special(line: ScheduleLine, index: int) -> bool:
    schedule = getattr(line, "schedule", None)
    if schedule is None or not schedule.week_start_date:
        return False
    day_date = schedule.week_start_date + timedelta(days=index)
    return bool(get_special_day_label(day_date))


def get_schedule_line_activity_indices(line: ScheduleLine) -> list[int]:
    activity_indices: list[int] = []
    for index in range(7):
        shift_1_label = normalize_shift_label(getattr(line, f"day_{index}_shift_1", "") or "")
        shift_2_label = normalize_shift_label(getattr(line, f"day_{index}_shift_2", "") or "")
        compensation_mode = str(getattr(line, f"day_{index}_compensation_mode", "") or "").strip()
        compensation_hours = decimal_hours(
            getattr(line, f"day_{index}_compensation_hours", Decimal("0.00")) or "0"
        )
        if shift_1_label or shift_2_label or compensation_mode or compensation_hours != Decimal("0.00"):
            activity_indices.append(index)
    return activity_indices


def schedule_line_has_loan_marker_on_day(line: ScheduleLine, index: int) -> bool:
    shift_1_label = getattr(line, f"day_{index}_shift_1", "") or ""
    shift_2_label = getattr(line, f"day_{index}_shift_2", "") or ""
    return any(
        normalize_shift_key(label) in LOAN_SHIFT_LABELS
        for label in (shift_1_label, shift_2_label)
        if normalize_shift_label(label)
    )


def get_schedule_line_loan_indexes(line: ScheduleLine) -> set[int]:
    return {
        index
        for index in range(7)
        if schedule_line_has_loan_marker_on_day(line, index)
    }


def get_schedule_line_progression_key(line: ScheduleLine) -> tuple[date, int, int, datetime, int, int]:
    schedule = getattr(line, "schedule", None)
    week_start = getattr(schedule, "week_start_date", None) or date.min
    schedule_created_at = getattr(schedule, "created_at", None) or datetime.min
    activity_indices = get_schedule_line_activity_indices(line)
    if activity_indices:
        first_activity_index = activity_indices[0]
        last_activity_index = activity_indices[-1]
    else:
        first_activity_index = 7
        last_activity_index = 7
    return (
        week_start,
        first_activity_index,
        last_activity_index,
        schedule_created_at,
        getattr(schedule, "pk", 0) or 0,
        getattr(line, "pk", 0) or 0,
    )


def get_schedule_line_scope_indexes(line: ScheduleLine) -> set[int]:
    schedule = getattr(line, "schedule", None)
    employee_identifier = (getattr(line, "employee_identifier", "") or "").strip()
    if schedule is None or not schedule.week_start_date or not employee_identifier:
        return set(range(7))

    same_week_lines = list(
        ScheduleLine.objects.select_related("schedule").filter(
            employee_identifier=employee_identifier,
            schedule__week_start_date=schedule.week_start_date,
        )
    )
    if not any(getattr(candidate, "pk", None) == getattr(line, "pk", None) for candidate in same_week_lines):
        same_week_lines.append(line)

    ordered_lines = sorted(same_week_lines, key=get_schedule_line_progression_key)
    if len(ordered_lines) <= 1:
        return set(range(7))

    current_loan_indexes = get_schedule_line_loan_indexes(line)
    other_loan_indexes: set[int] = set()
    for candidate in ordered_lines:
        if getattr(candidate, "pk", None) == getattr(line, "pk", None):
            continue
        other_loan_indexes.update(get_schedule_line_loan_indexes(candidate))

    if other_loan_indexes and not current_loan_indexes:
        current_activity_indexes = set(get_schedule_line_activity_indices(line))
        if current_activity_indexes:
            matched_indexes = current_activity_indexes & other_loan_indexes
            return matched_indexes or other_loan_indexes
        return set(other_loan_indexes)

    activity_meta = [
        {
            "line": candidate,
            "activity_indices": get_schedule_line_activity_indices(candidate),
        }
        for candidate in ordered_lines
    ]
    current_position = next(
        (
            index
            for index, item in enumerate(activity_meta)
            if getattr(item["line"], "pk", None) == getattr(line, "pk", None)
        ),
        None,
    )
    if current_position is None:
        return set(range(7))

    current_activity = activity_meta[current_position]["activity_indices"]
    if not current_activity:
        return set()

    has_previous_active_line = any(item["activity_indices"] for item in activity_meta[:current_position])
    scope_start = current_activity[0] if has_previous_active_line else 0
    next_active_first = next(
        (
            item["activity_indices"][0]
            for item in activity_meta[current_position + 1 :]
            if item["activity_indices"]
        ),
        None,
    )
    scope_end = 6 if next_active_first is None else max(current_activity[-1], next_active_first - 1)
    return {index for index in range(scope_start, scope_end + 1)}


def get_schedule_line_external_loan_hours(
    line: ScheduleLine,
    *,
    config: SystemConfiguration | None = None,
) -> Decimal:
    return sum(
        get_schedule_line_external_loan_hours_by_index(line, config=config).values(),
        Decimal("0.00"),
    ).quantize(TWO_DECIMALS)


def get_schedule_line_external_loan_hours_by_index(
    line: ScheduleLine,
    *,
    config: SystemConfiguration | None = None,
) -> dict[int, Decimal]:
    schedule = getattr(line, "schedule", None)
    employee_identifier = (getattr(line, "employee_identifier", "") or "").strip()
    loan_indexes = get_schedule_line_loan_indexes(line)
    if schedule is None or not schedule.week_start_date or not employee_identifier or not loan_indexes:
        return {}

    config = config or SystemConfiguration.load()
    same_week_lines = (
        ScheduleLine.objects.select_related("schedule")
        .filter(
            employee_identifier=employee_identifier,
            schedule__week_start_date=schedule.week_start_date,
        )
        .exclude(pk=getattr(line, "pk", None))
    )
    borrowed_hours_by_index: dict[int, Decimal] = {index: Decimal("0.00") for index in loan_indexes}
    for candidate in same_week_lines:
        candidate_templates = get_selected_shift_templates(candidate)
        candidate_breakdown = build_line_day_breakdown(
            candidate,
            config=config,
            shift_templates=candidate_templates,
        )
        for day_info in candidate_breakdown:
            index = int(day_info["index"])
            if index in loan_indexes:
                borrowed_hours_by_index[index] = (
                    borrowed_hours_by_index.get(index, Decimal("0.00"))
                    + decimal_hours(day_info["worked_hours"])
                ).quantize(TWO_DECIMALS)
    return {
        index: hours.quantize(TWO_DECIMALS)
        for index, hours in borrowed_hours_by_index.items()
        if hours > Decimal("0.00")
    }


def resolve_compensation_usage(
    compensation_entries: list[dict[str, Decimal | int | str]],
    *,
    available_day_balance: Decimal,
    starting_day_balance: Decimal | None = None,
    available_hour_balance: Decimal,
    available_advance_pending_balance: Decimal = Decimal("0.00"),
    day_reference_hours: Decimal,
    weekly_target_hours: Decimal | None = None,
    base_work_days: int | None = None,
) -> dict[str, Decimal | int | list[int] | dict[int, dict[str, Decimal | str | bool]]]:
    remaining_day_balance = decimal_hours(available_day_balance)
    starting_day_balance = decimal_hours(starting_day_balance if starting_day_balance is not None else available_day_balance)
    remaining_hour_balance = max(decimal_hours(available_hour_balance), Decimal("0.00"))
    weekly_target_hours = max(decimal_hours(weekly_target_hours or "0"), Decimal("0.00"))
    base_work_days = max(int(base_work_days or 6), 1)
    cumulative_counted_hours = Decimal("0.00")
    cumulative_overtime_hours = Decimal("0.00")
    repayment_plan = build_company_day_repayment_plan(
        compensation_entries,
        starting_day_balance=starting_day_balance,
        day_reference_hours=day_reference_hours,
        base_work_days=base_work_days,
    )
    valid_repayment_indexes = set(repayment_plan["valid_repayment_indexes"])
    automatic_repayment_indexes = set(repayment_plan["automatic_repayment_indexes"])
    invalid_repayment_reasons = dict(repayment_plan["invalid_repayment_reasons"])
    excluded_hours_by_index = {
        int(index): decimal_hours(hours)
        for index, hours in repayment_plan["excluded_hours_by_index"].items()
    }

    payment_days_used = 0
    advance_rest_days_used = 0
    additional_rest_days_used = 0
    company_day_repayments_used = 0
    money_payment_days_used = 0
    payment_hours_used = Decimal("0.00")
    money_payment_hours_used = Decimal("0.00")
    invalid_pay_day_indices: list[int] = []
    invalid_pay_hours_indices: list[int] = []
    invalid_pay_money_day_indices: list[int] = []
    invalid_pay_money_indices: list[int] = []
    invalid_advance_day_indices: list[int] = []
    invalid_auto_rest_day_indices: list[int] = []
    invalid_company_day_repayment_indices: list[int] = []
    generated_special_days = 0
    excluded_company_day_hours = Decimal("0.00")
    day_states: dict[int, dict[str, Decimal | str | bool]] = {}

    for entry in sorted(compensation_entries, key=lambda item: int(item.get("index", 0))):
        index = int(entry.get("index", 0))
        compensation_mode = str(entry.get("mode") or "")
        compensation_hours = decimal_hours(entry.get("hours", Decimal("0.00")) or "0")
        worked_hours = decimal_hours(entry.get("worked_hours", Decimal("0.00")) or "0")
        special_generated = bool(entry.get("special_generated"))
        is_additional_rest_day = bool(entry.get("is_additional_rest_day", False))
        is_absence_day = bool(entry.get("is_absence_day", False))
        day_state: dict[str, Decimal | str | bool] = {
            "mode": compensation_mode,
            "requested_hours": compensation_hours,
            "source": "",
            "valid": True,
            "available_day_balance": remaining_day_balance.quantize(TWO_DECIMALS),
            "available_hour_balance": remaining_hour_balance.quantize(TWO_DECIMALS),
            "available_advance_pending_balance": Decimal("0.00"),
            "remaining_day_balance": remaining_day_balance,
            "remaining_hour_balance": remaining_hour_balance,
            "remaining_advance_pending_balance": Decimal("0.00"),
            "generated_day": False,
            "generated_hours": Decimal("0.00"),
            "hour_difference": Decimal("0.00"),
            "day_movement_type": "",
            "day_movement_description": "",
            "hours_movement_type": "",
            "hours_movement_description": "",
            "applied_hours": Decimal("0.00"),
            "applied_day_delta": Decimal("0.00"),
            "repayment_reason": "",
            "is_complete_work_day": bool(
                is_complete_work_day_entry(entry, day_reference_hours=day_reference_hours)
            ),
            "excluded_hours_from_overtime": Decimal("0.00"),
        }

        requested_day_delta = Decimal("0.00")
        repays_company_day = False
        selected_automatic_repayment = (
            compensation_mode != COMPANY_DAY_REPAYMENT_MODE and index in automatic_repayment_indexes
        )
        if compensation_mode == COMPANY_DAY_REPAYMENT_MODE or selected_automatic_repayment:
            if index in valid_repayment_indexes:
                requested_day_delta = Decimal("1.00")
                repays_company_day = True
                day_state["source"] = "signed_day_balance" if compensation_mode == COMPANY_DAY_REPAYMENT_MODE else "automatic_repayment"
                day_state["day_movement_type"] = COMPANY_DAY_REPAYMENT_MOVEMENT
                day_state["day_movement_description"] = (
                    COMPANY_DAY_REPAYMENT_MANUAL_DESCRIPTION
                    if compensation_mode == COMPANY_DAY_REPAYMENT_MODE
                    else COMPANY_DAY_REPAYMENT_AUTO_DESCRIPTION
                )
            else:
                invalid_company_day_repayment_indices.append(index)
                day_state["source"] = "repayment_invalid"
                day_state["valid"] = False
                day_state["repayment_reason"] = str(invalid_repayment_reasons.get(index, ""))
        elif compensation_mode == ScheduleLine.CompensationMode.PAY_DAY:
            requested_day_delta = Decimal("-1.00")
            day_state["source"] = "signed_day_balance"
            day_state["day_movement_type"] = ScheduleBalanceMovement.MovementType.PAY_DAY
            day_state["day_movement_description"] = "Pago con descanso"
        elif compensation_mode in MONEY_DAY_COMPENSATION_MODES:
            if remaining_day_balance >= Decimal("1.00"):
                remaining_day_balance = (remaining_day_balance - Decimal("1.00")).quantize(TWO_DECIMALS)
                money_payment_days_used += 1
                day_state["source"] = "day_balance"
                day_state["day_movement_type"] = ScheduleBalanceMovement.MovementType.PAY_MONEY_DAY
                day_state["day_movement_description"] = "Pago en dinero por dia"
            else:
                invalid_pay_money_day_indices.append(index)
                day_state["source"] = "insufficient"
                day_state["valid"] = False
        elif compensation_mode in ADVANCE_DAY_COMPENSATION_MODES:
            requested_day_delta = Decimal("-1.00")
            day_state["source"] = "signed_day_balance"
            day_state["day_movement_type"] = ScheduleBalanceMovement.MovementType.ADVANCE_DAY
            day_state["day_movement_description"] = "Descanso adelantado"
        elif is_additional_rest_day:
            requested_day_delta = Decimal("-1.00")
            day_state["source"] = "signed_day_balance"
            day_state["day_movement_type"] = ScheduleBalanceMovement.MovementType.ADDITIONAL_REST
            day_state["day_movement_description"] = "Descanso adicional"
        elif is_absence_day:
            requested_day_delta = Decimal("-1.00")
            day_state["source"] = "signed_day_balance"
            day_state["day_movement_type"] = ScheduleBalanceMovement.MovementType.ADDITIONAL_REST
            day_state["day_movement_description"] = "Inasistencia"
        elif compensation_mode == ScheduleLine.CompensationMode.PAY_HOURS:
            if compensation_hours <= Decimal("0.00") or remaining_hour_balance < compensation_hours:
                invalid_pay_hours_indices.append(index)
                day_state["valid"] = False
                day_state["source"] = "insufficient"
            else:
                payment_hours_used += compensation_hours
                remaining_hour_balance = (remaining_hour_balance - compensation_hours).quantize(TWO_DECIMALS)
                day_state["source"] = "hour_balance"
                day_state["hours_movement_type"] = ScheduleBalanceMovement.MovementType.PAY_HOURS
                day_state["hours_movement_description"] = "Pago con horas"
                day_state["applied_hours"] = compensation_hours
        elif compensation_mode in MONEY_HOUR_COMPENSATION_MODES:
            if compensation_hours <= Decimal("0.00") or remaining_hour_balance < compensation_hours:
                invalid_pay_money_indices.append(index)
                day_state["valid"] = False
                day_state["source"] = "insufficient"
            else:
                money_payment_hours_used += compensation_hours
                remaining_hour_balance = (remaining_hour_balance - compensation_hours).quantize(TWO_DECIMALS)
                day_state["source"] = "hour_balance"
                day_state["hours_movement_type"] = ScheduleBalanceMovement.MovementType.PAY_MONEY_HOURS
                day_state["hours_movement_description"] = "Pago en dinero por horas"
                day_state["applied_hours"] = compensation_hours

        if requested_day_delta != Decimal("0.00"):
            projected_day_balance = (remaining_day_balance + requested_day_delta).quantize(TWO_DECIMALS)
            if projected_day_balance < (MAX_ADVANCE_REST_PENDING_DAYS * Decimal("-1.00")):
                day_state["valid"] = False
                day_state["source"] = "day_limit"
                if compensation_mode == COMPANY_DAY_REPAYMENT_MODE:
                    invalid_company_day_repayment_indices.append(index)
                    day_state["repayment_reason"] = "no_negative_balance"
                elif compensation_mode == ScheduleLine.CompensationMode.PAY_DAY:
                    invalid_pay_day_indices.append(index)
                elif compensation_mode in ADVANCE_DAY_COMPENSATION_MODES:
                    invalid_advance_day_indices.append(index)
                else:
                    invalid_auto_rest_day_indices.append(index)
            else:
                remaining_day_balance = projected_day_balance
                day_state["applied_day_delta"] = requested_day_delta
                if compensation_mode == COMPANY_DAY_REPAYMENT_MODE:
                    company_day_repayments_used += 1
                elif compensation_mode == ScheduleLine.CompensationMode.PAY_DAY:
                    payment_days_used += 1
                elif compensation_mode in ADVANCE_DAY_COMPENSATION_MODES:
                    advance_rest_days_used += 1
                    additional_rest_days_used += 1
                else:
                    additional_rest_days_used += 1

        if (
            special_generated
            and bool(day_state["is_complete_work_day"])
            and worked_hours > Decimal("0.00")
            and compensation_mode != COMPANY_DAY_REPAYMENT_MODE
            and not selected_automatic_repayment
        ):
            remaining_day_balance = (remaining_day_balance + Decimal("1.00")).quantize(TWO_DECIMALS)
            day_state["generated_day"] = True
            generated_special_days += 1

        excluded_hours = Decimal("0.00")
        if repays_company_day:
            excluded_hours = excluded_hours_by_index.get(index, Decimal("0.00"))
            excluded_company_day_hours = (excluded_company_day_hours + excluded_hours).quantize(TWO_DECIMALS)
            day_state["excluded_hours_from_overtime"] = excluded_hours

        counted_worked_hours = max(worked_hours - excluded_hours, Decimal("0.00")).quantize(TWO_DECIMALS)
        previous_overtime_hours = cumulative_overtime_hours
        cumulative_counted_hours = (cumulative_counted_hours + counted_worked_hours).quantize(TWO_DECIMALS)
        cumulative_overtime_hours = max(cumulative_counted_hours - weekly_target_hours, Decimal("0.00")).quantize(
            TWO_DECIMALS
        )
        generated_hours = max(cumulative_overtime_hours - previous_overtime_hours, Decimal("0.00")).quantize(
            TWO_DECIMALS
        )
        if generated_hours > Decimal("0.00"):
            remaining_hour_balance = (remaining_hour_balance + generated_hours).quantize(TWO_DECIMALS)
            day_state["generated_hours"] = generated_hours
        day_state["hour_difference"] = generated_hours

        day_state["remaining_day_balance"] = remaining_day_balance
        day_state["remaining_hour_balance"] = remaining_hour_balance
        day_states[index] = day_state

    return {
        "payment_days_used": payment_days_used,
        "advance_rest_days_used": advance_rest_days_used,
        "additional_rest_days_used": additional_rest_days_used,
        "company_day_repayments_used": company_day_repayments_used,
        "automatic_company_day_repayments_used": len(automatic_repayment_indexes),
        "automatic_repayment_indexes": sorted(automatic_repayment_indexes),
        "payment_days_from_day_balance": payment_days_used,
        "payment_days_from_hour_balance": 0,
        "uncovered_payment_days": len(invalid_pay_day_indices),
        "money_payment_days_used": money_payment_days_used,
        "payment_day_hour_equivalent": Decimal("0.00"),
        "payment_hours_used": payment_hours_used.quantize(TWO_DECIMALS),
        "money_payment_hours_used": money_payment_hours_used.quantize(TWO_DECIMALS),
        "invalid_pay_day_indices": invalid_pay_day_indices,
        "invalid_pay_hours_indices": invalid_pay_hours_indices,
        "invalid_pay_money_day_indices": invalid_pay_money_day_indices,
        "invalid_pay_money_indices": invalid_pay_money_indices,
        "invalid_advance_day_indices": invalid_advance_day_indices,
        "invalid_advance_day_with_balance_indices": [],
        "invalid_auto_rest_day_indices": invalid_auto_rest_day_indices,
        "invalid_company_day_repayment_indices": invalid_company_day_repayment_indices,
        "invalid_company_day_repayment_reasons": invalid_repayment_reasons,
        "generated_special_days": generated_special_days,
        "excluded_company_day_hours": excluded_company_day_hours.quantize(TWO_DECIMALS),
        "generated_overtime_hours": cumulative_overtime_hours.quantize(TWO_DECIMALS),
        "remaining_day_balance": remaining_day_balance.quantize(TWO_DECIMALS),
        "remaining_hour_balance": remaining_hour_balance.quantize(TWO_DECIMALS),
        "remaining_advance_pending_balance": Decimal("0.00"),
        "repayment_plan": repayment_plan,
        "day_states": day_states,
    }


def summarize_line_payments(line: ScheduleLine) -> tuple[int, int, Decimal, int, Decimal]:
    payment_days_used = 0
    advance_rest_days_used = 0
    payment_hours_used = Decimal("0.00")
    money_payment_days_used = 0
    money_payment_hours_used = Decimal("0.00")

    for index in range(7):
        compensation_mode = getattr(line, f"day_{index}_compensation_mode", "") or ""
        compensation_hours = decimal_hours(
            getattr(line, f"day_{index}_compensation_hours", Decimal("0.00")) or "0"
        )

        if compensation_mode == ScheduleLine.CompensationMode.PAY_DAY:
            payment_days_used += 1
        elif compensation_mode in ADVANCE_DAY_COMPENSATION_MODES:
            advance_rest_days_used += 1
        elif compensation_mode == ScheduleLine.CompensationMode.PAY_HOURS:
            payment_hours_used += compensation_hours
        elif compensation_mode in MONEY_DAY_COMPENSATION_MODES:
            money_payment_days_used += 1
        elif compensation_mode in MONEY_HOUR_COMPENSATION_MODES:
            money_payment_hours_used += compensation_hours

    return (
        payment_days_used,
        advance_rest_days_used,
        payment_hours_used.quantize(TWO_DECIMALS),
        money_payment_days_used,
        money_payment_hours_used.quantize(TWO_DECIMALS),
    )


def get_schedule_line_balance_snapshot(
    line: ScheduleLine,
    config: SystemConfiguration | None = None,
) -> dict[str, Decimal]:
    config = config or SystemConfiguration.load()
    zero_balance = {
        "prior_day_balance": Decimal("0.00"),
        "prior_hour_balance": Decimal("0.00"),
        "prior_total_balance": Decimal("0.00"),
        "prior_advance_pending_balance": Decimal("0.00"),
        "prior_day_equivalent_hours": Decimal("0.00"),
        "day_reference_hours": get_line_day_reference_hours(line, config=config),
    }

    employee_identifier = (line.employee_identifier or "").strip()
    schedule = getattr(line, "schedule", None)
    if not employee_identifier or schedule is None or not schedule.week_start_date:
        return zero_balance

    current_key = get_schedule_line_progression_key(line)
    previous_candidates = list(
        ScheduleLine.objects.filter(
            employee_identifier=employee_identifier,
            schedule__week_start_date__lte=schedule.week_start_date,
        )
        .exclude(pk=getattr(line, "pk", None))
        .select_related("schedule")
    )
    previous_line = None
    previous_key = None
    for candidate in previous_candidates:
        candidate_key = get_schedule_line_progression_key(candidate)
        if candidate_key >= current_key:
            continue
        if previous_key is None or candidate_key > previous_key:
            previous_line = candidate
            previous_key = candidate_key

    if previous_line is None:
        try:
            initial_balance = (
                EmployeeInitialBalance.objects.filter(employee_identifier=employee_identifier)
                .order_by("-updated_at", "-pk")
                .first()
            )
        except (ProgrammingError, OperationalError):
            return zero_balance
        if initial_balance is None:
            return zero_balance

        prior_day_balance = decimal_hours(initial_balance.initial_day_balance)
        prior_hour_balance = max(decimal_hours(initial_balance.initial_hour_balance), Decimal("0.00"))
        day_reference_hours = zero_balance["day_reference_hours"]
        return {
            "prior_day_balance": prior_day_balance,
            "prior_hour_balance": prior_hour_balance,
            "prior_total_balance": (
                prior_hour_balance + (prior_day_balance * day_reference_hours)
            ).quantize(TWO_DECIMALS),
            "prior_advance_pending_balance": Decimal("0.00"),
            "prior_day_equivalent_hours": (prior_day_balance * day_reference_hours).quantize(TWO_DECIMALS),
            "day_reference_hours": day_reference_hours,
        }

    prior_day_balance = decimal_hours(previous_line.accrued_day_balance)
    prior_hour_balance = max(decimal_hours(previous_line.accrued_hour_balance), Decimal("0.00"))
    day_reference_hours = zero_balance["day_reference_hours"]
    prior_total_balance = (
        prior_hour_balance + (prior_day_balance * day_reference_hours)
    ).quantize(TWO_DECIMALS)

    return {
        "prior_day_balance": prior_day_balance,
        "prior_hour_balance": prior_hour_balance,
        "prior_total_balance": prior_total_balance,
        "prior_advance_pending_balance": Decimal("0.00"),
        "prior_day_equivalent_hours": (prior_day_balance * day_reference_hours).quantize(TWO_DECIMALS),
        "day_reference_hours": day_reference_hours,
    }


def get_schedule_line_balance_snapshots(
    lines: list[ScheduleLine] | tuple[ScheduleLine, ...],
    config: SystemConfiguration | None = None,
) -> dict[int, dict[str, Decimal]]:
    config = config or SystemConfiguration.load()
    snapshots: dict[int, dict[str, Decimal]] = {}
    normalized_lines: list[ScheduleLine] = []
    identifiers: set[str] = set()
    current_pks: list[int] = []
    max_week_start: date | None = None

    for line in lines:
        employee_identifier = (getattr(line, "employee_identifier", "") or "").strip()
        schedule = getattr(line, "schedule", None)
        if not employee_identifier or schedule is None or not schedule.week_start_date:
            continue
        normalized_lines.append(line)
        identifiers.add(employee_identifier)
        if getattr(line, "pk", None):
            current_pks.append(line.pk)
        if max_week_start is None or schedule.week_start_date > max_week_start:
            max_week_start = schedule.week_start_date

    if not normalized_lines or not identifiers or max_week_start is None:
        return snapshots

    zero_snapshots = {
        getattr(line, "pk", 0): {
            "prior_day_balance": Decimal("0.00"),
            "prior_hour_balance": Decimal("0.00"),
            "prior_total_balance": Decimal("0.00"),
            "prior_advance_pending_balance": Decimal("0.00"),
            "prior_day_equivalent_hours": Decimal("0.00"),
            "day_reference_hours": get_line_day_reference_hours(line, config=config),
        }
        for line in normalized_lines
        if getattr(line, "pk", None) is not None
    }

    try:
        previous_candidates_queryset = (
            ScheduleLine.objects.filter(
                employee_identifier__in=identifiers,
                schedule__week_start_date__lte=max_week_start,
            )
            .select_related("schedule")
        )
        if current_pks:
            previous_candidates_queryset = previous_candidates_queryset.exclude(pk__in=current_pks)
        previous_candidates = list(previous_candidates_queryset)
        initial_balances = {
            balance.employee_identifier: balance
            for balance in EmployeeInitialBalance.objects.filter(employee_identifier__in=identifiers)
            .order_by("-updated_at", "-pk")
        }
    except (ProgrammingError, OperationalError):
        return zero_snapshots

    candidates_by_identifier: dict[str, list[tuple[tuple[object, ...], ScheduleLine]]] = {}
    for candidate in previous_candidates:
        cleaned_identifier = (candidate.employee_identifier or "").strip()
        if not cleaned_identifier:
            continue
        candidates_by_identifier.setdefault(cleaned_identifier, []).append(
            (get_schedule_line_progression_key(candidate), candidate)
        )

    for candidate_entries in candidates_by_identifier.values():
        candidate_entries.sort(key=lambda item: item[0])

    for line in normalized_lines:
        line_pk = getattr(line, "pk", None)
        if line_pk is None:
            continue
        zero_balance = zero_snapshots[line_pk]
        employee_identifier = (line.employee_identifier or "").strip()
        current_key = get_schedule_line_progression_key(line)
        previous_line: ScheduleLine | None = None
        previous_key: tuple[object, ...] | None = None

        for candidate_key, candidate in candidates_by_identifier.get(employee_identifier, []):
            if candidate_key >= current_key:
                continue
            if previous_key is None or candidate_key > previous_key:
                previous_line = candidate
                previous_key = candidate_key

        if previous_line is None:
            initial_balance = initial_balances.get(employee_identifier)
            if initial_balance is None:
                snapshots[line_pk] = zero_balance
                continue

            prior_day_balance = decimal_hours(initial_balance.initial_day_balance)
            prior_hour_balance = max(decimal_hours(initial_balance.initial_hour_balance), Decimal("0.00"))
            day_reference_hours = zero_balance["day_reference_hours"]
            snapshots[line_pk] = {
                "prior_day_balance": prior_day_balance,
                "prior_hour_balance": prior_hour_balance,
                "prior_total_balance": (
                    prior_hour_balance + (prior_day_balance * day_reference_hours)
                ).quantize(TWO_DECIMALS),
                "prior_advance_pending_balance": Decimal("0.00"),
                "prior_day_equivalent_hours": (prior_day_balance * day_reference_hours).quantize(TWO_DECIMALS),
                "day_reference_hours": day_reference_hours,
            }
            continue

        prior_day_balance = decimal_hours(previous_line.accrued_day_balance)
        prior_hour_balance = max(decimal_hours(previous_line.accrued_hour_balance), Decimal("0.00"))
        day_reference_hours = zero_balance["day_reference_hours"]
        snapshots[line_pk] = {
            "prior_day_balance": prior_day_balance,
            "prior_hour_balance": prior_hour_balance,
            "prior_total_balance": (
                prior_hour_balance + (prior_day_balance * day_reference_hours)
            ).quantize(TWO_DECIMALS),
            "prior_advance_pending_balance": Decimal("0.00"),
            "prior_day_equivalent_hours": (prior_day_balance * day_reference_hours).quantize(TWO_DECIMALS),
            "day_reference_hours": day_reference_hours,
        }

    return snapshots


def normalize_initial_balance_header(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    return normalized.strip("_")


def parse_initial_balance_decimal(value: object, label: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value).strip().replace(",", ".")).quantize(TWO_DECIMALS)
    except (InvalidOperation, AttributeError, ValueError):
        raise ValueError(f"Fila {row_number}: el valor de {label} no es numerico.")


def decode_uploaded_csv(uploaded_file) -> str:
    raw_content = uploaded_file.read()
    try:
        uploaded_file.seek(0)
    except (AttributeError, OSError):
        pass

    if not raw_content:
        raise ValueError("El archivo CSV esta vacio.")

    if isinstance(raw_content, str):
        return raw_content

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw_content.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise ValueError("No fue posible leer el archivo CSV. Revisa la codificacion del archivo.")


def load_initial_balance_rows(uploaded_file) -> tuple[tuple[object, ...], list[tuple[object, ...]]]:
    file_name = (getattr(uploaded_file, "name", "") or "").strip().lower()
    if file_name.endswith(".csv"):
        csv_text = decode_uploaded_csv(uploaded_file)
        sample = csv_text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(StringIO(csv_text), dialect)
        rows = [
            tuple(row)
            for row in reader
            if any(str(cell or "").strip() for cell in row)
        ]
        if not rows:
            raise ValueError("El archivo no contiene encabezados.")
        return tuple(rows[0]), rows[1:]

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    return header_row, data_rows


def build_schedule_flat_file_headers(schedule: WeeklySchedule) -> list[str]:
    headers = ["Cedula", "Empleado", "Cargo"]
    for column in schedule.get_day_columns():
        day_label = str(column["label"])
        headers.extend(
            [
                f"{day_label} turno 1",
                f"{day_label} turno 2",
                f"{day_label} modo pago",
                f"{day_label} horas pago",
                f"{day_label} inventario",
            ]
        )
    headers.extend(["Ajuste dias", "Ajuste horas"])
    return headers


def load_schedule_flat_file_rows(uploaded_file) -> tuple[tuple[object, ...], list[tuple[object, ...]]]:
    file_name = (getattr(uploaded_file, "name", "") or "").strip().lower()
    if file_name.endswith(".csv"):
        csv_text = decode_uploaded_csv(uploaded_file)
        sample = csv_text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(StringIO(csv_text), dialect)
        rows = [
            tuple(row)
            for row in reader
            if any(str(cell or "").strip() for cell in row)
        ]
        if not rows:
            raise ValueError("El archivo no contiene encabezados.")
        return tuple(rows[0]), rows[1:]

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    return header_row, data_rows


def parse_schedule_flat_decimal(value: object, label: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value).strip().replace(",", ".")).quantize(TWO_DECIMALS)
    except (InvalidOperation, AttributeError, ValueError):
        raise ValueError(f"Fila {row_number}: el valor de {label} no es numerico.")


def parse_schedule_flat_boolean(value: object) -> bool:
    normalized = normalize_initial_balance_header(value)
    return normalized in {"1", "si", "sí", "s", "true", "verdadero", "x", "yes"}


def line_has_frozen_workload_snapshot(line: ScheduleLine) -> bool:
    return (
        decimal_hours(line.weekly_target_hours) > Decimal("0.00")
        and decimal_hours(line.daily_max_hours) > Decimal("0.00")
        and int(line.base_work_days or 0) > 0
    )


def apply_job_role_snapshot(
    line: ScheduleLine,
    job_role: JobRole,
    config: SystemConfiguration,
    *,
    overwrite_workload: bool,
    role_code: str = "",
    role_name: str = "",
) -> None:
    line.job_role_code = role_code or job_role.code or line.job_role_code or ""
    line.job_role_name = role_name or job_role.name or line.job_role_name or ""
    if overwrite_workload or not line_has_frozen_workload_snapshot(line):
        line.weekly_target_hours = job_role.weekly_target_hours or config.default_weekly_hours
        line.daily_max_hours = job_role.daily_max_hours or config.default_daily_max_hours
        line.base_work_days = job_role.base_work_days or config.default_base_work_days


def stringify_upload_identifier(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    raw_value = str(value).strip()
    if raw_value.endswith(".0") and raw_value.replace(".", "", 1).isdigit():
        return raw_value[:-2]
    return raw_value


def import_employee_initial_balances(uploaded_file, *, updated_by=None) -> dict[str, object]:
    header_row, data_rows = load_initial_balance_rows(uploaded_file)
    if not header_row:
        raise ValueError("El archivo no contiene encabezados.")

    header_index = {
        normalize_initial_balance_header(header): index
        for index, header in enumerate(header_row)
        if str(header or "").strip()
    }
    identifier_index = next(
        (
            index
            for key, index in header_index.items()
            if key in {
                "cedula",
                "documento",
                "numero_documento",
                "numero_de_documento",
                "identificacion",
                "employee_identifier",
            }
        ),
        None,
    )
    name_index = next(
        (
            index
            for key, index in header_index.items()
            if key in {
                "nombre",
                "nombres",
                "empleado",
                "employee_name",
                "nombre_completo",
                "nombres_apellidos",
                "nombres_y_apellidos",
            }
        ),
        None,
    )
    day_index = next(
        (
            index
            for key, index in header_index.items()
            if key in {
                "dias",
                "dias_iniciales",
                "dias_extra",
                "dias_extras",
                "dias_extra_semana",
                "dias_extras_semana",
                "saldo_dias",
                "saldo_inicial_dias",
                "initial_day_balance",
            }
        ),
        None,
    )
    hour_index = next(
        (
            index
            for key, index in header_index.items()
            if key in {
                "horas",
                "horas_iniciales",
                "horas_extra",
                "horas_extras",
                "horas_extra_semana",
                "horas_extras_semana",
                "saldo_horas",
                "saldo_inicial_horas",
                "initial_hour_balance",
            }
        ),
        None,
    )

    if identifier_index is None:
        raise ValueError("La plantilla debe incluir una columna de Cedula o Documento.")
    if day_index is None and hour_index is None:
        raise ValueError("La plantilla debe incluir al menos una columna de Dias u Horas.")

    created_count = 0
    updated_count = 0
    touched_identifiers: list[str] = []

    with suppress_initial_balance_rebuild():
        for row_number, row in enumerate(data_rows, start=2):
            identifier = str(row[identifier_index] or "").strip().upper() if identifier_index < len(row) else ""
            if not identifier:
                continue
            employee_name = (
                str(row[name_index] or "").strip()
                if name_index is not None and name_index < len(row)
                else ""
            )
            initial_day_balance = (
                parse_initial_balance_decimal(row[day_index], "dias", row_number)
                if day_index is not None and day_index < len(row)
                else Decimal("0.00")
            )
            initial_hour_balance = (
                parse_initial_balance_decimal(row[hour_index], "horas", row_number)
                if hour_index is not None and hour_index < len(row)
                else Decimal("0.00")
            )

            balance, created = EmployeeInitialBalance.objects.get_or_create(
                employee_identifier=identifier,
                defaults={
                    "employee_name": employee_name,
                    "initial_day_balance": initial_day_balance,
                    "initial_hour_balance": initial_hour_balance,
                    "created_by": updated_by,
                    "updated_by": updated_by,
                },
            )
            if created:
                created_count += 1
            else:
                balance.employee_name = employee_name or balance.employee_name
                balance.initial_day_balance = initial_day_balance
                balance.initial_hour_balance = initial_hour_balance
                balance.updated_by = updated_by
                balance.save(
                    update_fields=[
                        "employee_name",
                        "initial_day_balance",
                        "initial_hour_balance",
                        "updated_by",
                        "updated_at",
                    ]
                )
                updated_count += 1

            touched_identifiers.append(identifier)

    touched_identifiers = sorted({identifier for identifier in touched_identifiers if identifier})

    if touched_identifiers:
        earliest_schedule = (
            ScheduleLine.objects.select_related("schedule")
            .filter(employee_identifier__in=touched_identifiers)
            .order_by("schedule__week_start_date", "pk")
            .first()
        )
        if earliest_schedule is not None:
            rebuild_balances_for_employees_from_week(
                earliest_schedule.schedule.week_start_date,
                touched_identifiers,
            )

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "processed_count": created_count + updated_count,
        "touched_identifiers": touched_identifiers,
    }


def normalize_schedule_upload_compensation_mode(value: object) -> str:
    normalized = normalize_initial_balance_header(value)
    if not normalized:
        return ""

    aliases = {
        "sin_pago": "",
        "sinpago": "",
        "ninguno": "",
        "pay_day": ScheduleLine.CompensationMode.PAY_DAY,
        "pago_dia": ScheduleLine.CompensationMode.PAY_DAY,
        "pagodia": ScheduleLine.CompensationMode.PAY_DAY,
        "pay_hours": ScheduleLine.CompensationMode.PAY_HOURS,
        "pago_horas": ScheduleLine.CompensationMode.PAY_HOURS,
        "pagohoras": ScheduleLine.CompensationMode.PAY_HOURS,
        "advance_day": ScheduleLine.CompensationMode.ADVANCE_DAY,
        "descanso_adelantado": ScheduleLine.CompensationMode.ADVANCE_DAY,
        "repay_company_day": COMPANY_DAY_REPAYMENT_MODE,
        "compensa_dia_empresa": COMPANY_DAY_REPAYMENT_MODE,
        "compensadiaempresa": COMPANY_DAY_REPAYMENT_MODE,
        "company_day_repayment": COMPANY_DAY_REPAYMENT_MODE,
        "pago_dinero_dia": ScheduleLine.CompensationMode.PAY_MONEY_DAY,
        "pagodinero_dia": ScheduleLine.CompensationMode.PAY_MONEY_DAY,
        "pago_en_dinero_por_dia": ScheduleLine.CompensationMode.PAY_MONEY_DAY,
        "pay_money_day": ScheduleLine.CompensationMode.PAY_MONEY_DAY,
        "pago_dinero_horas": ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
        "pagodinero_horas": ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
        "pago_en_dinero_por_horas": ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
        "pay_money_hours": ScheduleLine.CompensationMode.PAY_MONEY_HOURS,
        "pay_money": ScheduleLine.CompensationMode.PAY_MONEY,
        "pago_dinero": ScheduleLine.CompensationMode.PAY_MONEY,
        "pagodinero": ScheduleLine.CompensationMode.PAY_MONEY,
    }
    if normalized in aliases:
        return aliases[normalized]

    if normalized in {
        normalize_initial_balance_header(COMPANY_DAY_REPAYMENT_MODE),
        normalize_initial_balance_header(COMPANY_DAY_REPAYMENT_LABEL),
    }:
        return COMPANY_DAY_REPAYMENT_MODE

    for code, label in ScheduleLine.CompensationMode.choices:
        if normalized in {normalize_initial_balance_header(code), normalize_initial_balance_header(label)}:
            return code

    raise ValueError(f"Modo de pago no reconocido: {value}.")


def import_schedule_flat_file(
    schedule: WeeklySchedule,
    uploaded_file,
    *,
    updated_by=None,
    allow_money_payment: bool = False,
) -> dict[str, object]:
    from schedules.forms import ScheduleLineForm, build_shift_choices

    header_row, data_rows = load_schedule_flat_file_rows(uploaded_file)
    if not header_row:
        raise ValueError("El archivo no contiene encabezados.")

    expected_headers = build_schedule_flat_file_headers(schedule)
    header_index = {
        normalize_initial_balance_header(header): index
        for index, header in enumerate(header_row)
        if str(header or "").strip()
    }
    missing_headers = [
        header
        for header in expected_headers
        if normalize_initial_balance_header(header) not in header_index
    ]
    if missing_headers:
        raise ValueError(
            "Faltan columnas obligatorias en la plantilla del horario: "
            + ", ".join(missing_headers[:6])
            + ("." if len(missing_headers) <= 6 else ", ...")
        )

    day_columns = schedule.get_day_columns()
    line_by_identifier = {
        (line.employee_identifier or "").strip(): line
        for line in schedule.lines.all().order_by("job_role_name", "employee_name", "employee_identifier")
    }
    config = SystemConfiguration.load()
    role_by_name = {
        normalize_initial_balance_header(role.name): role
        for role in JobRole.objects.filter(is_active=True)
    }
    shift_choices = build_shift_choices(second_slot=False)
    secondary_shift_choices = build_shift_choices(second_slot=True)
    shared_shift_template_map = {
        shift.label: shift
        for shift in ShiftTemplate.objects.all()
    }
    shared_rest_shift_label = get_rest_shift_label()
    balance_snapshots_by_line_pk = get_schedule_line_balance_snapshots(list(line_by_identifier.values()), config=config)
    overtime_restrictions_by_identifier = {
        restriction.employee_identifier: restriction
        for restriction in EmployeeOvertimeRestriction.objects.filter(
            employee_identifier__in=list(line_by_identifier.keys()),
            is_active=True,
        ).only(
            "employee_identifier",
            "employee_name",
            "max_daily_overtime_hours",
            "max_weekly_overtime_hours",
        )
    }

    prepared_lines: list[ScheduleLine] = []
    touched_identifiers: list[str] = []
    seen_identifiers: set[str] = set()
    errors: list[str] = []

    def get_row_value(row: tuple[object, ...], header: str) -> object:
        return row[header_index[normalize_initial_balance_header(header)]]

    def serialize_decimal(value: Decimal) -> str:
        return "" if value == Decimal("0.00") else format(value.normalize(), "f")

    for row_number, row in enumerate(data_rows, start=2):
        if not any(str(cell or "").strip() for cell in row):
            continue

        identifier = stringify_upload_identifier(get_row_value(row, "Cedula"))
        if not identifier:
            errors.append(f"Fila {row_number}: debes indicar la cedula.")
            continue
        if identifier in seen_identifiers:
            errors.append(f"Fila {row_number}: la cedula {identifier} esta repetida en el archivo.")
            continue
        seen_identifiers.add(identifier)

        line = line_by_identifier.get(identifier)
        if line is None:
            errors.append(
                f"Fila {row_number}: la cedula {identifier} no existe en el horario. "
                "Primero cargala en la semana y luego vuelve a importar."
            )
            continue

        role_name = str(get_row_value(row, "Cargo") or "").strip()
        if role_name:
            matched_role = role_by_name.get(normalize_initial_balance_header(role_name))
            if matched_role is None:
                errors.append(f"Fila {row_number}: el cargo '{role_name}' no existe en la parametrizacion.")
                continue
            apply_job_role_snapshot(
                line,
                matched_role,
                config,
                overwrite_workload=not line_has_frozen_workload_snapshot(line),
            )

        employee_name = str(get_row_value(row, "Empleado") or "").strip()
        if employee_name:
            line.employee_name = employee_name

        form_data: dict[str, object] = {}
        for index in range(7):
            form_data[f"day_{index}_shift_1"] = getattr(line, f"day_{index}_shift_1", "") or ""
            form_data[f"day_{index}_shift_2"] = getattr(line, f"day_{index}_shift_2", "") or ""
            form_data[f"day_{index}_compensation_mode"] = getattr(line, f"day_{index}_compensation_mode", "") or ""
            current_hours = decimal_hours(getattr(line, f"day_{index}_compensation_hours", Decimal("0.00")) or "0")
            form_data[f"day_{index}_compensation_hours"] = serialize_decimal(current_hours)
            if bool(getattr(line, f"day_{index}_inventory", False)):
                form_data[f"day_{index}_inventory"] = "on"

        current_manual_days = decimal_hours(line.manual_day_adjustment)
        current_manual_hours = decimal_hours(line.manual_hour_adjustment)
        form_data["manual_day_adjustment"] = serialize_decimal(current_manual_days)
        form_data["manual_hour_adjustment"] = serialize_decimal(current_manual_hours)

        try:
            for day_index, column in enumerate(day_columns):
                day_label = str(column["label"])
                form_data[f"day_{day_index}_shift_1"] = str(get_row_value(row, f"{day_label} turno 1") or "").strip()
                form_data[f"day_{day_index}_shift_2"] = str(get_row_value(row, f"{day_label} turno 2") or "").strip()
                form_data[f"day_{day_index}_compensation_mode"] = normalize_schedule_upload_compensation_mode(
                    get_row_value(row, f"{day_label} modo pago")
                )
                imported_hours = parse_schedule_flat_decimal(
                    get_row_value(row, f"{day_label} horas pago"),
                    f"{day_label} horas pago",
                    row_number,
                )
                form_data[f"day_{day_index}_compensation_hours"] = serialize_decimal(imported_hours)
                if parse_schedule_flat_boolean(get_row_value(row, f"{day_label} inventario")):
                    form_data[f"day_{day_index}_inventory"] = "on"
                else:
                    form_data.pop(f"day_{day_index}_inventory", None)

            form_data["manual_day_adjustment"] = serialize_decimal(
                parse_schedule_flat_decimal(get_row_value(row, "Ajuste dias"), "ajuste dias", row_number)
            )
            form_data["manual_hour_adjustment"] = serialize_decimal(
                parse_schedule_flat_decimal(get_row_value(row, "Ajuste horas"), "ajuste horas", row_number)
            )
        except ValueError as exc:
            errors.append(str(exc))
            continue

        line_form = ScheduleLineForm(
            data=form_data,
            instance=line,
            schedule=schedule,
            shift_choices=shift_choices,
            secondary_shift_choices=secondary_shift_choices,
            readonly=False,
            allow_money_payment=allow_money_payment,
            show_admin_fields=True,
            config=config,
            balance_snapshot=balance_snapshots_by_line_pk.get(getattr(line, "pk", 0)),
            overtime_restriction=overtime_restrictions_by_identifier.get(identifier),
            shift_template_map=shared_shift_template_map,
            rest_shift_label=shared_rest_shift_label,
        )

        if not line_form.is_valid():
            row_errors: list[str] = []
            for field_name, field_errors in line_form.errors.items():
                for error in field_errors:
                    row_errors.append(f"{field_name}: {error}")
            for error in line_form.non_field_errors():
                row_errors.append(str(error))
            if not row_errors:
                row_errors.append("Hay datos invalidos en la fila.")
            errors.append(f"Fila {row_number} ({identifier}): " + " | ".join(row_errors))
            continue

        prepared_line = line_form.save(commit=False)
        prepared_lines.append(prepared_line)
        touched_identifiers.append(identifier)

    if errors:
        message = " ".join(errors[:5])
        if len(errors) > 5:
            message += f" Hay {len(errors) - 5} error(es) adicional(es) en el archivo."
        raise ValueError(message)

    if not prepared_lines:
        raise ValueError("El archivo no contiene filas de horario para procesar.")

    touched_identifiers = sorted({identifier for identifier in touched_identifiers if identifier})
    with transaction.atomic():
        for prepared_line in prepared_lines:
            prepared_line.save()
        if updated_by is not None:
            schedule.updated_by = updated_by
            schedule.save(update_fields=["updated_by", "updated_at"])
        rebuild_balances_for_employees_from_week(schedule.week_start_date, touched_identifiers)

    return {
        "processed_count": len(prepared_lines),
        "touched_identifiers": touched_identifiers,
    }


def get_schedule_line_compact_alert_summary(
    line: ScheduleLine,
    balance_snapshot: dict[str, Decimal] | None = None,
    config: SystemConfiguration | None = None,
) -> str:
    config = config or SystemConfiguration.load()
    balance_snapshot = balance_snapshot or get_schedule_line_balance_snapshot(line, config=config)
    shift_templates = get_selected_shift_templates(line)
    day_breakdown = build_line_day_breakdown(line, config=config, shift_templates=shift_templates)
    expected_plan = build_expected_week_plan(
        line,
        day_breakdown=day_breakdown,
        config=config,
        shift_templates=shift_templates,
    )
    compensation_entries = build_compensation_entries(
        line,
        day_breakdown=day_breakdown,
        expected_plan=expected_plan,
        config=config,
        shift_templates=shift_templates,
    )
    total_worked_hours = sum(decimal_hours(entry["worked_hours"]) for entry in compensation_entries).quantize(TWO_DECIMALS)
    manual_day_adjustment = decimal_hours(line.manual_day_adjustment)
    manual_hour_adjustment = decimal_hours(line.manual_hour_adjustment)
    payment_resolution = resolve_compensation_usage(
        compensation_entries,
        available_day_balance=balance_snapshot["prior_day_balance"] + manual_day_adjustment,
        starting_day_balance=balance_snapshot["prior_day_balance"],
        available_hour_balance=balance_snapshot["prior_hour_balance"] + manual_hour_adjustment,
        available_advance_pending_balance=Decimal("0.00"),
        day_reference_hours=balance_snapshot["day_reference_hours"],
        weekly_target_hours=line.weekly_target_hours or config.default_weekly_hours,
        base_work_days=int(expected_plan["base_work_days"]),
    )
    external_loan_hours = get_schedule_line_external_loan_hours(line, config=config)
    validation_metrics = build_schedule_validation_metrics(
        line,
        expected_plan=expected_plan,
        day_breakdown=day_breakdown,
        total_hours=total_worked_hours,
        credited_hours=decimal_hours(payment_resolution["payment_hours_used"]) + external_loan_hours,
        excluded_compensation_hours=decimal_hours(payment_resolution["excluded_company_day_hours"]),
        config=config,
    )
    movement_summary = summarize_schedule_day_movements(day_breakdown, payment_resolution["day_states"])
    status_value = str(validation_metrics["status"])
    difference_hours = decimal_hours(validation_metrics["difference_hours"])
    status_blocker_message = get_schedule_line_status_blocker_message(
        ScheduleLine(
            validation_status=status_value,
            weekly_hour_difference=difference_hours,
        )
    )
    tolerated_hour_difference = is_non_blocking_hour_difference(status_value, difference_hours)
    categories: list[str] = []
    daily_limit = line.daily_max_hours or config.default_daily_max_hours
    overtime_restriction = get_active_overtime_restriction(line.employee_identifier)
    daily_restriction_limit = (
        decimal_hours(overtime_restriction.max_daily_overtime_hours)
        if overtime_restriction is not None
        else Decimal("0.00")
    )

    if any(decimal_hours(day_info["worked_hours"]) > daily_limit for day_info in day_breakdown):
        categories.append("limites del dia")

    if overtime_restriction and any(
        get_daily_overtime_hours(
            day_info["worked_hours"],
            balance_snapshot["day_reference_hours"],
        ) > daily_restriction_limit
        for day_info in day_breakdown
    ):
        categories.append("restriccion medica")

    if (
        overtime_restriction
        and decimal_hours(line.overtime_hours) > decimal_hours(overtime_restriction.max_weekly_overtime_hours)
    ):
        categories.append("restriccion medica")

    if (
        payment_resolution["invalid_pay_day_indices"]
        or payment_resolution["invalid_pay_hours_indices"]
        or payment_resolution["invalid_pay_money_day_indices"]
        or payment_resolution["invalid_pay_money_indices"]
        or payment_resolution["invalid_advance_day_indices"]
        or payment_resolution["invalid_advance_day_with_balance_indices"]
        or payment_resolution["invalid_auto_rest_day_indices"]
        or payment_resolution["invalid_company_day_repayment_indices"]
    ):
        categories.append("saldo previo")

    if status_blocker_message:
        categories.append(status_blocker_message)
    elif validation_metrics["status"] in {
        ScheduleLine.ValidationStatus.INCOMPLETE,
        ScheduleLine.ValidationStatus.OVERPLANNED,
    } and not tolerated_hour_difference:
        categories.append("jornada semanal")
    elif validation_metrics["status"] in {
        ScheduleLine.ValidationStatus.IMPOSSIBLE,
        ScheduleLine.ValidationStatus.INCONSISTENT,
    }:
        categories.append("jornada semanal")

    if not categories:
        categories_text = "Sin alertas"
    else:
        unique_categories = list(dict.fromkeys(categories))
        if status_blocker_message:
            extra_categories = [
                category
                for category in unique_categories
                if category != status_blocker_message
            ]
            categories_text = status_blocker_message
            if extra_categories:
                categories_text += f" Revisa {', '.join(extra_categories)}."
        else:
            categories_text = f"Revisa {', '.join(unique_categories)}."

    movement_notes: list[str] = []
    if movement_summary["generated_sunday_days"] > 0:
        movement_notes.append(
            f"{movement_summary['generated_sunday_days']} dia(s) generado(s) por domingo trabajado"
        )
    if movement_summary["generated_holiday_days"] > 0:
        movement_notes.append(
            f"{movement_summary['generated_holiday_days']} dia(s) generado(s) por festivo trabajado"
        )
    if movement_summary["paid_days"] > 0:
        movement_notes.append(
            f"{movement_summary['paid_days']} dia(s) consumido(s) mediante Pago dia"
        )
    if movement_summary["advance_days"] > 0:
        movement_notes.append(
            f"{movement_summary['advance_days']} dia(s) registrado(s) como descanso adelantado"
        )
    if movement_summary["absence_days"] > 0:
        movement_notes.append(
            f"{movement_summary['absence_days']} dia(s) a favor de la empresa por inasistencia"
        )
    if movement_summary["company_day_repayments"] > 0:
        movement_notes.append(
            f"{movement_summary['company_day_repayments']} dia(s) compensado(s) a favor de la empresa"
        )

    summary_parts = [
        (
            "Resultado estimado: "
            f"{format_signed_day_balance(decimal_hours(payment_resolution['remaining_day_balance']))} "
            f"y {format_weekly_difference_compact(validation_metrics['difference_hours'])}."
        )
    ]
    if movement_notes:
        summary_parts.append(f"Movimientos: {'; '.join(movement_notes)}.")
    if external_loan_hours > Decimal("0.00"):
        summary_parts.append(f"Horas de prestamo en otra sede: {external_loan_hours} h.")
    unlinked_loan_days = [
        day_plan
        for day_plan in expected_plan.get("day_plans", [])
        if str(day_plan.get("expected_reason") or "") == "prestamo_sin_destino"
    ]
    if unlinked_loan_days:
        summary_parts.append(
            f"Prestamo sin sede destino: reduce {len(unlinked_loan_days)} dia(s) de la jornada y no mueve saldos."
        )
    if categories_text != "Sin alertas":
        summary_parts.append(categories_text)
    return " ".join(summary_parts)


def build_schedule_validation_metrics(
    line: ScheduleLine,
    *,
    expected_plan: dict[str, object],
    day_breakdown: list[dict[str, Decimal | date | str | int]],
    total_hours: Decimal,
    credited_hours: Decimal = Decimal("0.00"),
    excluded_compensation_hours: Decimal = Decimal("0.00"),
    config: SystemConfiguration | None = None,
) -> dict[str, object]:
    config = config or SystemConfiguration.load()
    daily_limit = decimal_hours(line.daily_max_hours or config.default_daily_max_hours or "0")
    day_plans = list(expected_plan.get("day_plans", []))
    adjusted_weekly_hours = decimal_hours(expected_plan.get("expected_weekly_hours", Decimal("0.00")) or "0")
    adjusted_weekly_exact_minutes = Decimal(
        str(expected_plan.get("expected_weekly_exact_minutes", Decimal("0")) or "0")
    )
    adjusted_weekly_programmable_minutes = Decimal(
        str(expected_plan.get("expected_weekly_programmable_minutes", Decimal("0")) or "0")
    )
    rounding_adjustment_minutes = Decimal(
        str(expected_plan.get("rounding_adjustment_minutes", Decimal("0")) or "0")
    )
    capacity_hours = Decimal("0.00")
    additional_rest_days = 0
    reducer_days = 0
    excluded_compensation_hours = decimal_hours(excluded_compensation_hours)

    for day_plan in day_plans:
        expected_reason = str(day_plan.get("expected_reason") or "")
        if expected_reason == "laborable":
            capacity_hours += daily_limit
        elif expected_reason in {
            "descanso_compensatorio",
            "descanso_adelantado",
            "descanso_adicional",
            "festivo_no_trabajado",
            "novedad_no_laborable",
            "prestamo_sin_destino",
            "inasistencia",
        }:
            reducer_days += 1
            if expected_reason in {
                "descanso_compensatorio",
                "descanso_adelantado",
                "descanso_adicional",
            }:
                additional_rest_days += 1

    capacity_hours = capacity_hours.quantize(TWO_DECIMALS)
    raw_total_hours = decimal_hours(total_hours)
    credited_hours = decimal_hours(credited_hours)
    effective_total_hours = max(
        raw_total_hours + credited_hours - excluded_compensation_hours,
        Decimal("0.00"),
    ).quantize(TWO_DECIMALS)
    difference_hours = (effective_total_hours - adjusted_weekly_hours).quantize(TWO_DECIMALS)
    mandatory_rest_index = int(expected_plan.get("mandatory_rest_index", 0) or 0)

    if adjusted_weekly_hours == Decimal("0.00") and decimal_hours(total_hours) == Decimal("0.00"):
        status = ScheduleLine.ValidationStatus.VALID
    elif daily_limit <= Decimal("0.00") and adjusted_weekly_hours > Decimal("0.00"):
        status = ScheduleLine.ValidationStatus.INCONSISTENT
    elif difference_hours == Decimal("0.00"):
        status = ScheduleLine.ValidationStatus.VALID
    elif difference_hours > Decimal("0.00"):
        status = ScheduleLine.ValidationStatus.OVERPLANNED
    elif capacity_hours >= adjusted_weekly_hours:
        status = ScheduleLine.ValidationStatus.INCOMPLETE
    else:
        status = ScheduleLine.ValidationStatus.IMPOSSIBLE

    return {
        "status": status,
        "capacity_hours": capacity_hours,
        "difference_hours": difference_hours,
        "raw_total_hours": raw_total_hours,
        "credited_hours": credited_hours,
        "effective_total_hours": effective_total_hours,
        "excluded_compensation_hours": excluded_compensation_hours,
        "adjusted_weekly_hours": adjusted_weekly_hours,
        "adjusted_weekly_exact_hours": minutes_to_decimal_hours(adjusted_weekly_exact_minutes),
        "adjusted_weekly_exact_minutes": adjusted_weekly_exact_minutes,
        "adjusted_weekly_exact_label": format_minutes_duration(adjusted_weekly_exact_minutes),
        "adjusted_weekly_programmable_minutes": adjusted_weekly_programmable_minutes,
        "adjusted_weekly_programmable_label": format_minutes_duration(adjusted_weekly_programmable_minutes),
        "rounding_adjustment_minutes": rounding_adjustment_minutes,
        "rounding_adjustment_label": format_minutes_duration(rounding_adjustment_minutes),
        "programming_interval_minutes": int(expected_plan.get("programming_interval_minutes", get_programming_interval_minutes(config=config))),
        "base_work_days": int(expected_plan.get("base_work_days", get_line_base_work_days(line, config=config))),
        "daily_equivalent_hours": decimal_hours(expected_plan.get("day_reference_hours", Decimal("0.00")) or "0"),
        "mandatory_rest_index": mandatory_rest_index,
        "mandatory_rest_label": SystemConfiguration.day_name(
            ((line.schedule.first_day_index if line.schedule_id else config.week_start_day) + mandatory_rest_index) % 7
        ),
        "additional_rest_days": additional_rest_days,
        "reducer_days": reducer_days,
    }


def recalculate_schedule_line(line: ScheduleLine) -> ScheduleLine:
    config = SystemConfiguration.load()
    shift_templates = get_selected_shift_templates(line)
    daily_limit = decimal_hours(line.daily_max_hours or config.default_daily_max_hours)
    weekly_target = decimal_hours(line.weekly_target_hours or config.default_weekly_hours)
    day_reference_hours = get_line_day_reference_hours(line, config=config)
    balance_snapshot = get_schedule_line_balance_snapshot(line, config=config)
    day_breakdown = build_line_day_breakdown(line, config=config, shift_templates=shift_templates)
    expected_plan = build_expected_week_plan(
        line,
        day_breakdown=day_breakdown,
        config=config,
        shift_templates=shift_templates,
    )
    compensation_entries = build_compensation_entries(
        line,
        day_breakdown=day_breakdown,
        expected_plan=expected_plan,
        config=config,
        shift_templates=shift_templates,
    )

    total_hours = Decimal("0.00")
    total_night_bonus = Decimal("0.00")
    warnings: list[str] = []
    overtime_restriction = get_active_overtime_restriction(line.employee_identifier)
    daily_restriction_limit = (
        decimal_hours(overtime_restriction.max_daily_overtime_hours)
        if overtime_restriction is not None
        else Decimal("0.00")
    )
    weekly_restriction_limit = (
        decimal_hours(overtime_restriction.max_weekly_overtime_hours)
        if overtime_restriction is not None
        else Decimal("0.00")
    )

    for day_info in day_breakdown:
        index = int(day_info["index"])
        daily_hours = decimal_hours(day_info["worked_hours"])
        daily_night_bonus = decimal_hours(day_info["night_hours"])
        setattr(line, f"day_{index}_hours", daily_hours)
        total_hours += daily_hours
        total_night_bonus += daily_night_bonus

        if daily_hours > daily_limit:
            warnings.append(f"Dia {index + 1}: supera el maximo diario ({daily_hours} h).")

        if overtime_restriction:
            daily_overtime_hours = get_daily_overtime_hours(daily_hours, day_reference_hours)
            if daily_overtime_hours > daily_restriction_limit:
                warnings.append(
                    f"Dia {index + 1}: restriccion medica supera el tope diario de extras "
                    f"({daily_overtime_hours} h vs {daily_restriction_limit} h)."
                )

        compensation_mode = day_info["compensation_mode"]
        compensation_hours = decimal_hours(day_info["compensation_hours"])
        is_special_day = bool(day_info["special_label"])

        if compensation_mode == ScheduleLine.CompensationMode.PAY_HOURS:
            compensated_day_hours = daily_hours + compensation_hours
            if compensation_hours <= Decimal("0.00"):
                warnings.append(f"Dia {index + 1}: pago horas requiere una cantidad mayor que cero.")
            elif daily_hours >= daily_limit:
                warnings.append(f"Dia {index + 1}: la jornada del dia ya esta completa y no requiere pago horas.")
            elif compensated_day_hours > daily_limit:
                warnings.append(f"Dia {index + 1}: pago horas supera la jornada del dia.")
        elif compensation_mode == ScheduleLine.CompensationMode.PAY_DAY and is_special_day:
            warnings.append(f"Dia {index + 1}: no se pueden pagar dias acumulados en domingo o festivo.")
        elif compensation_mode == COMPANY_DAY_REPAYMENT_MODE and compensation_hours > Decimal("0.00"):
            warnings.append(
                f"Dia {index + 1}: la compensacion del dia a favor de la empresa no utiliza horas parciales."
            )
        elif compensation_mode in MONEY_DAY_COMPENSATION_MODES:
            continue
        elif compensation_mode in MONEY_HOUR_COMPENSATION_MODES and compensation_hours <= Decimal("0.00"):
            warnings.append(f"Dia {index + 1}: pago en dinero por horas requiere una cantidad mayor que cero.")

    line.total_hours = total_hours.quantize(TWO_DECIMALS)
    line.expected_work_days = int(expected_plan["expected_work_days"])
    line.expected_weekly_hours = decimal_hours(expected_plan["expected_weekly_hours"])
    line.night_bonus_hours = total_night_bonus.quantize(TWO_DECIMALS)
    worked_total_hours = line.total_hours

    manual_day_adjustment = decimal_hours(line.manual_day_adjustment)
    manual_hour_adjustment = decimal_hours(line.manual_hour_adjustment)
    prior_day_balance = balance_snapshot["prior_day_balance"]
    prior_hour_balance = balance_snapshot["prior_hour_balance"]
    payment_resolution = resolve_compensation_usage(
        compensation_entries,
        available_day_balance=prior_day_balance + manual_day_adjustment,
        starting_day_balance=prior_day_balance,
        available_hour_balance=prior_hour_balance + manual_hour_adjustment,
        available_advance_pending_balance=Decimal("0.00"),
        day_reference_hours=day_reference_hours,
        weekly_target_hours=weekly_target,
        base_work_days=int(expected_plan["base_work_days"]),
    )
    external_loan_hours = get_schedule_line_external_loan_hours(line, config=config)
    validation_metrics = build_schedule_validation_metrics(
        line,
        expected_plan=expected_plan,
        day_breakdown=day_breakdown,
        total_hours=worked_total_hours,
        credited_hours=decimal_hours(payment_resolution["payment_hours_used"]) + external_loan_hours,
        excluded_compensation_hours=decimal_hours(payment_resolution["excluded_company_day_hours"]),
        config=config,
    )
    movement_summary = summarize_schedule_day_movements(day_breakdown, payment_resolution["day_states"])
    line.total_hours = (worked_total_hours + decimal_hours(payment_resolution["payment_hours_used"])).quantize(
        TWO_DECIMALS
    )
    line.weekly_hour_difference = decimal_hours(validation_metrics["difference_hours"])
    line.overtime_hours = decimal_hours(payment_resolution["generated_overtime_hours"])
    line.special_days_generated = int(payment_resolution["generated_special_days"])
    line.available_capacity_hours = decimal_hours(validation_metrics["capacity_hours"])
    line.validation_status = str(validation_metrics["status"])
    line.payment_days_used = int(payment_resolution["payment_days_used"])
    line.advance_rest_days_used = int(payment_resolution["advance_rest_days_used"])
    line.payment_hours_used = decimal_hours(payment_resolution["payment_hours_used"])
    line.money_payment_days_used = int(payment_resolution["money_payment_days_used"])
    line.money_payment_hours_used = decimal_hours(payment_resolution["money_payment_hours_used"])
    payment_hours = decimal_hours(payment_resolution["payment_hours_used"])
    money_payment_hours = decimal_hours(payment_resolution["money_payment_hours_used"])

    line.accrued_day_balance = decimal_hours(payment_resolution["remaining_day_balance"])
    line.accrued_hour_balance = max(
        (
            prior_hour_balance
            + line.overtime_hours
            + manual_hour_adjustment
            - payment_hours
            - money_payment_hours
        ),
        Decimal("0.00"),
    ).quantize(TWO_DECIMALS)
    line.advance_rest_pending_balance = max(line.accrued_day_balance * Decimal("-1.00"), Decimal("0.00")).quantize(
        TWO_DECIMALS
    )
    line.accrued_total_hours_balance = (
        line.accrued_hour_balance + (line.accrued_day_balance * day_reference_hours)
    ).quantize(TWO_DECIMALS)

    tolerated_hour_difference = is_non_blocking_hour_difference(
        line.validation_status,
        line.weekly_hour_difference,
    )
    effective_weekly_hours = decimal_hours(validation_metrics["effective_total_hours"])
    if effective_weekly_hours > weekly_target and not tolerated_hour_difference:
        warnings.append(f"Supera el objetivo semanal: {effective_weekly_hours} h vs {weekly_target} h.")

    if (
        overtime_restriction
        and line.overtime_hours > weekly_restriction_limit
    ):
        warnings.append(
            "Restriccion medica: no puede superar "
            f"{weekly_restriction_limit} h extra en la semana."
        )

    if payment_resolution["invalid_pay_day_indices"]:
        warnings.append(ADVANCE_REST_LIMIT_ERROR_MESSAGE)

    if payment_resolution["invalid_pay_money_day_indices"]:
        warnings.append("Se intento aplicar pago en dinero por dia sin un dia acumulado disponible.")

    if payment_resolution["invalid_advance_day_indices"] or payment_resolution["invalid_auto_rest_day_indices"]:
        warnings.append(ADVANCE_REST_LIMIT_ERROR_MESSAGE)

    if payment_resolution["invalid_pay_hours_indices"] or payment_resolution["invalid_pay_money_indices"]:
        warnings.append("Las horas descontadas superan el saldo acumulado disponible.")

    if payment_resolution["invalid_company_day_repayment_indices"]:
        repayment_messages = []
        for index in payment_resolution["invalid_company_day_repayment_indices"]:
            reason = str(payment_resolution["invalid_company_day_repayment_reasons"].get(index, ""))
            repayment_messages.append(f"Dia {int(index) + 1}: {get_company_day_repayment_error_message(reason)}")
        warnings.extend(repayment_messages)

    if (
        payment_resolution["invalid_pay_day_indices"]
        or payment_resolution["invalid_pay_money_day_indices"]
        or payment_resolution["invalid_advance_day_indices"]
        or payment_resolution["invalid_auto_rest_day_indices"]
        or payment_resolution["invalid_pay_hours_indices"]
        or payment_resolution["invalid_pay_money_indices"]
        or payment_resolution["invalid_company_day_repayment_indices"]
    ):
        line.validation_status = ScheduleLine.ValidationStatus.INCONSISTENT

    status_blocker_message = get_schedule_line_status_blocker_message(line)
    if status_blocker_message:
        warnings.append(status_blocker_message)
    elif line.validation_status == ScheduleLine.ValidationStatus.OVERPLANNED and not tolerated_hour_difference:
        warnings.append(
            "La programacion supera la jornada ajustada del cargo en "
            f"{line.weekly_hour_difference} h."
        )
    elif line.validation_status == ScheduleLine.ValidationStatus.INCOMPLETE and not tolerated_hour_difference:
        warnings.append(
            "La programacion esta "
            f"{abs(line.weekly_hour_difference)} h por debajo de la jornada ajustada y aun existe capacidad para completarla."
        )

    status_label = line.get_validation_status_display()
    if tolerated_hour_difference:
        status_label = "Valida con diferencia permitida"

    summary_parts = [
        f"Estado: {status_label}.",
        f"Jornada cargo: {weekly_target} h.",
        f"Dias base: {validation_metrics['base_work_days']}.",
        f"Equivalente diario: {validation_metrics['daily_equivalent_hours']} h.",
        f"Descanso obligatorio: {validation_metrics['mandatory_rest_label']}.",
        f"Descansos adicionales: {validation_metrics['additional_rest_days']}.",
        f"Jornada ajustada: {line.expected_weekly_hours} h.",
        f"Programadas: {line.total_hours} h.",
        f"Validadas para jornada: {validation_metrics['effective_total_hours']} h.",
        f"Diferencia: {line.weekly_hour_difference} h.",
        f"Capacidad disponible: {line.available_capacity_hours} h.",
        f"Saldo neto dias: {format_signed_day_balance(line.accrued_day_balance)}.",
        f"Saldo horas: {line.accrued_hour_balance} h.",
    ]
    if decimal_hours(validation_metrics["excluded_compensation_hours"]) > Decimal("0.00"):
        summary_parts.append(
            f"Horas excluidas por compensacion de deuda: {validation_metrics['excluded_compensation_hours']} h."
        )
    if payment_hours > Decimal("0.00"):
        summary_parts.append(f"Horas pagas aplicadas a jornada: {payment_hours} h.")
    if external_loan_hours > Decimal("0.00"):
        summary_parts.append(f"Horas de prestamo en otra sede: {external_loan_hours} h.")
    unlinked_loan_days = [
        day_plan
        for day_plan in expected_plan.get("day_plans", [])
        if str(day_plan.get("expected_reason") or "") == "prestamo_sin_destino"
    ]
    if unlinked_loan_days:
        summary_parts.append(
            f"Prestamo sin sede destino: reduce {len(unlinked_loan_days)} dia(s) de la jornada y no mueve saldos."
        )
    if int(payment_resolution["company_day_repayments_used"]) > 0:
        summary_parts.append(
            f"Dias compensados a favor de la empresa: {int(payment_resolution['company_day_repayments_used'])}."
        )
    movement_notes: list[str] = []
    if movement_summary["generated_sunday_days"] > 0:
        movement_notes.append(
            f"{movement_summary['generated_sunday_days']} dia(s) generado(s) por domingo trabajado"
        )
    if movement_summary["generated_holiday_days"] > 0:
        movement_notes.append(
            f"{movement_summary['generated_holiday_days']} dia(s) generado(s) por festivo trabajado"
        )
    if movement_summary["paid_days"] > 0:
        movement_notes.append(
            f"{movement_summary['paid_days']} dia(s) consumido(s) mediante Pago dia"
        )
    if movement_summary["advance_days"] > 0:
        movement_notes.append(
            f"{movement_summary['advance_days']} dia(s) registrado(s) como descanso adelantado"
        )
    if movement_summary["absence_days"] > 0:
        movement_notes.append(
            f"{movement_summary['absence_days']} dia(s) a favor de la empresa por inasistencia"
        )
    if movement_summary["company_day_repayments"] > 0:
        movement_notes.append(
            f"{movement_summary['company_day_repayments']} dia(s) compensado(s) a favor de la empresa"
        )
    if movement_notes:
        summary_parts.append(f"Movimientos: {'; '.join(movement_notes)}.")
    automatic_repayment_indexes = [int(index) for index in payment_resolution.get("automatic_repayment_indexes", [])]
    if automatic_repayment_indexes:
        automatic_labels = ", ".join(
            f"{(line.schedule.week_start_date + timedelta(days=index)):%d/%m/%Y}"
            for index in automatic_repayment_indexes
        )
        summary_parts.append(
            f"Compensacion automatica aplicada en: {automatic_labels}."
        )
    if validation_metrics["rounding_adjustment_minutes"] != Decimal("0"):
        summary_parts.extend(
            [
                f"Calculo exacto previo al redondeo: {validation_metrics['adjusted_weekly_exact_label']}.",
                f"Ajuste tecnico de redondeo: {validation_metrics['rounding_adjustment_label']}.",
            ]
        )
    summary_parts.extend(warnings)
    line.validation_summary = " ".join(summary_parts)
    line.warnings_count = len(warnings)
    return line


def build_schedule_balance_movement_idempotency_key(
    *,
    line: ScheduleLine,
    movement_date: date,
    movement_type: str,
    quantity_days: Decimal,
    quantity_hours: Decimal,
    description: str,
) -> str:
    raw_value = "|".join(
        [
            str(getattr(line.schedule, "pk", "")),
            str(getattr(line, "pk", "")),
            movement_date.isoformat(),
            movement_type,
            format(decimal_hours(quantity_days), "f"),
            format(decimal_hours(quantity_hours), "f"),
            description.strip(),
        ]
    )
    return hashlib.sha1(raw_value.encode("utf-8")).hexdigest()


def schedule_balance_movement_matches_payload(
    movement: ScheduleBalanceMovement,
    payload: dict[str, object],
) -> bool:
    return (
        movement.movement_date == payload["movement_date"]
        and movement.movement_type == payload["movement_type"]
        and decimal_hours(movement.quantity_days) == decimal_hours(payload["quantity_days"])
        and decimal_hours(movement.quantity_hours) == decimal_hours(payload["quantity_hours"])
        and decimal_hours(movement.equivalent_hours) == decimal_hours(payload["equivalent_hours"])
        and decimal_hours(movement.balance_before_days) == decimal_hours(payload["balance_before_days"])
        and decimal_hours(movement.balance_after_days) == decimal_hours(payload["balance_after_days"])
        and (movement.description or "") == str(payload["description"] or "")
    )


def build_schedule_balance_reversal_key(movement: ScheduleBalanceMovement) -> str:
    base_key = str(movement.idempotency_key or "").strip()
    if not base_key:
        base_key = f"legacy:{movement.pk}:{movement.movement_type}:{movement.movement_date:%Y%m%d}"
    return f"{base_key}:reversal"


def rebuild_schedule_line_movements(line: ScheduleLine) -> None:
    if not line.pk or not line.schedule_id:
        return

    config = SystemConfiguration.load()
    day_reference_hours = get_line_day_reference_hours(line, config=config)
    shift_templates = get_selected_shift_templates(line)
    day_breakdown = build_line_day_breakdown(line, config=config, shift_templates=shift_templates)
    expected_plan = build_expected_week_plan(
        line,
        day_breakdown=day_breakdown,
        config=config,
        shift_templates=shift_templates,
    )
    compensation_entries = build_compensation_entries(
        line,
        day_breakdown=day_breakdown,
        expected_plan=expected_plan,
        config=config,
        shift_templates=shift_templates,
    )
    balance_snapshot = get_schedule_line_balance_snapshot(line, config=config)
    payment_resolution = resolve_compensation_usage(
        compensation_entries,
        available_day_balance=balance_snapshot["prior_day_balance"] + decimal_hours(line.manual_day_adjustment),
        starting_day_balance=balance_snapshot["prior_day_balance"],
        available_hour_balance=balance_snapshot["prior_hour_balance"] + decimal_hours(line.manual_hour_adjustment),
        available_advance_pending_balance=Decimal("0.00"),
        day_reference_hours=day_reference_hours,
        weekly_target_hours=line.weekly_target_hours or config.default_weekly_hours,
        base_work_days=int(expected_plan["base_work_days"]),
    )
    movement_date_default = line.schedule.week_end_date or line.schedule.week_start_date
    movement_start_date = line.schedule.week_start_date
    desired_movements: list[dict[str, object]] = []
    day_states = payment_resolution["day_states"]

    manual_day_adjustment = decimal_hours(line.manual_day_adjustment)
    if manual_day_adjustment != Decimal("0.00"):
        desired_movements.append(
            {
                "movement_date": movement_start_date,
                "movement_type": ScheduleBalanceMovement.MovementType.MANUAL_DAY,
                "quantity_days": manual_day_adjustment,
                "quantity_hours": Decimal("0.00"),
                "equivalent_hours": (manual_day_adjustment * day_reference_hours).quantize(TWO_DECIMALS),
                "description": "Ajuste manual de dias",
            }
        )

    manual_hour_adjustment = decimal_hours(line.manual_hour_adjustment)
    if manual_hour_adjustment != Decimal("0.00"):
        desired_movements.append(
            {
                "movement_date": movement_start_date,
                "movement_type": ScheduleBalanceMovement.MovementType.MANUAL_HOUR,
                "quantity_days": Decimal("0.00"),
                "quantity_hours": manual_hour_adjustment,
                "equivalent_hours": manual_hour_adjustment,
                "description": "Ajuste manual de horas",
            }
        )

    for day_info in day_breakdown:
        day_date = day_info["date"]
        worked_hours = decimal_hours(day_info["worked_hours"])
        special_label = str(day_info["special_label"] or "")
        day_state = day_states.get(int(day_info["index"]), {})

        if bool(day_state.get("generated_day")) and worked_hours > Decimal("0.00") and special_label:
            desired_movements.append(
                {
                    "movement_date": day_date,
                    "movement_type": ScheduleBalanceMovement.MovementType.SPECIAL_DAY,
                    "quantity_days": Decimal("1.00"),
                    "quantity_hours": Decimal("0.00"),
                    "equivalent_hours": day_reference_hours,
                    "description": f"{special_label} laborado",
                }
            )

        day_movement_type = str(day_state.get("day_movement_type") or "")
        applied_day_delta = decimal_hours(day_state.get("applied_day_delta", Decimal("0.00")) or "0")
        if day_movement_type and applied_day_delta != Decimal("0.00"):
            desired_movements.append(
                {
                    "movement_date": day_date,
                    "movement_type": day_movement_type,
                    "quantity_days": applied_day_delta,
                    "quantity_hours": Decimal("0.00"),
                    "equivalent_hours": (applied_day_delta * day_reference_hours).quantize(TWO_DECIMALS),
                    "description": str(day_state.get("day_movement_description") or ""),
                }
            )

        hours_movement_type = str(day_state.get("hours_movement_type") or "")
        applied_hours = decimal_hours(day_state.get("applied_hours", Decimal("0.00")) or "0")
        if hours_movement_type == ScheduleBalanceMovement.MovementType.PAY_HOURS and applied_hours != Decimal("0.00"):
            desired_movements.append(
                {
                    "movement_date": day_date,
                    "movement_type": ScheduleBalanceMovement.MovementType.PAY_HOURS,
                    "quantity_days": Decimal("0.00"),
                    "quantity_hours": (applied_hours * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    "equivalent_hours": (applied_hours * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    "description": str(day_state.get("hours_movement_description") or "Pago con horas"),
                }
            )
        elif (
            hours_movement_type == ScheduleBalanceMovement.MovementType.PAY_MONEY_HOURS
            and applied_hours != Decimal("0.00")
        ):
            desired_movements.append(
                {
                    "movement_date": day_date,
                    "movement_type": ScheduleBalanceMovement.MovementType.PAY_MONEY_HOURS,
                    "quantity_days": Decimal("0.00"),
                    "quantity_hours": (applied_hours * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    "equivalent_hours": (applied_hours * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    "description": str(day_state.get("hours_movement_description") or "Pago en dinero por horas"),
                }
            )

    if line.overtime_hours > Decimal("0.00"):
        desired_movements.append(
            {
                "movement_date": movement_date_default,
                "movement_type": ScheduleBalanceMovement.MovementType.OVERTIME,
                "quantity_days": Decimal("0.00"),
                "quantity_hours": line.overtime_hours,
                "equivalent_hours": line.overtime_hours,
                "description": "Horas extras de la semana",
            }
        )

    running_day_balance = decimal_hours(balance_snapshot["prior_day_balance"])
    active_payloads: list[dict[str, object]] = []
    for payload in desired_movements:
        quantity_days = decimal_hours(payload.get("quantity_days", Decimal("0.00")) or "0")
        balance_before_days = running_day_balance
        if quantity_days != Decimal("0.00"):
            running_day_balance = (running_day_balance + quantity_days).quantize(TWO_DECIMALS)
        payload["balance_before_days"] = balance_before_days.quantize(TWO_DECIMALS)
        payload["balance_after_days"] = running_day_balance.quantize(TWO_DECIMALS)
        payload["idempotency_key"] = build_schedule_balance_movement_idempotency_key(
            line=line,
            movement_date=payload["movement_date"],
            movement_type=str(payload["movement_type"]),
            quantity_days=quantity_days,
            quantity_hours=decimal_hours(payload.get("quantity_hours", Decimal("0.00")) or "0"),
            description=str(payload.get("description") or ""),
        )
        active_payloads.append(payload)

    existing_active_movements = list(
        line.balance_movements.filter(is_reversal=False, is_reversed=False).order_by("pk")
    )
    existing_active_by_key: dict[str, list[ScheduleBalanceMovement]] = {}
    for movement in existing_active_movements:
        movement_key = str(movement.idempotency_key or "").strip()
        existing_active_by_key.setdefault(movement_key, []).append(movement)
    movements_to_create: list[ScheduleBalanceMovement] = []
    movements_to_reverse: list[ScheduleBalanceMovement] = []
    movements_to_mark_reversed: list[ScheduleBalanceMovement] = []
    recorded_by = line.schedule.updated_by or line.schedule.created_by
    matched_existing_ids: set[int] = set()

    for payload in active_payloads:
        idempotency_key = str(payload["idempotency_key"])
        existing = next(
            (
                candidate
                for candidate in existing_active_by_key.get(idempotency_key, [])
                if candidate.pk not in matched_existing_ids
                and schedule_balance_movement_matches_payload(candidate, payload)
            ),
            None,
        )
        if existing is not None:
            matched_existing_ids.add(existing.pk)
            continue

        movements_to_create.append(
            ScheduleBalanceMovement(
                schedule=line.schedule,
                line=line,
                site=line.schedule.site,
                employee_identifier=line.employee_identifier,
                employee_name=line.employee_name,
                job_role_name=line.job_role_name,
                movement_date=payload["movement_date"],
                movement_type=payload["movement_type"],
                quantity_days=payload["quantity_days"],
                quantity_hours=payload["quantity_hours"],
                equivalent_hours=payload["equivalent_hours"],
                balance_before_days=payload["balance_before_days"],
                balance_after_days=payload["balance_after_days"],
                movement_origin="horario",
                idempotency_key=idempotency_key,
                recorded_by=recorded_by,
                description=str(payload["description"] or ""),
            )
        )

    for existing in existing_active_movements:
        if existing.pk in matched_existing_ids:
            continue
        if existing.is_reversed:
            continue
        reversal_key = build_schedule_balance_reversal_key(existing)
        if not line.balance_movements.filter(idempotency_key=reversal_key, is_reversal=True).exists():
            movements_to_reverse.append(
                ScheduleBalanceMovement(
                    schedule=line.schedule,
                    line=line,
                    site=line.schedule.site,
                    employee_identifier=line.employee_identifier,
                    employee_name=line.employee_name,
                    job_role_name=line.job_role_name,
                    movement_date=movement_date_default,
                    movement_type=ScheduleBalanceMovement.MovementType.REVERSAL,
                    quantity_days=(decimal_hours(existing.quantity_days) * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    quantity_hours=(decimal_hours(existing.quantity_hours) * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    equivalent_hours=(decimal_hours(existing.equivalent_hours) * Decimal("-1.00")).quantize(TWO_DECIMALS),
                    balance_before_days=decimal_hours(existing.balance_after_days),
                    balance_after_days=decimal_hours(existing.balance_before_days),
                    movement_origin="recalculo_horario",
                    idempotency_key=reversal_key,
                    is_reversal=True,
                    reversed_movement=existing,
                    recorded_by=recorded_by,
                    description=f"Reverso de {get_schedule_balance_movement_label(existing.movement_type).lower()}",
                )
            )
        existing.is_reversed = True
        movements_to_mark_reversed.append(existing)

    if movements_to_mark_reversed:
        ScheduleBalanceMovement.objects.bulk_update(movements_to_mark_reversed, ["is_reversed"])
    if movements_to_reverse:
        ScheduleBalanceMovement.objects.bulk_create(movements_to_reverse)
    if movements_to_create:
        ScheduleBalanceMovement.objects.bulk_create(movements_to_create)


def save_schedule_line_with_balances(line: ScheduleLine) -> ScheduleLine:
    recalculate_schedule_line(line)
    line.save()
    rebuild_schedule_line_movements(line)
    return line


def rebuild_balances_for_employee_from_earliest_schedule(employee_identifier: str) -> bool:
    cleaned_identifier = (employee_identifier or "").strip()
    if not cleaned_identifier:
        return False

    first_line = (
        ScheduleLine.objects.select_related("schedule")
        .filter(employee_identifier=cleaned_identifier)
        .order_by("schedule__week_start_date", "pk")
        .first()
    )
    if first_line is None or first_line.schedule is None:
        return False

    rebuild_balances_for_employees_from_week(
        first_line.schedule.week_start_date,
        [cleaned_identifier],
    )
    return True


def release_schedule_balance_reversal_links(schedule: WeeklySchedule) -> int:
    movement_ids = list(schedule.balance_movements.values_list("pk", flat=True))
    if not movement_ids:
        return 0

    return ScheduleBalanceMovement.objects.filter(
        reversed_movement_id__in=movement_ids,
    ).update(reversed_movement=None)


def release_schedule_lines_balance_reversal_links(lines) -> int:
    line_ids = list(lines.values_list("pk", flat=True))
    if not line_ids:
        return 0

    movement_ids = list(
        ScheduleBalanceMovement.objects.filter(line_id__in=line_ids).values_list("pk", flat=True)
    )
    if not movement_ids:
        return 0

    return ScheduleBalanceMovement.objects.filter(
        reversed_movement_id__in=movement_ids,
    ).update(reversed_movement=None)


def release_schedule_line_balance_reversal_links(line: ScheduleLine) -> int:
    return release_schedule_lines_balance_reversal_links(
        ScheduleLine.objects.filter(pk=line.pk)
    )


def rebuild_balances_for_employees_from_week(
    week_start_date: date,
    employee_identifiers: list[str] | set[str] | tuple[str, ...] | None = None,
) -> None:
    queryset = ScheduleLine.objects.select_related("schedule").filter(
        schedule__week_start_date__gte=week_start_date,
    )
    if employee_identifiers:
        queryset = queryset.filter(employee_identifier__in=set(employee_identifiers))

    ordered_lines = sorted(
        list(queryset),
        key=lambda line: (
            (line.employee_identifier or "").strip(),
            *get_schedule_line_progression_key(line),
        ),
    )
    for line in ordered_lines:
        save_schedule_line_with_balances(line)


def get_latest_schedule_lines_by_employee(
    employee_identifiers: list[str] | set[str] | tuple[str, ...] | None = None,
) -> dict[str, ScheduleLine]:
    queryset = ScheduleLine.objects.select_related("schedule", "schedule__site")
    if employee_identifiers:
        queryset = queryset.filter(employee_identifier__in=set(employee_identifiers))

    latest_by_employee: dict[str, tuple[tuple[object, ...], ScheduleLine]] = {}
    for line in queryset:
        cleaned_identifier = (line.employee_identifier or "").strip()
        if not cleaned_identifier:
            continue
        candidate_key = get_schedule_line_progression_key(line)
        current_entry = latest_by_employee.get(cleaned_identifier)
        if current_entry is None or candidate_key > current_entry[0]:
            latest_by_employee[cleaned_identifier] = (candidate_key, line)

    return {
        employee_identifier: entry[1]
        for employee_identifier, entry in latest_by_employee.items()
    }


def build_schedule_balance_audit_rows(
    employee_identifiers: list[str] | set[str] | tuple[str, ...] | None = None,
) -> list[dict[str, object]]:
    latest_lines = get_latest_schedule_lines_by_employee(employee_identifiers)
    if not latest_lines:
        return []

    identifiers = list(latest_lines.keys())
    initial_balances = {
        balance.employee_identifier: balance
        for balance in EmployeeInitialBalance.objects.filter(employee_identifier__in=identifiers)
    }
    movement_totals = {
        row["employee_identifier"]: {
            "days": decimal_hours(row["total_days"]),
            "hours": decimal_hours(row["total_hours"]),
        }
        for row in (
            ScheduleBalanceMovement.objects.filter(
                employee_identifier__in=identifiers,
                is_reversal=False,
                is_reversed=False,
            )
            .values("employee_identifier")
            .annotate(
                total_days=Coalesce(
                    Sum("quantity_days"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=8, decimal_places=2)),
                ),
                total_hours=Coalesce(
                    Sum("quantity_hours"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=10, decimal_places=2)),
                ),
            )
        )
    }

    rows: list[dict[str, object]] = []
    for employee_identifier, line in latest_lines.items():
        initial_balance = initial_balances.get(employee_identifier)
        totals = movement_totals.get(
            employee_identifier,
            {"days": Decimal("0.00"), "hours": Decimal("0.00")},
        )
        initial_day_balance = decimal_hours(
            getattr(initial_balance, "initial_day_balance", Decimal("0.00"))
        )
        initial_hour_balance = decimal_hours(
            getattr(initial_balance, "initial_hour_balance", Decimal("0.00"))
        )
        audited_day_balance = (initial_day_balance + totals["days"]).quantize(TWO_DECIMALS)
        audited_hour_balance = (initial_hour_balance + totals["hours"]).quantize(TWO_DECIMALS)
        stored_day_balance = decimal_hours(line.accrued_day_balance)
        stored_hour_balance = decimal_hours(line.accrued_hour_balance)
        day_difference = (stored_day_balance - audited_day_balance).quantize(TWO_DECIMALS)
        hour_difference = (stored_hour_balance - audited_hour_balance).quantize(TWO_DECIMALS)

        rows.append(
            {
                "site_code": getattr(line.schedule.site, "code", ""),
                "site_name": getattr(line.schedule.site, "name", ""),
                "job_role_name": line.job_role_name,
                "employee_identifier": employee_identifier,
                "employee_name": line.employee_name,
                "audited_day_balance": audited_day_balance,
                "stored_day_balance": stored_day_balance,
                "day_difference": day_difference,
                "audited_hour_balance": audited_hour_balance,
                "stored_hour_balance": stored_hour_balance,
                "hour_difference": hour_difference,
                "has_difference": (
                    day_difference != Decimal("0.00")
                    or hour_difference != Decimal("0.00")
                ),
            }
        )

    return sorted(
        rows,
        key=lambda row: (
            str(row["site_code"] or ""),
            str(row["job_role_name"] or "").casefold(),
            str(row["employee_name"] or "").casefold(),
            str(row["employee_identifier"] or ""),
        ),
    )


def purge_blacklisted_lines_from_schedule(schedule: WeeklySchedule) -> list[str]:
    if schedule_accepts_blacklisted_staff(schedule):
        return []

    blacklisted_identifiers = get_blacklisted_employee_identifiers(
        schedule.lines.values_list("employee_identifier", flat=True)
    )
    if not blacklisted_identifiers:
        return []

    blacklisted_lines = schedule.lines.filter(employee_identifier__in=blacklisted_identifiers)
    release_schedule_lines_balance_reversal_links(blacklisted_lines)
    blacklisted_lines.delete()
    rebuild_balances_for_employees_from_week(schedule.week_start_date, list(blacklisted_identifiers))
    return sorted(blacklisted_identifiers)


def copy_schedule_template(source_schedule: WeeklySchedule, target_schedule: WeeklySchedule) -> tuple[int, int]:
    if source_schedule.pk == target_schedule.pk:
        return 0, 0

    blacklisted_identifiers = set()
    if not schedule_accepts_blacklisted_staff(target_schedule):
        blacklisted_identifiers = get_blacklisted_employee_identifiers(
            source_schedule.lines.values_list("employee_identifier", flat=True)
        )
    source_lines = {
        (line.employee_identifier or "").strip(): line
        for line in source_schedule.lines.all()
        if (line.employee_identifier or "").strip() not in blacklisted_identifiers
    }
    copied_count = 0
    touched_employee_identifiers: list[str] = []

    for line in target_schedule.lines.all():
        employee_identifier = (line.employee_identifier or "").strip()
        if employee_identifier in blacklisted_identifiers:
            continue
        source_line = source_lines.get(employee_identifier)
        if source_line is None:
            continue

        for index in range(7):
            setattr(line, f"day_{index}_shift_1", getattr(source_line, f"day_{index}_shift_1", "") or "")
            setattr(line, f"day_{index}_shift_2", getattr(source_line, f"day_{index}_shift_2", "") or "")
            setattr(line, f"day_{index}_compensation_mode", "")
            setattr(line, f"day_{index}_compensation_hours", Decimal("0.00"))

        line.save()
        copied_count += 1
        touched_employee_identifiers.append(employee_identifier)

    if touched_employee_identifiers:
        rebuild_balances_for_employees_from_week(target_schedule.week_start_date, touched_employee_identifiers)

    return copied_count, len(source_lines)


def sync_schedule_from_legacy(schedule: WeeklySchedule) -> tuple[int, int]:
    config = SystemConfiguration.load()
    purge_blacklisted_lines_from_schedule(schedule)
    if schedule.site.is_personal_vario:
        staff = build_personal_vario_staff_records()
        blacklisted_identifiers: set[str] = set()
    else:
        staff = fetch_active_staff_for_site(
            schedule.site.code,
            week_start_date=schedule.week_start_date,
        )
        blacklisted_identifiers = get_blacklisted_employee_identifiers(
            [employee.employee_id for employee in staff]
        )
    existing_lines = {
        line.employee_identifier: line for line in schedule.lines.all()
    }
    created_count = 0
    updated_count = 0
    touched_employee_identifiers: list[str] = []

    for employee in staff:
        if employee.employee_id in blacklisted_identifiers:
            continue
        job_role, _ = JobRole.objects.get_or_create(
            name=employee.role_name or "PERSONAL VARIO",
            defaults={
                "code": employee.role_code,
                "weekly_target_hours": config.default_weekly_hours,
                "daily_max_hours": config.default_daily_max_hours,
                "base_work_days": config.default_base_work_days,
            },
        )

        line = existing_lines.get(employee.employee_id)
        line_is_new = line is None
        if line is None:
            line = ScheduleLine(
                schedule=schedule,
                employee_identifier=employee.employee_id,
            )
            created_count += 1
        else:
            updated_count += 1

        line.employee_name = employee.employee_name
        line.department_code = employee.department_code
        line.department_name = employee.department_name
        role_snapshot_missing = not line_has_frozen_workload_snapshot(line)
        if schedule.site.is_personal_vario:
            if not (line.job_role_name or "").strip() or (line.job_role_name or "").strip().upper() == "PERSONAL VARIO":
                apply_job_role_snapshot(
                    line,
                    job_role,
                    config,
                    overwrite_workload=line_is_new or role_snapshot_missing,
                    role_code=employee.role_code,
                    role_name=employee.role_name,
                )
            else:
                line.job_role_code = line.job_role_code or employee.role_code or job_role.code
                line.weekly_target_hours = line.weekly_target_hours or job_role.weekly_target_hours
                line.daily_max_hours = line.daily_max_hours or job_role.daily_max_hours
                line.base_work_days = line.base_work_days or job_role.base_work_days or config.default_base_work_days
        else:
            if line_is_new or role_snapshot_missing or not (line.job_role_name or "").strip():
                apply_job_role_snapshot(
                    line,
                    job_role,
                    config,
                    overwrite_workload=line_is_new or role_snapshot_missing,
                    role_code=employee.role_code,
                    role_name=employee.role_name,
                )
        line.save()
        touched_employee_identifiers.append(line.employee_identifier)

    if touched_employee_identifiers:
        rebuild_balances_for_employees_from_week(schedule.week_start_date, touched_employee_identifiers)

    return created_count, updated_count
