from __future__ import annotations
from dataclasses import dataclass
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import re

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.access import get_accessible_schedules_queryset
from core.models import ShiftTemplate
from schedules.models import ScheduleBalanceMovement, ScheduleLine, WeeklySchedule
from schedules.services import (
    build_line_day_breakdown,
    build_schedule_flat_file_headers,
    get_schedule_line_activity_indices,
    get_schedule_line_progression_key,
)

SPANISH_MONTH_NAMES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
SHIFT_RANGE_PATTERN = re.compile(r"^(?P<start>\d{1,2}:\d{2})-(?P<end>\d{1,2}:\d{2})$")
HOURLY_COVERAGE_START_HOUR = 6
HOURLY_COVERAGE_END_HOUR = 21


@dataclass(slots=True)
class InventoryReportEntry:
    site_code: str
    site_name: str
    inventory_date: date
    employee_identifier: str
    employee_name: str
    job_role_name: str


def build_excel_response(title: str, headers: list[str], rows: list[list[object]], filename: str) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(row)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 38)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def excel_number(value) -> int | float:
    decimal_value = Decimal(str(value or "0"))
    if decimal_value == decimal_value.to_integral():
        return int(decimal_value)
    return float(decimal_value)


def excel_number_or_blank(value):
    decimal_value = Decimal(str(value or "0"))
    if decimal_value == Decimal("0"):
        return ""
    return excel_number(decimal_value)


def workbook_to_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def workbook_to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_schedule_excel_response(schedule: WeeklySchedule) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Horario"

    day_columns = schedule.get_day_columns()
    worksheet["A1"] = f"Horario de {schedule.site.name}"
    worksheet["A2"] = f"Semana del {schedule.week_start_date:%d/%m/%Y} al {schedule.week_end_date:%d/%m/%Y}"
    worksheet["A3"] = f"Estado: {schedule.get_status_display()}"
    worksheet["A4"] = f"Notas: {schedule.notes or 'Sin notas'}"
    total_columns = 3 + (len(day_columns) * 3) + 5
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_columns)

    for cell_ref in ("A1", "A2", "A3", "A4"):
        worksheet[cell_ref].font = Font(bold=True)
        worksheet[cell_ref].alignment = Alignment(horizontal="left")

    header_group_row = 6
    header_detail_row = 7

    worksheet.cell(row=header_group_row, column=1, value="Cedula")
    worksheet.cell(row=header_group_row, column=2, value="Empleado")
    worksheet.cell(row=header_group_row, column=3, value="Cargo")
    worksheet.merge_cells(start_row=header_group_row, start_column=1, end_row=header_detail_row, end_column=1)
    worksheet.merge_cells(start_row=header_group_row, start_column=2, end_row=header_detail_row, end_column=2)
    worksheet.merge_cells(start_row=header_group_row, start_column=3, end_row=header_detail_row, end_column=3)

    current_column = 4
    for column in day_columns:
        worksheet.cell(
            row=header_group_row,
            column=current_column,
            value=f"{column['label']}\n{column['date']}",
        )
        worksheet.merge_cells(
            start_row=header_group_row,
            start_column=current_column,
            end_row=header_group_row,
            end_column=current_column + 2,
        )
        worksheet.cell(row=header_detail_row, column=current_column, value="Turno 1")
        worksheet.cell(row=header_detail_row, column=current_column + 1, value="Turno 2")
        worksheet.cell(row=header_detail_row, column=current_column + 2, value="Horas")
        current_column += 3

    summary_headers = ["Total", "Extras", "Rec. noct.", "Dias acum.", "Horas acum."]
    for summary_header in summary_headers:
        worksheet.cell(row=header_group_row, column=current_column, value=summary_header)
        worksheet.merge_cells(
            start_row=header_group_row,
            start_column=current_column,
            end_row=header_detail_row,
            end_column=current_column,
        )
        current_column += 1

    for row_number in (header_group_row, header_detail_row):
        for cell in worksheet[row_number]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ordered_lines = schedule.lines.all().order_by("job_role_name", "employee_name", "employee_identifier")
    for line in ordered_lines:
        row = [
            line.employee_identifier,
            line.employee_name,
            line.job_role_name,
        ]
        for index in range(7):
            row.extend(
                [
                    getattr(line, f"day_{index}_shift_1", "") or "",
                    getattr(line, f"day_{index}_shift_2", "") or "",
                    excel_number(getattr(line, f"day_{index}_hours", Decimal("0.00"))),
                ]
            )
        row.extend(
            [
                excel_number(line.total_hours),
                excel_number(line.overtime_hours),
                excel_number(line.night_bonus_hours),
                excel_number(line.accrued_day_balance),
                excel_number(line.accrued_hour_balance),
            ]
        )
        worksheet.append(row)

    for column_index, column_cells in enumerate(
        worksheet.iter_cols(min_row=header_group_row, max_row=worksheet.max_row),
        start=1,
    ):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 26)

    worksheet.freeze_panes = "D8"
    worksheet.sheet_view.zoomScale = 85
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1

    filename = f"horario_{schedule.site.code}_{schedule.week_start_date:%Y%m%d}.xlsx"
    return workbook_to_response(workbook, filename)


def build_schedule_flat_file_template_response(schedule: WeeklySchedule) -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PlantillaHorario"

    headers = build_schedule_flat_file_headers(schedule)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    compensation_labels = dict(ScheduleLine.CompensationMode.choices)
    ordered_lines = schedule.lines.all().order_by("job_role_name", "employee_name", "employee_identifier")
    for line in ordered_lines:
        row = [
            line.employee_identifier,
            line.employee_name,
            line.job_role_name,
        ]
        for index in range(7):
            compensation_mode = getattr(line, f"day_{index}_compensation_mode", "") or ""
            row.extend(
                [
                    getattr(line, f"day_{index}_shift_1", "") or "",
                    getattr(line, f"day_{index}_shift_2", "") or "",
                    compensation_labels.get(compensation_mode, "") if compensation_mode else "",
                    excel_number_or_blank(getattr(line, f"day_{index}_compensation_hours", Decimal("0.00"))),
                    "Si" if bool(getattr(line, f"day_{index}_inventory", False)) else "",
                ]
            )
        row.extend(
            [
                excel_number_or_blank(line.manual_day_adjustment),
                excel_number_or_blank(line.manual_hour_adjustment),
            ]
        )
        worksheet.append(row)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 26)

    info_sheet = workbook.create_sheet("Catalogos")
    info_sheet.append(["Tipo", "Valor"])
    info_sheet.append(["Instruccion", "Edita la hoja PlantillaHorario y vuelve a cargar el archivo desde el portal."])
    info_sheet.append(["Instruccion", "Los modos de pago admitidos son: Sin pago, Pago dia, Pago horas, Descanso adelantado, Pago dinero dia y Pago dinero horas."])
    info_sheet.append(["Instruccion", "La columna Inventario admite Si o vacio."])
    info_sheet.append([])
    info_sheet.append(["Turnos disponibles", ""])
    for shift in ShiftTemplate.objects.filter(is_active=True).order_by("display_order", "label"):
        info_sheet.append(["Turno", shift.label])

    for cell in info_sheet[1]:
        cell.font = Font(bold=True)
    for column_cells in info_sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        info_sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 16), 60)

    filename = f"plantilla_horario_{schedule.site.code}_{schedule.week_start_date:%Y%m%d}.xlsx"
    return workbook_to_response(workbook, filename)


def build_initial_balance_template_response() -> HttpResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plantilla"
    worksheet.append(["Cedula", "Nombres y apellidos", "Dias extras", "Horas extras"])
    worksheet.append(["1000123456", "Empleado Ejemplo", 2, 4.5])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 24)

    return workbook_to_response(workbook, "plantilla_saldos_iniciales.xlsx")


def shift_label_to_minutes(label: str, shift_templates: dict[str, ShiftTemplate]) -> tuple[int, int] | None:
    normalized = str(label or "").strip()
    if not normalized:
        return None

    template = shift_templates.get(normalized)
    if template is not None:
        if not (template.counts_as_worked_time and template.start_time and template.end_time):
            return None
        start_minutes = template.start_time.hour * 60 + template.start_time.minute
        end_minutes = template.end_time.hour * 60 + template.end_time.minute
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60
        return start_minutes, end_minutes

    match = SHIFT_RANGE_PATTERN.match(normalized)
    if not match:
        return None
    start_text = match.group("start")
    end_text = match.group("end")
    start_hour, start_minute = [int(value) for value in start_text.split(":")]
    end_hour, end_minute = [int(value) for value in end_text.split(":")]
    start_minutes = (start_hour * 60) + start_minute
    end_minutes = (end_hour * 60) + end_minute
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    return start_minutes, end_minutes


def build_hourly_coverage_schedule_workbook(schedule: WeeklySchedule) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    summary_sheet.append(["Reporte", "Cobertura por franjas horarias"])
    summary_sheet.append(["Sede", schedule.site.name])
    summary_sheet.append(["Semana inicio", schedule.week_start_date.strftime("%d/%m/%Y")])
    summary_sheet.append(["Semana fin", schedule.week_end_date.strftime("%d/%m/%Y")])
    summary_sheet.append(["Estado", schedule.get_status_display()])
    summary_sheet.append(["Rango horario", f"{HOURLY_COVERAGE_START_HOUR:02d}:00 a {HOURLY_COVERAGE_END_HOUR:02d}:00"])
    for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
        row[0].font = Font(bold=True)
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 40

    lines = list(schedule.lines.all().order_by("job_role_name", "employee_name", "employee_identifier"))
    role_names = sorted(
        {
            ((line.job_role_name or "").strip() or "SIN CARGO")
            for line in lines
        },
        key=str.casefold,
    )
    selected_labels = {
        str(getattr(line, f"day_{index}_shift_{slot}", "") or "").strip()
        for line in lines
        for index in range(7)
        for slot in (1, 2)
        if str(getattr(line, f"day_{index}_shift_{slot}", "") or "").strip()
    }
    shift_templates = {
        shift.label: shift
        for shift in ShiftTemplate.objects.filter(label__in=selected_labels, is_active=True)
    }

    coverage_counts: dict[tuple[int, str, int], int] = defaultdict(int)
    for line in lines:
        role_name = ((line.job_role_name or "").strip() or "SIN CARGO")
        for day_index in range(7):
            covered_hours: set[int] = set()
            for slot in (1, 2):
                interval = shift_label_to_minutes(
                    getattr(line, f"day_{day_index}_shift_{slot}", "") or "",
                    shift_templates,
                )
                if interval is None:
                    continue
                start_minutes, end_minutes = interval
                for hour in range(HOURLY_COVERAGE_START_HOUR, HOURLY_COVERAGE_END_HOUR):
                    slot_start = hour * 60
                    slot_end = (hour + 1) * 60
                    if min(end_minutes, slot_end) > max(start_minutes, slot_start):
                        covered_hours.add(hour)
            for hour in covered_hours:
                coverage_counts[(day_index, role_name, hour)] += 1

    for day_index, column in enumerate(schedule.get_day_columns()):
        sheet = workbook.create_sheet(f"{column['label'][:20]}")
        title = f"{column['label']} {column['date']} - {schedule.site.name}"
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(role_names) + 1, 2))
        headers = ["Franja horaria", *role_names]
        sheet.append(headers)
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for hour in range(HOURLY_COVERAGE_START_HOUR, HOURLY_COVERAGE_END_HOUR):
            row = [f"{hour:02d}:00-{hour + 1:02d}:00"]
            for role_name in role_names:
                row.append(coverage_counts.get((day_index, role_name, hour), 0))
            sheet.append(row)

        for column_index, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 14), 34)
        sheet.freeze_panes = "A3"

    return workbook


def build_hourly_coverage_excel_response(schedule: WeeklySchedule) -> HttpResponse:
    workbook = build_hourly_coverage_schedule_workbook(schedule)
    filename = f"cobertura_horaria_{schedule.site.code}_{schedule.week_start_date:%Y%m%d}.xlsx"
    return workbook_to_response(workbook, filename)


def build_hourly_coverage_report_attachment(schedule: WeeklySchedule) -> tuple[str, bytes]:
    workbook = build_hourly_coverage_schedule_workbook(schedule)
    filename = f"cobertura_horaria_{schedule.site.code}_{schedule.week_start_date:%Y%m%d}.xlsx"
    return filename, workbook_to_bytes(workbook)


def get_accessible_schedule_queryset_for_range(user, date_from: date, date_to: date, site=None):
    queryset = WeeklySchedule.objects.select_related("site").prefetch_related("lines")
    queryset = get_accessible_schedules_queryset(user, queryset)
    queryset = queryset.filter(week_end_date__gte=date_from, week_start_date__lte=date_to)
    if site is not None:
        queryset = queryset.filter(site=site)
    return queryset.order_by("week_start_date", "site__code")


def build_special_days_report_rows(schedules: list[WeeklySchedule], date_from: date, date_to: date) -> list[list[object]]:
    rows: list[list[object]] = []
    for schedule in schedules:
        for line in schedule.lines.all():
            for day_info in build_line_day_breakdown(line):
                if not (date_from <= day_info["date"] <= date_to):
                    continue
                if Decimal(str(day_info["worked_hours"])) <= Decimal("0.00"):
                    continue
                label = str(day_info["special_label"] or "")
                if not label:
                    continue
                rows.append(
                    [
                        day_info["date"],
                        schedule.site.name,
                        line.employee_identifier,
                        line.employee_name,
                        label,
                    ]
                )
    return rows


def build_night_bonus_report_rows(schedules: list[WeeklySchedule], date_from: date, date_to: date) -> list[list[object]]:
    rows: list[list[object]] = []
    for schedule in schedules:
        for line in schedule.lines.all():
            for day_info in build_line_day_breakdown(line):
                if not (date_from <= day_info["date"] <= date_to):
                    continue
                night_hours = Decimal(str(day_info["night_hours"]))
                if night_hours <= Decimal("0.00"):
                    continue
                rows.append(
                    [
                        day_info["date"],
                        schedule.site.name,
                        line.employee_identifier,
                        line.employee_name,
                        float(night_hours),
                    ]
                )
    return rows


def build_overtime_balance_report_rows(user, date_from: date, date_to: date, site=None) -> list[list[object]]:
    movements = (
        ScheduleBalanceMovement.objects.select_related("site", "schedule")
        .filter(
            schedule__in=get_accessible_schedules_queryset(
                user,
                WeeklySchedule.objects.filter(week_end_date__gte=date_from, week_start_date__lte=date_to),
            ),
            movement_date__gte=date_from,
            movement_date__lte=date_to,
            equivalent_hours__gt=0,
        )
        .exclude(
            movement_type__in=[
                ScheduleBalanceMovement.MovementType.PAY_DAY,
                ScheduleBalanceMovement.MovementType.PAY_HOURS,
                ScheduleBalanceMovement.MovementType.PAY_MONEY_DAY,
                ScheduleBalanceMovement.MovementType.PAY_MONEY_HOURS,
                ScheduleBalanceMovement.MovementType.PAY_MONEY,
            ]
        )
        .order_by("movement_date", "site__code", "employee_name")
    )
    if site is not None:
        movements = movements.filter(site=site)

    return [
        [
            movement.movement_date,
            movement.site.name,
            movement.employee_identifier,
            movement.employee_name,
            float(movement.equivalent_hours),
        ]
        for movement in movements
    ]


def get_weekly_balance_lines(user, week_start_date: date, site=None):
    queryset = WeeklySchedule.objects.select_related("site").prefetch_related("lines")
    queryset = get_accessible_schedules_queryset(user, queryset).filter(week_start_date=week_start_date)
    if site is not None:
        queryset = queryset.filter(site=site)
    schedules = list(queryset.order_by("site__code"))
    lines: list[ScheduleLine] = []
    for schedule in schedules:
        for line in schedule.lines.all():
            line.schedule_scope = schedule
            lines.append(line)
    return lines


def build_weekly_balance_report_rows(lines: list[ScheduleLine]) -> list[list[object]]:
    return [
        [
            line.schedule_scope.site.name,
            line.job_role_name,
            line.employee_identifier,
            line.employee_name,
            excel_number(line.accrued_day_balance),
            excel_number(line.accrued_hour_balance),
            excel_number(line.night_bonus_hours),
        ]
        for line in lines
    ]


def build_inventory_report_entries(lines: list[ScheduleLine]) -> list[InventoryReportEntry]:
    entries: list[InventoryReportEntry] = []
    for line in lines:
        schedule = getattr(line, "schedule_scope", None) or getattr(line, "schedule", None)
        if schedule is None or not schedule.week_start_date:
            continue

        for index in range(7):
            if not bool(getattr(line, f"day_{index}_inventory", False)):
                continue

            entries.append(
                InventoryReportEntry(
                    site_code=(schedule.site.code or "").strip(),
                    site_name=(schedule.site.name or "").strip(),
                    inventory_date=schedule.week_start_date + timedelta(days=index),
                    employee_identifier=(line.employee_identifier or "").strip(),
                    employee_name=(line.employee_name or "").strip(),
                    job_role_name=(line.job_role_name or "SIN CARGO").strip() or "SIN CARGO",
                )
            )

    return sorted(
        entries,
        key=lambda entry: (
            entry.inventory_date,
            entry.site_code,
            entry.job_role_name.casefold(),
            entry.employee_name.casefold(),
            entry.employee_identifier,
        ),
    )


def build_inventory_report_rows(lines: list[ScheduleLine]) -> list[list[object]]:
    rows: list[list[object]] = []
    for entry in build_inventory_report_entries(lines):
        rows.append(
            [
                entry.inventory_date,
                SPANISH_MONTH_NAMES[entry.inventory_date.month],
                entry.site_name,
                entry.employee_identifier,
                entry.employee_name,
                entry.job_role_name,
                "",
            ]
        )
    return rows


def get_latest_visible_lines_by_employee(user) -> list[ScheduleLine]:
    queryset = (
        get_accessible_schedules_queryset(
            user,
            WeeklySchedule.objects.select_related("site").prefetch_related("lines"),
        )
        .order_by("-week_start_date", "site__code")
    )
    latest_by_employee: dict[str, tuple[tuple[int, tuple[object, ...]], ScheduleLine]] = {}
    for schedule in queryset:
        for line in schedule.lines.all():
            employee_identifier = (line.employee_identifier or "").strip()
            if not employee_identifier:
                continue
            candidate_key = (
                1 if schedule.status == WeeklySchedule.Status.PUBLISHED or get_schedule_line_activity_indices(line) else 0,
                get_schedule_line_progression_key(line),
            )
            current_entry = latest_by_employee.get(employee_identifier)
            if current_entry is None or candidate_key > current_entry[0]:
                latest_by_employee[employee_identifier] = (candidate_key, line)
    return [entry[1] for entry in latest_by_employee.values()]


def build_balance_role_breakdown(lines: list[ScheduleLine]) -> dict[str, object]:
    days_by_role: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0.00"))
    hours_by_role: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0.00"))
    total_days = Decimal("0.00")
    total_hours = Decimal("0.00")

    for line in lines:
        role_name = (line.job_role_name or "Sin cargo").strip() or "Sin cargo"
        day_value = max(Decimal(str(line.accrued_day_balance or "0")), Decimal("0.00"))
        hour_value = max(Decimal(str(line.accrued_hour_balance or "0")), Decimal("0.00"))
        days_by_role[role_name] += day_value
        hours_by_role[role_name] += hour_value
        total_days += day_value
        total_hours += hour_value

    return {
        "total_days": total_days,
        "total_hours": total_hours,
        "days_by_role": sorted(days_by_role.items(), key=lambda item: item[0].casefold()),
        "hours_by_role": sorted(hours_by_role.items(), key=lambda item: item[0].casefold()),
    }
